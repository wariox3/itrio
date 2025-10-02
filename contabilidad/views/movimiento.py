from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from contabilidad.models.movimiento import ConMovimiento
from contabilidad.models.comprobante import ConComprobante
from contabilidad.models.cuenta import ConCuenta
from contabilidad.models.grupo import ConGrupo
from contabilidad.models.periodo import ConPeriodo
from contabilidad.models.cuenta_clase import ConCuentaClase
from contabilidad.models.cuenta_grupo import ConCuentaGrupo
from contabilidad.models.cuenta_cuenta import ConCuentaCuenta
from general.models.contacto import GenContacto
from contabilidad.serializers.movimiento import ConMovimientoSerializador, ConMovimientoExcelSerializador, ConMovimientoListaSerializador
from datetime import datetime
from io import BytesIO
from django.db.models import F,Sum
from io import BytesIO
from django.http import HttpResponse
from django.db import transaction
from utilidades.workbook_estilos_deprecated import WorkbookEstilos
from utilidades.excel_funciones import ExcelFunciones
from openpyxl import Workbook
from general.formatos.balance_prueba import FormatoBalancePrueba
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import OrderingFilter
from contabilidad.filters.movimiento import MovimientoFilter
from utilidades.excel_exportar import ExcelExportar
import base64
import openpyxl

class MovimientoViewSet(viewsets.ModelViewSet):
    queryset = ConMovimiento.objects.all()
    serializer_class = ConMovimientoSerializador
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_class = MovimientoFilter 
    serializadores = {
        'lista': ConMovimientoListaSerializador,
        'informe_movimiento': ConMovimientoExcelSerializador,
    }

    def get_serializer_class(self):
        serializador_parametro = self.request.query_params.get('serializador', None)
        if not serializador_parametro or serializador_parametro not in self.serializadores:
            return ConMovimientoSerializador
        return self.serializadores[serializador_parametro]

    def get_queryset(self):
        queryset = super().get_queryset()
        serializer_class = self.get_serializer_class()        
        select_related = getattr(serializer_class.Meta, 'select_related_fields', [])
        if select_related:
            queryset = queryset.select_related(*select_related)        
        campos = serializer_class.Meta.fields        
        if campos and campos != '__all__':
            queryset = queryset.only(*campos) 
        return queryset 
    
    def list(self, request, *args, **kwargs):
        if request.query_params.get('excel'):
            queryset = self.filter_queryset(self.get_queryset())
            serializer = self.get_serializer(queryset, many=True)
            titulo = 'Informe movimiento detalles'
            nombre_hoja = "movimiento_detalles"
            nombre_archivo = "movimiento_detalles.xlsx"
            if request.query_params.get('excel_masivo'):
                exporter = ExcelExportar(serializer.data, nombre_hoja, nombre_archivo)
                return exporter.exportar() 
            elif request.query_params.get('excel_informe'): 
                serializador_parametro = self.request.query_params.get('serializador', None)                
                if serializador_parametro == 'informe_movimiento':
                    titulo = 'Movimientos' 
                    nombre_archivo = "movimientos.xlsx"  
                    nombre_hoja = 'movimientos'    
                exporter = ExcelExportar(serializer.data, nombre_hoja, nombre_archivo, titulo)
                return exporter.exportar_informe()                    
            else:
                exporter = ExcelExportar(serializer.data, nombre_hoja, nombre_archivo)
                return exporter.exportar_estilo()            
        return super().list(request, *args, **kwargs)

    @action(detail=False, methods=["post"], url_path=r'importar',)
    def importar(self, request):
        raw = request.data        
        archivo_base64 = raw.get('archivo_base64')
        if archivo_base64:
            try:
                archivo_data = base64.b64decode(archivo_base64)
                archivo = BytesIO(archivo_data)
                wb = openpyxl.load_workbook(archivo)
                sheet = wb.active    
            except Exception as e:     
                return Response({f'mensaje':'Error procesando el archivo, valide que es un archivo de excel .xlsx', 'codigo':15}, status=status.HTTP_400_BAD_REQUEST)  
            cuentas_map = {c.codigo: c.id for c in ConCuenta.objects.all()}
            grupos_map = {g.codigo: g.id for g in ConGrupo.objects.all()}
            contactos_map = {ct.numero_identificacion: ct.id for ct in GenContacto.objects.all()}
            data_modelo = []
            errores = False
            errores_datos = []
            registros_importados = 0
            for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                data = {
                    'numero': row[0],
                    'fecha': None,
                    'debito': row[2] if row[2] is not None else 0,
                    'credito': row[3] if row[3] is not None else 0,
                    'base': row[4] if row[4] is not None else 0, 
                    'naturaleza': None,
                    'comprobante': row[5],
                    'cuenta': str(row[6]),
                    'grupo': row[7],
                    'contacto': row[8],
                    'detalle': row[9],
                    'periodo': row[10],
                    'cierre': row[11],
                    'documento': None
                }  

                if row[1]:       
                    fecha_str = str(row[1])
                    fecha_valida = datetime.strptime(fecha_str, "%Y%m%d").date()
                    data['fecha'] = fecha_valida                 
                
                data['cuenta'] = cuentas_map.get(data['cuenta'])
                data['grupo'] = grupos_map.get(data['grupo'])
                data['contacto'] = contactos_map.get(data['contacto'])

                naturaleza = 'D'
                if data['credito'] > 0:
                    naturaleza = 'C'
                data['naturaleza'] = naturaleza

                serializer = ConMovimientoSerializador(data=data)
                if serializer.is_valid():
                    data_modelo.append(serializer.validated_data)
                    registros_importados += 1
                else:
                    errores = True
                    error_dato = {
                        'fila': i,
                        'errores': serializer.errors
                    }                                    
                    errores_datos.append(error_dato)

            if not errores:
                with transaction.atomic():
                    batch_size = 1000  # Define el tamaño del lote
                    for i in range(0, len(data_modelo), batch_size):
                        batch = data_modelo[i:i + batch_size]
                        ConMovimiento.objects.bulk_create([ConMovimiento(**detalle) for detalle in batch])
                #ConMovimiento.objects.bulk_create([ConMovimiento(**detalle) for detalle in data_modelo])
                return Response({'registros_importados': registros_importados}, status=status.HTTP_200_OK)
            else:                    
                return Response({'mensaje':'Errores de validacion', 'codigo':1, 'errores_validador': errores_datos}, status=status.HTTP_400_BAD_REQUEST)        
        else:
            return Response({'mensaje':'Faltan parametros', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)    
        
    def obtener_saldo_cuenta(self, fecha_desde, fecha_hasta, cierre = False, cuenta_con_movimiento = False, cuenta_desde = None, cuenta_hasta = None):
        parametro_cierre = ' AND m.cierre = false '
        if cierre == True:
            parametro_cierre = ''
        filtro_cuenta_desde = ''
        if cuenta_desde:
            filtro_cuenta_desde = f"AND c.codigo >= '{cuenta_desde}' "  
        filtro_cuenta_hasta = ''
        if cuenta_hasta:
            filtro_cuenta_hasta = f"AND c.codigo <= '{cuenta_hasta}' "               
        query = f'''
            SELECT
                c.id,
                c.codigo,
                c.nombre,
                c.cuenta_clase_id,
                c.cuenta_grupo_id,
                c.cuenta_cuenta_id,
                c.nivel,
                COALESCE(SUM(CASE WHEN m.fecha < '{fecha_desde}' THEN m.debito ELSE 0 END), 0) AS debito_anterior,
                COALESCE(SUM(CASE WHEN m.fecha < '{fecha_desde}' THEN m.credito ELSE 0 END), 0) AS credito_anterior,
                COALESCE(SUM(CASE WHEN m.fecha BETWEEN '{fecha_desde}' AND '{fecha_hasta}' {parametro_cierre} THEN m.debito ELSE 0 END), 0) AS debito,
                COALESCE(SUM(CASE WHEN m.fecha BETWEEN '{fecha_desde}' AND '{fecha_hasta}' {parametro_cierre} THEN m.credito ELSE 0 END), 0) AS credito
            FROM
                con_cuenta c
            LEFT JOIN
                con_movimiento m ON m.cuenta_id = c.id
            WHERE
                1=1 
                {filtro_cuenta_desde}
                {filtro_cuenta_hasta}                                
            GROUP BY
                c.id, c.codigo, c.nombre, c.cuenta_clase_id, c.cuenta_grupo_id, c.cuenta_cuenta_id, c.nivel
            ORDER BY
                c.cuenta_clase_id,
                c.cuenta_grupo_id,
                c.cuenta_cuenta_id;
        '''
        resultados = ConCuenta.objects.raw(query)
        
        # Procesar resultados
        resultados_json = []
        clases_dict = {}
        grupos_dict = {}
        cuentas_dict = {}
        for cuenta in resultados:
            saldo_anterior = cuenta.debito_anterior - cuenta.credito_anterior
            saldo_actual = saldo_anterior + (cuenta.debito - cuenta.credito)
            if cuenta_con_movimiento and saldo_anterior == 0 and cuenta.debito == 0 and cuenta.credito == 0:
                continue 
            resultados_json.append({
                'tipo': 'AUXILIAR',
                'id': cuenta.id,
                'codigo': cuenta.codigo,
                'nombre': cuenta.nombre,
                'contacto_id': None,
                'contacto_nombre_corto': None,
                'contacto_numero_identificacion': None,
                'cuenta_clase_id': cuenta.cuenta_clase_id,
                'cuenta_grupo_id': cuenta.cuenta_grupo_id,
                'cuenta_cuenta_id': cuenta.cuenta_cuenta_id,
                'nivel': cuenta.nivel,
                'saldo_anterior': saldo_anterior,
                'debito': cuenta.debito,
                'credito': cuenta.credito,
                'saldo_actual': saldo_actual,
                'comprobante_id': None,
                'comprobante_nombre': None,
                'numero': None,
                'fecha': None,  
            })
            
            # Agrupación en clases, grupos y cuentas
            clase_id = cuenta.cuenta_clase_id
            if clase_id not in clases_dict:
                clases_dict[clase_id] = {
                    'tipo': 'CLASE',
                    'id': clase_id,
                    'codigo': str(clase_id),
                    'nombre': "",
                    'contacto_id': None,
                    'contacto_nombre_corto': None,                    
                    'contacto_numero_identificacion': None,    
                    'cuenta_clase_id': clase_id,
                    'nivel': 1,
                    'saldo_anterior': 0,
                    'debito': 0,
                    'credito': 0,
                    'saldo_actual': 0,
                    'comprobante_id': None,
                    'comprobante_nombre': None,
                    'numero': None,
                    'fecha': None,  
                }
            clases_dict[clase_id]['saldo_anterior'] += saldo_anterior
            clases_dict[clase_id]['debito'] += cuenta.debito
            clases_dict[clase_id]['credito'] += cuenta.credito
            clases_dict[clase_id]['saldo_actual'] += saldo_actual
            
            grupo_id = cuenta.cuenta_grupo_id
            if grupo_id not in grupos_dict:
                grupos_dict[grupo_id] = {
                    'tipo': 'GRUPO',
                    'id': grupo_id,
                    'codigo': str(grupo_id),
                    'nombre': "",
                    'contacto_id': None,
                    'contacto_nombre_corto': None,                    
                    'contacto_numero_identificacion': None,                
                    'cuenta_grupo_id': grupo_id,
                    'saldo_anterior': 0,
                    'debito': 0,
                    'credito': 0,
                    'saldo_actual': 0,
                    'comprobante_id': None,
                    'comprobante_nombre': None,
                    'numero': None,
                    'fecha': None,  
                }
            grupos_dict[grupo_id]['saldo_anterior'] += saldo_anterior
            grupos_dict[grupo_id]['debito'] += cuenta.debito
            grupos_dict[grupo_id]['credito'] += cuenta.credito
            grupos_dict[grupo_id]['saldo_actual'] += saldo_actual

            cuenta_id = cuenta.cuenta_cuenta_id
            if cuenta_id not in cuentas_dict:
                cuentas_dict[cuenta_id] = {
                    'tipo': 'CUENTA',
                    'id': cuenta_id,
                    'codigo': str(cuenta_id),
                    'nombre': "",
                    'contacto_id': None,
                    'contacto_nombre_corto': None,                    
                    'contacto_numero_identificacion': None,    
                    'cuenta_cuenta_id': cuenta_id,
                    'saldo_anterior': 0,
                    'debito': 0,
                    'credito': 0,
                    'saldo_actual': 0,
                    'comprobante_id': None,
                    'comprobante_nombre': None,
                    'numero': None,
                    'fecha': None,  
                }
            cuentas_dict[cuenta_id]['saldo_anterior'] += saldo_anterior
            cuentas_dict[cuenta_id]['debito'] += cuenta.debito
            cuentas_dict[cuenta_id]['credito'] += cuenta.credito
            cuentas_dict[cuenta_id]['saldo_actual'] += saldo_actual
        
            cuentas_clases_map = {c.id: c.nombre for c in ConCuentaClase.objects.all()}
            for clase_id, clase_data in clases_dict.items():
                if clase_id in cuentas_clases_map and clase_id is not None:
                    clase_data['nombre'] = cuentas_clases_map[clase_id]

            cuentas_grupos_map = {g.id: g.nombre for g in ConCuentaGrupo.objects.all()}
            for grupo_id, grupo_data in grupos_dict.items():
                if grupo_id in cuentas_grupos_map and grupo_id is not None:
                    grupo_data['nombre'] = cuentas_grupos_map[grupo_id]

            cuentas_cuentas_map = {cc.id: cc.nombre for cc in ConCuentaCuenta.objects.all()}
            for cuenta_cuenta_id, cuenta_cuenta_data in cuentas_dict.items():
                if cuenta_cuenta_id in cuentas_cuentas_map and cuenta_cuenta_id is not None:
                    cuenta_cuenta_data['nombre'] = cuentas_cuentas_map[cuenta_cuenta_id]
        # Agregar clases, grupos y cuentas al resultado final
        resultados_json.extend(clases_dict.values())
        resultados_json.extend(grupos_dict.values())
        resultados_json.extend(cuentas_dict.values())
        resultados_json.sort(key=lambda x: str(x['codigo']))
        return resultados_json

    def obtener_balance_prueba_tercero(self, fecha_desde, fecha_hasta, cierre = False, numero_identificacion = '', nombre_corto = ''):
        parametro_cierre = ' AND ma.cierre = false '
        if cierre == True:
            parametro_cierre = ''
        filtro_identificacion = ''
        if numero_identificacion:
            filtro_identificacion = f" AND co.numero_identificacion = '{numero_identificacion}'"                
        filtro_nombre_corto = ''
        if nombre_corto:
                filtro_nombre_corto = f" AND co.nombre_corto LIKE '%%{nombre_corto}%%'"
        query = f'''
            SELECT
                MIN(m.id) AS id,
                m.cuenta_id,    
                c.codigo,
                c.nombre,
                m.contacto_id,
                co.nombre_corto,
                co.numero_identificacion,
                c.cuenta_clase_id,
                c.cuenta_grupo_id,
                c.cuenta_cuenta_id,
                c.nivel,   
                COALESCE((SELECT SUM(debito) FROM con_movimiento ma WHERE ma.cuenta_id = m.cuenta_id AND ma.contacto_id = m.contacto_id and ma.fecha < '{fecha_desde}'), 0) AS debito_anterior,
                COALESCE((SELECT SUM(credito) FROM con_movimiento ma WHERE ma.cuenta_id = m.cuenta_id AND ma.contacto_id = m.contacto_id and ma.fecha < '{fecha_desde}'), 0) AS credito_anterior,
                COALESCE((SELECT SUM(debito) FROM con_movimiento ma WHERE ma.cuenta_id = m.cuenta_id AND ma.contacto_id = m.contacto_id {parametro_cierre} and ma.fecha BETWEEN '{fecha_desde}' AND '{fecha_hasta}'), 0) AS debito,
                COALESCE((SELECT SUM(credito) FROM con_movimiento ma WHERE ma.cuenta_id = m.cuenta_id AND ma.contacto_id = m.contacto_id {parametro_cierre} and ma.fecha BETWEEN '{fecha_desde}' AND '{fecha_hasta}'), 0) AS credito
            FROM
                con_movimiento m
            LEFT JOIN
                con_cuenta c ON m.cuenta_id = c.id
            LEFT JOIN
                gen_contacto co ON m.contacto_id = co.id  
            WHERE 
                m.fecha <= '{fecha_hasta}' and m.contacto_id is not null {filtro_identificacion} {filtro_nombre_corto}
            GROUP BY
                m.cuenta_id, m.contacto_id, co.nombre_corto, co.numero_identificacion ,c.codigo, c.nombre, c.cuenta_clase_id, c.cuenta_grupo_id, c.cuenta_cuenta_id, c.nivel
            ORDER BY
                c.cuenta_clase_id,
                c.cuenta_grupo_id,
                c.cuenta_cuenta_id;
        '''
        resultados = ConMovimiento.objects.raw(query)
        
        # Procesar resultados
        resultados_json = []        
        for cuenta in resultados:
            saldo_anterior = cuenta.debito_anterior - cuenta.credito_anterior
            saldo_actual = saldo_anterior + (cuenta.debito - cuenta.credito)
            resultados_json.append({
                'tipo': 'TERCERO',
                'cuenta_id': cuenta.cuenta_id,
                'codigo': cuenta.codigo,
                'nombre': cuenta.nombre,
                'contacto_id': cuenta.contacto_id,
                'contacto_nombre_corto': cuenta.nombre_corto,
                'contacto_numero_identificacion': cuenta.numero_identificacion,
                'cuenta_clase_id': cuenta.cuenta_clase_id,
                'cuenta_grupo_id': cuenta.cuenta_grupo_id,
                'cuenta_cuenta_id': cuenta.cuenta_cuenta_id,
                'nivel': cuenta.nivel,
                'saldo_anterior': saldo_anterior,
                'debito': cuenta.debito,
                'credito': cuenta.credito,
                'saldo_actual': saldo_actual,
                'comprobante_id': None,
                'comprobante_nombre': None,
                'numero': None,
                'fecha': None,    
            })                                        
        #resultados_json.sort(key=lambda x: str(x['codigo']))
        return resultados_json

    def obtener_movimiento(self, fecha_desde, fecha_hasta, cierre = False, comprobante = None, cuenta = None):
        parametro_cierre = ' AND m.cierre = false '
        if cierre == True:
            parametro_cierre = ''
        filtro_comprobante = ''
        if comprobante:
            filtro_comprobante = f" AND m.comprobante_id = {comprobante}"   
        filtro_cuenta = ''     
        if cuenta:
            filtro_cuenta = f"AND c.codigo = '{cuenta}'"           
        query = f'''            
            SELECT
                m.id,
                m.debito,
                m.credito,
                m.comprobante_id,
                m.contacto_id,
                m.numero,
                m.fecha,
                c.codigo,
                c.nombre,
                c.cuenta_clase_id,
                c.cuenta_grupo_id,
                c.cuenta_cuenta_id,
                c.nivel,
                co.nombre AS comprobante_nombre,
                con.numero_identificacion AS contacto_numero_identificacion,
                con.nombre_corto AS contacto_nombre_corto
            FROM
                con_movimiento m
            LEFT JOIN
                con_cuenta c ON m.cuenta_id = c.id
            LEFT JOIN
                con_comprobante co ON m.comprobante_id = co.id
            LEFT JOIN
                gen_contacto con ON m.contacto_id = con.id
            WHERE
        		m.fecha BETWEEN '{fecha_desde}' AND '{fecha_hasta}' {parametro_cierre}  {filtro_comprobante} {filtro_cuenta}
            ORDER BY
                m.fecha ASC;
        '''
        resultados = ConMovimiento.objects.raw(query)   
        resultados_json = []
        for movimiento in resultados:            
            resultados_json.append({
                'tipo': 'MOVIMIENTO',
                'id': movimiento.id,
                'codigo': movimiento.codigo,
                'nombre': movimiento.nombre,
                'contacto_id': movimiento.contacto_id,
                'contacto_nombre_corto': movimiento.contacto_nombre_corto,
                'contacto_numero_identificacion': movimiento.contacto_numero_identificacion,
                'cuenta_clase_id': movimiento.cuenta_clase_id,
                'cuenta_grupo_id': movimiento.cuenta_grupo_id,
                'cuenta_cuenta_id': movimiento.cuenta_cuenta_id,
                'nivel': movimiento.nivel,
                'saldo_anterior': 0,
                'debito': movimiento.debito,
                'credito': movimiento.credito,
                'saldo_actual': 0,
                'comprobante_id': movimiento.comprobante_id,
                'comprobante_nombre': movimiento.comprobante_nombre,
                'numero': movimiento.numero,
                'fecha': movimiento.fecha,  
            })         
        return resultados_json

    def obtener_movimiento_detallado(self, fecha_desde, fecha_hasta, cierre = False):
        parametro_cierre = ' AND m.cierre = false '
        if cierre == True:
            parametro_cierre = ''
            
        query = f'''            
            SELECT
                m.id,
                m.numero,
                m.fecha,                
                m.debito,
                m.credito,
                m.detalle, 
                m.cuenta_id,
                m.contacto_id,                               
                m.comprobante_id,
                c.codigo as cuenta_codigo,
                c.nombre as cuenta_nombre,
                c.cuenta_clase_id,
                c.cuenta_grupo_id,
                c.cuenta_cuenta_id,
                c.nivel as cuenta_nivel,
                cp.nombre as comprobante_nombre,
                co.nombre_corto as contacto_nombre_corto,
                co.numero_identificacion as contacto_numero_identificacion
            FROM
                con_movimiento m
            LEFT JOIN
                con_cuenta c ON m.cuenta_id = c.id
            LEFT JOIN
                con_comprobante cp ON m.comprobante_id = cp.id                
            LEFT JOIN
                gen_contacto co ON m.contacto_id = co.id
            WHERE
        		m.fecha BETWEEN '{fecha_desde}' AND '{fecha_hasta}' {parametro_cierre}   
            ORDER BY
                m.fecha ASC;
        '''
        resultados = ConMovimiento.objects.raw(query)   
        resultados_json = []
        for movimiento in resultados:            
            resultados_json.append({
                'id': movimiento.id,
                'numero': movimiento.numero,
                'fecha': movimiento.fecha,
                'detalle': movimiento.detalle,
                'debito': movimiento.debito,
                'credito': movimiento.credito,
                'base': movimiento.base,                
                'cuenta_codigo': movimiento.cuenta_codigo,
                'cuenta_nombre': movimiento.cuenta_nombre,
                'cuenta_clase_id': movimiento.cuenta_clase_id,
                'cuenta_grupo_id': movimiento.cuenta_grupo_id,
                'cuenta_cuenta_id': movimiento.cuenta_cuenta_id,
                'cuenta_nivel': movimiento.cuenta_nivel,
                'contacto_id': movimiento.contacto_id,
                'contacto_nombre_corto': movimiento.contacto_nombre_corto,  
                'contacto_numero_identificacion': movimiento.contacto_numero_identificacion,  
                'comprobante_nombre': movimiento.comprobante_nombre              
            })         
        return resultados_json

    def obtener_movimiento_certificado_retencion(self, fecha_desde, fecha_hasta, cuenta_desde = None, cuenta_hasta = None, contacto_id = None):
        where_contacto = ''
        if contacto_id:
            where_contacto = f" AND m.contacto_id >= {contacto_id}"

        where_cuenta_desde = ''
        if cuenta_desde:
            where_cuenta_desde = f" AND c.codigo >= '{cuenta_desde}'"

        where_cuenta_hasta = ''
        if cuenta_hasta:
            where_cuenta_hasta = f" AND c.codigo <= '{cuenta_hasta}'"            
            
        query = f'''
            select
            	MIN(m.id) AS id,
                m.cuenta_id,
            	m.contacto_id,            	
                c.codigo as cuenta_codigo,
                c.nombre as cuenta_nombre,                              
                co.numero_identificacion  as contacto_numero_identificacion,
                co.nombre_corto as contacto_nombre_corto,
                SUM(m.debito) as debito,
                SUM(m.credito) as credito,
                SUM(m.base) as base,
                SUM(CASE WHEN m.debito > 0 THEN m.base ELSE 0 END) as base_debito,
                SUM(CASE WHEN m.credito > 0 THEN m.base ELSE 0 END) as base_credito                
            FROM
                con_movimiento m
            LEFT JOIN
                con_cuenta c ON m.cuenta_id = c.id               
            LEFT JOIN
                gen_contacto co ON m.contacto_id = co.id
            WHERE
        		m.fecha BETWEEN '{fecha_desde}' AND '{fecha_hasta}' {where_cuenta_desde} {where_cuenta_hasta} {where_contacto}
        	group by 
        		m.cuenta_id,
        		m.contacto_id,
        		c.codigo,
        		c.nombre,
        		co.numero_identificacion,
        		co.nombre_corto            
        '''
        resultados = ConMovimiento.objects.raw(query)   
        resultados_json = []
        for movimiento in resultados: 
            retenido = movimiento.debito - movimiento.credito
            base_retenido = movimiento.base_debito - movimiento.base_credito
            resultados_json.append({
                'debito': movimiento.debito,
                'credito': movimiento.credito,
                'base': movimiento.base,     
                'retenido': retenido,  
                'base_retenido': base_retenido,         
                'cuenta_codigo': movimiento.cuenta_codigo,
                'cuenta_nombre': movimiento.cuenta_nombre,
                'contacto_id': movimiento.contacto_id,
                'contacto_numero_identificacion': movimiento.contacto_numero_identificacion,  
                'contacto_nombre_corto': movimiento.contacto_nombre_corto,  
                
            })         
        return resultados_json

    def obtener_movimiento_base(self, fecha_desde, fecha_hasta, cuenta_desde = None, cuenta_hasta = None, contacto_id = None):
        where_contacto = ''
        if contacto_id:
            where_contacto = f" AND m.contacto_id >= {contacto_id}"

        where_cuenta_desde = ''
        if cuenta_desde:
            where_cuenta_desde = f" AND c.codigo >= '{cuenta_desde}'"

        where_cuenta_hasta = ''
        if cuenta_hasta:
            where_cuenta_hasta = f" AND c.codigo <= '{cuenta_hasta}'" 
            
        query = f'''            
            SELECT
                m.id,
                m.numero,
                m.fecha,                
                m.debito,
                m.credito,
                m.detalle, 
                m.cuenta_id,
                m.contacto_id,                               
                m.comprobante_id,
                c.codigo as cuenta_codigo,
                c.nombre as cuenta_nombre,
                c.cuenta_clase_id,
                c.cuenta_grupo_id,
                c.cuenta_cuenta_id,
                c.nivel as cuenta_nivel,
                cp.nombre as comprobante_nombre,
                co.nombre_corto as contacto_nombre_corto,
                co.numero_identificacion as contacto_numero_identificacion
            FROM
                con_movimiento m
            LEFT JOIN
                con_cuenta c ON m.cuenta_id = c.id
            LEFT JOIN
                con_comprobante cp ON m.comprobante_id = cp.id                
            LEFT JOIN
                gen_contacto co ON m.contacto_id = co.id
            WHERE
        		m.fecha BETWEEN '{fecha_desde}' AND '{fecha_hasta}' {where_cuenta_desde} {where_cuenta_hasta} {where_contacto}   
                AND c.exige_base = true 
            ORDER BY
                m.fecha ASC;
        '''
        resultados = ConMovimiento.objects.raw(query)   
        resultados_json = []
        for movimiento in resultados:            
            resultados_json.append({
                'id': movimiento.id,
                'comprobante_nombre': movimiento.comprobante_nombre, 
                'numero': movimiento.numero,
                'fecha': movimiento.fecha,    
                'cuenta_codigo': movimiento.cuenta_codigo,
                'cuenta_nombre': movimiento.cuenta_nombre,
                'contacto_id': movimiento.contacto_id,
                'contacto_nombre_corto': movimiento.contacto_nombre_corto,  
                'contacto_numero_identificacion': movimiento.contacto_numero_identificacion,                  
                'debito': movimiento.debito,
                'credito': movimiento.credito,
                'base': movimiento.base,  
                'detalle': movimiento.detalle                                                                           
            })         
        return resultados_json

    @action(detail=False, methods=["post"], url_path=r'informe-balance-prueba',)
    def informe_balance_prueba(self, request):    
        raw = request.data
        parametros = raw.get("parametros", [])
        excel = raw.get('excel', False)
        pdf = raw.get('pdf', False)
        fecha_desde = parametros['fecha_desde']
        fecha_hasta = parametros['fecha_hasta']
        cuenta_codigo_desde = parametros.get('cuenta_codigo_desde', None)
        cuenta_codigo_hasta = parametros.get('cuenta_codigo_hasta', None)       
        cierre = parametros['incluir_cierre']
        cuenta_con_movimiento = parametros['cuenta_con_movimiento'] 
        resultados_json = self.obtener_saldo_cuenta(fecha_desde, fecha_hasta, cierre, cuenta_con_movimiento, cuenta_codigo_desde, cuenta_codigo_hasta)
        if pdf:
            formato = FormatoBalancePrueba()
            pdf = formato.generar_pdf(fecha_desde, fecha_hasta, resultados_json)
            nombre_archivo = f"balance_prueba{fecha_desde}.pdf"
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
            return response
        if excel:            
            excel_funciones = ExcelFunciones()
            wb = Workbook()
            ws = wb.active
            ws.title = "balance_prueba"                                                
            excel_funciones.agregar_titulo(ws, "Balance de prueba", "A", "G")                                   
            ws['A4'] = f"Fecha desde: {fecha_desde}"
            ws['A4'].font = excel_funciones.fuente_general           
            ws['A5'] = f"Fecha hasta: {fecha_hasta}"
            ws['A5'].font = excel_funciones.fuente_general 
            ws.append([])

            headers = ["Tipo", "Cuenta", "Nombre de cuenta", "Saldo anterior ($)", "Debitos ($)", "Creditos ($)", "Saldo actual ($)"]
            ws.append(headers)
            for registro in resultados_json:
                ws.append([
                    registro['tipo'],
                    registro['codigo'],
                    registro['nombre'],                    
                    registro['saldo_anterior'],
                    registro['debito'],
                    registro['credito'],
                    registro['saldo_actual']
                ])    
            excel_funciones.aplicar_estilos(ws, 7, [4,5,6,7])                                
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'
            response['Content-Disposition'] = f'attachment; filename=balance_prueba.xlsx'
            wb.save(response)
            return response
        else:
            return Response({'registros': resultados_json}, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=["post"], url_path=r'informe-balance-prueba-tercero',)
    def informe_balance_prueba_tercero(self, request):    
        raw = request.data
        parametros = raw.get("parametros", [])
        excel = raw.get('excel', False)
        pdf = raw.get('pdf', False)
        fecha_desde = parametros['fecha_desde']
        fecha_hasta = parametros['fecha_hasta']
        cuenta_codigo_desde = parametros.get('cuenta_codigo_desde', None)
        cuenta_codigo_hasta = parametros.get('cuenta_codigo_hasta', None)
        numero_identificacion = parametros['numero_identificacion']
        nombre_corto = parametros['nombre_corto']
        cierre = parametros['incluir_cierre']
        cuenta_con_movimiento = parametros['cuenta_con_movimiento'] 
        resultados_cuenta = self.obtener_saldo_cuenta(fecha_desde, fecha_hasta, cierre, cuenta_con_movimiento, cuenta_codigo_desde, cuenta_codigo_hasta)
        resultados_tercero = self.obtener_balance_prueba_tercero(fecha_desde, fecha_hasta, cierre ,numero_identificacion, nombre_corto)
        resultados = resultados_cuenta + resultados_tercero
        resultados.sort(key=lambda x: str(x['codigo']))
        if excel:
            excel_funciones = ExcelFunciones()
            wb = Workbook()
            ws = wb.active
            ws.title = "balance_prueba_contacto"                                                
            excel_funciones.agregar_titulo(ws, "Balance de prueba por contacto", "A", "I")                                   
            ws['A4'] = f"Fecha desde: {fecha_desde}"
            ws['A4'].font = excel_funciones.fuente_general           
            ws['A5'] = f"Fecha hasta: {fecha_hasta}"
            ws['A5'].font = excel_funciones.fuente_general 
            ws.append([])

            headers = ["Tipo", "Cuenta", "Nombre Cuenta", "Identificación", "Contacto", "Saldo anterior ($)", "Debitos ($)", "Creditos ($)", "Saldo actual ($)"]
            ws.append(headers)
            for registro in resultados:
                ws.append([
                    registro['tipo'],
                    registro['codigo'],
                    registro['nombre'],
                    registro['contacto_numero_identificacion'],
                    registro['contacto_nombre_corto'],
                    registro['saldo_anterior'],
                    registro['debito'],
                    registro['credito'],
                    registro['saldo_actual']
                ])    
            
            excel_funciones.aplicar_estilos(ws, 7, [6,7,8,9])                                      
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'
            response['Content-Disposition'] = f'attachment; filename=balance_prueba_contacto.xlsx'
            wb.save(response)
            return response
        else:
            return Response({'registros': resultados}, status=status.HTTP_200_OK)

    @action(detail=False, methods=["post"], url_path=r'informe-auxiliar-cuenta',)
    def informe_auxiliar_cuenta(self, request):    
        raw = request.data
        parametros = raw.get("parametros", [])
        excel = raw.get('excel', False)
        pdf = raw.get('pdf', False)
        fecha_desde = parametros['fecha_desde']
        fecha_hasta = parametros['fecha_hasta']
        cuenta_codigo_desde = parametros.get('cuenta_codigo_desde', None)
        cuenta_codigo_hasta = parametros.get('cuenta_codigo_hasta', None)
        cierre = parametros['incluir_cierre']
        cuenta_con_movimiento = parametros['cuenta_con_movimiento'] 
        comprobante = parametros['comprobante']
        cuenta = parametros['cuenta']         
        resultados_cuenta = self.obtener_saldo_cuenta(fecha_desde, fecha_hasta, cierre, cuenta_con_movimiento, cuenta_codigo_desde, cuenta_codigo_hasta)
        movimientos = self.obtener_movimiento(fecha_desde, fecha_hasta, cierre, comprobante, cuenta)
        resultados_json = resultados_cuenta + movimientos
        resultados_json.sort(key=lambda x: str(x['codigo']))
        if excel:
            excel_funciones = ExcelFunciones()
            wb = Workbook()
            ws = wb.active
            ws.title = "auxiliar_cuenta"                                                
            excel_funciones.agregar_titulo(ws, "Auxiliar cuenta", "A", "G")                                   
            ws['A4'] = f"Fecha desde: {fecha_desde}"
            ws['A4'].font = excel_funciones.fuente_general           
            ws['A5'] = f"Fecha hasta: {fecha_hasta}"
            ws['A5'].font = excel_funciones.fuente_general 
            ws.append([])
            headers = ["Tipo", "Cuenta", "Nombre de cuenta", "Saldo anterior ($)", "Debitos ($)", "Creditos ($)", "Saldo actual ($)"]
            ws.append(headers)
            for registro in resultados_json:
                ws.append([
                    registro['tipo'],
                    registro['codigo'],
                    registro['nombre'],                    
                    registro['saldo_anterior'],
                    registro['debito'],
                    registro['credito'],
                    registro['saldo_actual']
                ])    
            
            excel_funciones.aplicar_estilos(ws, 7, [4,5,6,7])                            
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'
            response['Content-Disposition'] = f'attachment; filename=auxiliar_cuenta.xlsx'
            wb.save(response)
            return response
        else:
            return Response({'registros': resultados_json}, status=status.HTTP_200_OK)
        
    @action(detail=False, methods=["post"], url_path=r'informe-auxiliar-tercero',)
    def informe_auxiliar_tercero(self, request):                 
        raw = request.data
        parametros = raw.get("parametros", [])
        excel = raw.get('excel', False)
        pdf = raw.get('pdf', False)
        fecha_desde = parametros['fecha_desde']
        fecha_hasta = parametros['fecha_hasta']
        cuenta_codigo_desde = parametros.get('cuenta_codigo_desde', None)
        cuenta_codigo_hasta = parametros.get('cuenta_codigo_hasta', None)
        numero_identificacion = parametros['numero_identificacion']
        nombre_corto = parametros['nombre_corto']
        cierre = parametros['incluir_cierre']
        cuenta_con_movimiento = parametros['cuenta_con_movimiento'] 
        resultados_cuenta = self.obtener_saldo_cuenta(fecha_desde, fecha_hasta, cierre, cuenta_con_movimiento, cuenta_codigo_desde, cuenta_codigo_hasta)
        resultados_tercero = self.obtener_balance_prueba_tercero(fecha_desde, fecha_hasta, cierre, numero_identificacion, nombre_corto)        
        resultados_json = resultados_cuenta + resultados_tercero
        resultados_json.sort(key=lambda x: str(x['codigo']))
        if excel:
            excel_funciones = ExcelFunciones()
            wb = Workbook()
            ws = wb.active
            ws.title = "auxiliar_contacto"                                                
            excel_funciones.agregar_titulo(ws, "Auxiliar por contacto", "A", "I")                                   
            ws['A4'] = f"Fecha desde: {fecha_desde}"
            ws['A4'].font = excel_funciones.fuente_general           
            ws['A5'] = f"Fecha hasta: {fecha_hasta}"
            ws['A5'].font = excel_funciones.fuente_general 
            ws.append([])

            headers = ["Tipo", "Cuenta", "Nombre Cuenta", "Identificación", "Contacto", "Saldo anterior ($)", "Debitos ($)", "Creditos ($)", "Saldo actual ($)"]
            ws.append(headers)
            for registro in resultados_json:
                ws.append([
                    registro['tipo'],
                    registro['codigo'],
                    registro['nombre'],
                    registro['contacto_numero_identificacion'],
                    registro['contacto_nombre_corto'],
                    registro['saldo_anterior'],
                    registro['debito'],
                    registro['credito'],
                    registro['saldo_actual']
                ])    
            
            excel_funciones.aplicar_estilos(ws, 7, [6,7,8,9])                              
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'
            response['Content-Disposition'] = f'attachment; filename=auxiliar_contacto.xlsx'
            wb.save(response)
            return response
        else:
            return Response({'registros': resultados_json}, status=status.HTTP_200_OK) 

    @action(detail=False, methods=["post"], url_path=r'informe-auxiliar-general',)
    def informe_auxiliar_general(self, request):             
        raw = request.data
        parametros = raw.get("parametros", [])
        excel = raw.get('excel', False)
        pdf = raw.get('pdf', False)
        fecha_desde = parametros['fecha_desde']
        fecha_hasta = parametros['fecha_hasta']
        cuenta_codigo_desde = parametros.get('cuenta_codigo_desde', None)
        cuenta_codigo_hasta = parametros.get('cuenta_codigo_hasta', None)
        cierre = parametros['incluir_cierre']
        cuenta_con_movimiento = parametros['cuenta_con_movimiento'] 
        comprobante = parametros['comprobante']   
        resultados_cuenta = self.obtener_saldo_cuenta(fecha_desde, fecha_hasta, cierre, cuenta_con_movimiento, cuenta_codigo_desde, cuenta_codigo_hasta)
        resultados_tercero = self.obtener_balance_prueba_tercero(fecha_desde, fecha_hasta, cierre)
        resultados_movimiento = self.obtener_movimiento(fecha_desde, fecha_hasta, cierre, comprobante)
        resultados_json = resultados_cuenta + resultados_tercero + resultados_movimiento
        resultados_json.sort(key=lambda x: str(x['codigo']))
        if excel:
            excel_funciones = ExcelFunciones()
            wb = Workbook()
            ws = wb.active
            ws.title = "auxiliar_general"                                                
            excel_funciones.agregar_titulo(ws, "Auxiliar general", "A", "L")                                   
            ws['A4'] = f"Fecha desde: {fecha_desde}"
            ws['A4'].font = excel_funciones.fuente_general           
            ws['A5'] = f"Fecha hasta: {fecha_hasta}"
            ws['A5'].font = excel_funciones.fuente_general 
            ws.append([])
            headers = ["Tipo", "Cuenta", "Nombre Cuenta", "Identificación", "Contacto", "Comprobante", "Numero", "Fecha", "Saldo anterior ($)", "Debitos ($)", "Creditos ($)", "Saldo actual ($)"]
            ws.append(headers)
            for registro in resultados_json:
                ws.append([
                    registro['tipo'],
                    registro['codigo'],
                    registro['nombre'],
                    registro['contacto_numero_identificacion'],
                    registro['contacto_nombre_corto'],
                    registro['comprobante_nombre'],
                    registro['numero'],
                    registro['fecha'],
                    registro['saldo_anterior'],
                    registro['debito'],
                    registro['credito'],
                    registro['saldo_actual']
                ])    
            
            excel_funciones.aplicar_estilos(ws, 7, [9,10,11,12])                       
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'
            response['Content-Disposition'] = f'attachment; filename=auxiliar_general.xlsx'
            wb.save(response)
            return response
        else:
            return Response({'registros': resultados_json}, status=status.HTTP_200_OK) 

    @action(detail=False, methods=["post"], url_path=r'informe-base',)
    def informe_base(self, request):    
        raw = request.data
        parametros = raw.get("parametros", [])
        excel = raw.get('excel', False)
        pdf = raw.get('pdf', False)
        fecha_desde = parametros['fecha_desde']
        fecha_hasta = parametros['fecha_hasta']  
        cuenta_codigo_desde = parametros.get('cuenta_codigo_desde', None)
        cuenta_codigo_hasta = parametros.get('cuenta_codigo_hasta', None)
        contacto_id = parametros.get('contacto_id', None)                     
        resultados_movimiento = self.obtener_movimiento_base(fecha_desde, fecha_hasta, cuenta_codigo_desde, cuenta_codigo_hasta, contacto_id)
        resultados_json = resultados_movimiento
        resultados_json.sort(key=lambda x: str(x['cuenta_codigo']))

        if excel:
            excel_funciones = ExcelFunciones()
            wb = Workbook()
            ws = wb.active
            ws.title = "bases"                                                
            excel_funciones.agregar_titulo(ws, "Informe de bases", "A", "K")                                   
            ws['A4'] = f"Fecha desde: {fecha_desde}"
            ws['A4'].font = excel_funciones.fuente_general           
            ws['A5'] = f"Fecha hasta: {fecha_hasta}"
            ws['A5'].font = excel_funciones.fuente_general 
            ws.append([])
            headers = ["Cuenta", "Nombre Cuenta", "Identificación", "Contacto", "Comprobante", "Numero", "Fecha", "Detalle", "Debitos ($)", "Creditos ($)", "Base ($)"]
            ws.append(headers)
            for registro in resultados_json:
                ws.append([
                    registro['cuenta_codigo'],
                    registro['cuenta_nombre'],
                    registro['contacto_numero_identificacion'],
                    registro['contacto_nombre_corto'],
                    registro['comprobante_nombre'],
                    registro['numero'],                    
                    registro['fecha'],
                    registro['detalle'],
                    registro['debito'],
                    registro['credito'],
                    registro['base'],
                ])    
            excel_funciones.aplicar_estilos(ws, 7, [9,10,11])                                
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'
            response['Content-Disposition'] = f'attachment; filename=bases.xlsx'
            wb.save(response)
            return response
        else:
            return Response({'registros': resultados_json}, status=status.HTTP_200_OK)  

    @action(detail=False, methods=["post"], url_path=r'informe-certificado-retencion',)
    def informe_certificado_retencion(self, request):    
        raw = request.data
        parametros = raw.get("parametros", [])
        excel = raw.get('excel', False)
        pdf = raw.get('pdf', False)
        fecha_desde = parametros['fecha_desde']
        fecha_hasta = parametros['fecha_hasta']
        cuenta_codigo_desde = parametros.get('cuenta_codigo_desde', None)
        cuenta_codigo_hasta = parametros.get('cuenta_codigo_hasta', None)
        contacto_id = parametros.get('contacto_id', None)
        resultados_movimiento = self.obtener_movimiento_certificado_retencion(fecha_desde, fecha_hasta, cuenta_codigo_desde, cuenta_codigo_hasta, contacto_id)
        resultados_json = resultados_movimiento

        if excel:
            excel_funciones = ExcelFunciones()
            wb = Workbook()
            ws = wb.active
            ws.title = "Certificado de retencion"                                                
            excel_funciones.agregar_titulo(ws, "Certificado retenciones", "A", "F")                                   
            ws['A4'] = f"Fecha desde: {fecha_desde}"
            ws['A4'].font = excel_funciones.fuente_general           
            ws['A5'] = f"Fecha hasta: {fecha_hasta}"
            ws['A5'].font = excel_funciones.fuente_general 
            ws.append([])
            headers = ["Identificación", "Contacto", "Cuenta", "Nombre cuenta", "Monto del pago sujeto a retención ($)", "Retenido y consignado ($)"]
            ws.append(headers)
            for registro in resultados_json:
                ws.append([
                    registro['contacto_numero_identificacion'],
                    registro['contacto_nombre_corto'],
                    registro['cuenta_codigo'],
                    registro['cuenta_nombre'],                                        
                    registro['base_retenido'],
                    registro['retenido'],
                ])            
            excel_funciones.aplicar_estilos(ws, 7, [7,8,9])                             
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'
            response['Content-Disposition'] = f'attachment; filename=certificado_retencion.xlsx'
            wb.save(response)
            return response
        else:
            return Response({'registros': resultados_json}, status=status.HTTP_200_OK)                            