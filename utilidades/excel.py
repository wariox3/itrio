from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, numbers
from openpyxl.utils import get_column_letter

class WorkbookEstilos(Workbook):
    def __init__(self, wb):
        self.wb = wb
        self.font = Font(name='Arial', size=10, bold=False)
        self.bold_font = Font(name='Arial', size=10, bold=True)
        self.alignment = Alignment(horizontal='left')
        self.fill = PatternFill(start_color="9CDFEB", end_color="9CDFEB", fill_type="solid")

    def aplicar_estilos(self, formato_numero=None):
        for sheet in self.wb.sheetnames:
            ws = self.wb[sheet]
            for row in ws.iter_rows():
                for cell in row:
                    cell.font = self.font                    
                    if cell.row == 1: 
                        cell.font = self.bold_font
                        cell.fill = self.fill
                    else:
                        if formato_numero is not None and cell.column in formato_numero:
                            cell.number_format = '#,##0.00'
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal='right')                                                                                                
                    if isinstance(cell.value, bool):
                        cell.value = "SI" if cell.value else "NO"

            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value is not None:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception as e:
                        print(f"Error al calcular el ancho para la celda {cell.coordinate}: {e}")
                
                adjusted_width = (max_length + 1)
                ws.column_dimensions[column_letter].width = adjusted_width

