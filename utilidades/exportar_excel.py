from openpyxl import Workbook
from openpyxl.styles import Font, NamedStyle
from django.utils.encoding import smart_str
from django.http import HttpResponse

class ExportarExcel:
    def __init__(self, data, sheet_name="Datos", filename="export.xlsx"):
        self.data = data
        self.sheet_name = sheet_name
        self.filename = filename

    def export(self):
        wb = Workbook()
        ws = wb.active
        ws.title = self.sheet_name        
        estilo = Font(name='Arial', size=10)
        if self.data:
            headers = list(self.data[0].keys())
            ws.append(headers)                        
            estilo_encabezado = Font(name='Arial', size=10, bold=True)
            for cell in ws[1]:
                cell.font = estilo_encabezado

            for row_data in self.data:
                row = [smart_str(row_data.get(col, '')) for col in headers]
                ws.append(row)
                for cell in ws[ws.max_row]:
                    cell.font = estilo

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Access-Control-Expose-Headers'] = 'Content-Disposition'
        response['Content-Disposition'] = f'attachment; filename="{self.filename}"'
        wb.save(response)
        return response

