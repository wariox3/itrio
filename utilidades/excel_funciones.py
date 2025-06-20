from general.models.empresa import GenEmpresa
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from datetime import datetime

class ExcelFunciones:
    def __init__(self):
        self.fill = PatternFill(start_color="9CDFEB", end_color="9CDFEB", fill_type="solid")
        self.fuente_general = Font(name='Arial', size=10, bold=False)
        self.fuente_encabezado = Font(name='Arial', size=10, bold=True)
        
    def agregar_titulo(self, ws, titulo, columna_desde, columna_hasta):
        rango_titulo = f"{columna_desde}1:{columna_hasta}1"
        ws.merge_cells(rango_titulo)
        celda_titulo = ws[f"{columna_desde}1"]
        celda_titulo.value = titulo        
        celda_titulo.font = Font(name='Arial', size=14, bold=True)
        celda_titulo.alignment = Alignment(horizontal="center", vertical="center")
        celda_titulo.fill = self.fill

        empresa = GenEmpresa.objects.filter(pk=1).values('nombre_corto')[0]
        rango_empresa = f"{columna_desde}2:{columna_hasta}2"
        ws.merge_cells(rango_empresa)
        celda_empresa = ws[f"{columna_desde}2"]
        celda_empresa.value = empresa['nombre_corto']
        celda_empresa.font = Font(name='Arial', size=12, bold=True)
        celda_empresa.alignment = Alignment(horizontal="center", vertical="center")
        celda_empresa.fill = self.fill

        
        rango_datos = f"{columna_desde}3:{columna_hasta}3"
        ws.merge_cells(rango_datos)
        celda_datos = ws[f"{columna_desde}3"]
        celda_datos.value = f"Fecha generación: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        celda_datos.font = Font(name='Arial', size=8, bold=False)
        celda_datos.alignment = Alignment(horizontal="right")
        celda_datos.fill = self.fill        
        ws.append([])

    def aplicar_estilos(self, ws, numero_fila_titulo = 1, formato_numero = None):
        for row in ws.iter_rows(min_row=numero_fila_titulo):
            for cell in row:
                cell.font = self.fuente_general                    
                if cell.row == numero_fila_titulo: 
                    cell.font = self.fuente_encabezado
                    cell.fill = self.fill
                else:
                    if formato_numero is not None and cell.column in formato_numero:
                        cell.number_format = '#,##0.00'
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal='right')                                                                                                
                if isinstance(cell.value, bool):
                    cell.value = "SI" if cell.value else "NO"

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception as e:
                    print(f"Error al calcular el ancho para la celda {cell.coordinate}: {e}")                
            adjusted_width = (max_length + 1)
            ws.column_dimensions[column_letter].width = adjusted_width       