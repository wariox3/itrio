from general.models.empresa import GenEmpresa
from openpyxl import Workbook
from openpyxl.styles import Font, NamedStyle, PatternFill, Alignment, Font
from django.utils.encoding import smart_str
from django.http import HttpResponse
from decimal import Decimal
from io import BytesIO
from datetime import datetime

class ExcelExportar:
    HEADERS_MAP = {
        'id': 'ID',
        'numero': 'Numero',
        'fecha': 'Fecha',
        'fecha_vence': 'Vencimiento',
        'documento_tipo_id': 'Documento ID',
        'documento_tipo__nombre': 'Documento',
        'contacto_id': 'Contacto ID',
        'contacto__numero_identificacion': 'Numero identificacion',
        'contacto__nombre_corto': 'Contacto',
        'documento__contacto__nombre_corto': 'Contacto',
        'documento__fecha': 'Fecha',
        'documento__numero': 'Numero',
        'documento__documento_tipo__nombre': 'Documento',
        'item_id': 'Item ID',
        'item__nombre': 'Item',
        'cantidad': 'Cantidad',
        'precio': 'Precio',
        'descuento': 'Descuento',
        'subtotal': 'Subtotal',
        'base_impuesto': 'Base',
        'impuesto': 'Impuesto',
        'total': 'Total',
        'afectado': 'Abono',
        'pendiente': 'Pendiente',
        'estado_aprobado': 'Aprobado',
        'estado_anulado': 'Anulado',
        'estado_electronico': 'Electronico',
        'estado_electronico_enviado': 'Electronico enviado',
        'estado_electronico_notificado': 'Electronico notificado'
    }
			
    def __init__(self, data, nombre_hoja="documentos", nombre_archivo="documentos.xlsx", titulo='Documentos'):
        self.data = data
        self.nombre_hoja = nombre_hoja
        self.nombre_archivo = nombre_archivo
        self.titulo = titulo
        self.fill = PatternFill(start_color="9CDFEB", end_color="9CDFEB", fill_type="solid")

    def _get_encabezados(self):        
        if not self.data:
            return []
        
        original_headers = list(self.data[0].keys())
        return [self.HEADERS_MAP.get(header, header) for header in original_headers]

    def _get_encabezados_originales(self):
        return list(self.data[0].keys()) if self.data else []

    def agregar_titulo(self, ws, titulo, columna_desde, columna_hasta):
        rango_titulo = f"{columna_desde}1:{columna_hasta}1"
        ws.merge_cells(rango_titulo)
        celda_titulo = ws[f"{columna_desde}1"]
        celda_titulo.value = titulo        
        celda_titulo.font = Font(name='Arial', size=14, bold=True)
        celda_titulo.alignment = Alignment(horizontal="center", vertical="center")
        celda_titulo.fill = self.fill

        empresa = GenEmpresa.objects.filter(pk=1).values('nombre_corto')[0]
        rango_empresa = f"{columna_desde}2:{columna_hasta}2"
        ws.merge_cells(rango_empresa)
        celda_empresa = ws[f"{columna_desde}2"]
        celda_empresa.value = empresa['nombre_corto']
        celda_empresa.font = Font(name='Arial', size=12, bold=True)
        celda_empresa.alignment = Alignment(horizontal="center", vertical="center")
        celda_empresa.fill = self.fill

        
        rango_datos = f"{columna_desde}3:{columna_hasta}3"
        ws.merge_cells(rango_datos)
        celda_datos = ws[f"{columna_desde}3"]
        celda_datos.value = f"Fecha generación: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        celda_datos.font = Font(name='Arial', size=8, bold=False)
        celda_datos.alignment = Alignment(horizontal="right")
        celda_datos.fill = self.fill        
        ws.append([])

    def exportar(self):
        # Debe quedar write_only y no permite estilos para procesamiento de grandes volumenes
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title=self.nombre_hoja)        
        if self.data:
            headers = list(self.data[0].keys())
            ws.append(headers)                       
            for row_data in self.data:
                row = []
                for col in headers:
                    value = row_data.get(col, '')
                    if value is None:
                        value = ''
                    row.append(str(value))
                ws.append(row)    
        virtual_workbook = BytesIO()
        wb.save(virtual_workbook)
        virtual_workbook.seek(0)        
        response = HttpResponse(
            virtual_workbook.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Access-Control-Expose-Headers'] = 'Content-Disposition'
        response['Content-Disposition'] = f'attachment; filename="{self.nombre_archivo}"'
        return response
    
    def exportar_estilo(self):
        wb = Workbook()
        ws = wb.active
        ws.title = self.nombre_hoja        
        estilo = Font(name='Arial', size=10)        
        estilo_decimal = '#,##0.00'
        if self.data:
            headers = list(self.data[0].keys())
            ws.append(headers)                        
            estilo_encabezado = Font(name='Arial', size=10, bold=True)
            for cell in ws[1]:
                cell.font = estilo_encabezado

            for row_data in self.data:
                row = []
                for col in headers:
                    value = row_data.get(col, '')     
                    if value is None:
                        value = ''                               
                    if not isinstance(value, (int, float, Decimal)):
                        value = smart_str(value)
                    row.append(value)                
                ws.append(row)                
                
                for cell in ws[ws.max_row]:
                    if isinstance(cell.value, (float, Decimal)):
                        cell.number_format = estilo_decimal
                    if isinstance(cell.value, bool):
                        cell.value = "SI" if cell.value else "NO"                        
                    cell.font = estilo

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Access-Control-Expose-Headers'] = 'Content-Disposition'
        response['Content-Disposition'] = f'attachment; filename="{self.nombre_archivo}"'
        wb.save(response)
        return response    

    def exportar_informe(self):
        wb = Workbook()
        ws = wb.active
        ws.title = self.nombre_hoja

        if self.data:
            estilo_normal = Font(name='Arial', size=10)
            estilo_encabezado = Font(name='Arial', size=10, bold=True)
            estilo_decimal = NamedStyle(name="decimal_style")
            estilo_decimal.font = estilo_normal
            estilo_decimal.number_format = '#,##0.00'
            if "decimal_style" not in wb.named_styles:
                wb.add_named_style(estilo_decimal)

            encabezados = self._get_encabezados()
            encabezados_originales = self._get_encabezados_originales()
            num_columnas = len(encabezados)
            columna_final = chr(ord('A') + num_columnas - 1)
            self.agregar_titulo(ws, self.titulo, 'A', columna_final)

            # Encabezados
            for col_idx, col_name in enumerate(encabezados, start=1):
                cell = ws.cell(row=5, column=col_idx, value=col_name)
                cell.font = estilo_encabezado

            # Filas de datos
            for row_idx, row_data in enumerate(self.data, start=6):
                for col_idx, col in enumerate(encabezados_originales, start=1):
                    value = row_data.get(col, '')
                    if value is None:
                        value = ''
                    elif isinstance(value, bool):
                        value = 'SI' if value else 'NO'                        
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = estilo_normal

                    if isinstance(value, (float, Decimal)):
                        cell.style = estilo_decimal

        virtual_workbook = BytesIO()
        wb.save(virtual_workbook)
        virtual_workbook.seek(0)
        response = HttpResponse(
            virtual_workbook.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Access-Control-Expose-Headers'] = 'Content-Disposition'
        response['Content-Disposition'] = f'attachment; filename="{self.nombre_archivo}"'
        return response
    