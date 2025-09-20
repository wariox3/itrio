from rest_framework import viewsets, permissions, status
from rest_framework.response import Response
from rest_framework.decorators import action
from ruteo.models.visita import RutVisita
from ruteo.models.despacho import RutDespacho
from ruteo.models.franja import RutFranja
from ruteo.models.flota import RutFlota
from general.models.configuracion import GenConfiguracion
from general.models.archivo import GenArchivo
from contenedor.models import CtnDireccion
from ruteo.serializers.visita import RutVisitaSerializador, RutVistaTraficoSerializador, RutVistaListaSerializador, RutVisitaExcelSerializador, RutVisitaDetalleSerializador
from ruteo.servicios.visita import VisitaServicio
from datetime import datetime
from django.utils import timezone
from utilidades.holmio import Holmio
from utilidades.google import Google
from utilidades.backblaze import Backblaze
from utilidades.imagen import Imagen
from utilidades.utilidades import UtilidadGeneral
from django.db.models import Sum, Count, F
from django.db.models.functions import Coalesce
from django.db import transaction
from io import BytesIO
from rest_framework.filters import OrderingFilter
from django_filters.rest_framework import DjangoFilterBackend
from ruteo.filters.visita import VisitaFilter
from utilidades.excel_exportar import ExcelExportar
from decimal import Decimal, ROUND_HALF_UP
import re
import gc
import base64
import openpyxl

class RutVisitaViewSet(viewsets.ModelViewSet):
    queryset = RutVisita.objects.all()
    serializer_class = RutVisitaSerializador
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_class = VisitaFilter 
    serializadores = {
        'lista': RutVistaListaSerializador,
        'lista_completa' : RutVistaListaSerializador,
        'trafico' : RutVistaTraficoSerializador,
        'excel': RutVisitaExcelSerializador,
        'detalle': RutVisitaDetalleSerializador
    }

    def get_serializer_class(self):
        serializador_parametro = self.request.query_params.get('serializador', None)
        if not serializador_parametro or serializador_parametro not in self.serializadores:
            return RutVisitaSerializador
        return self.serializadores[serializador_parametro]

    def get_queryset(self):
        queryset = super().get_queryset()
        serializer_class = self.get_serializer_class()        
        select_related = getattr(serializer_class.Meta, 'select_related_fields', [])
        if select_related:
            queryset = queryset.select_related(*select_related)        
        campos = serializer_class.Meta.fields        
        if campos and campos != '__all__':
            queryset = queryset.only(*campos) 
        return queryset 

    def list(self, request, *args, **kwargs):
        #Ya no se va a usar lista deprecated
        if request.query_params.get('lista', '').lower() == 'true':
            self.pagination_class = None
        if request.query_params.get('lista_completa', '').lower() == 'true':
            self.pagination_class = None
        if request.query_params.get('excel') or request.query_params.get('excel_masivo'):
            queryset = self.filter_queryset(self.get_queryset())
            serializer = self.get_serializer(queryset, many=True)
            exporter = ExcelExportar(serializer.data, nombre_hoja="visitas", nombre_archivo="visitas.xlsx", titulo="Visitas")
            if request.query_params.get('excel'):
                return exporter.exportar_estilo()
            if request.query_params.get('excel_masivo'):
                return exporter.exportar()
        return super().list(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):        
        visita = self.get_object()
        if visita.despacho_id:
                despacho = RutDespacho.objects.get(pk=visita.despacho_id)                
                if despacho.estado_aprobado == True:
                    return Response({'mensaje': 'No se puede eliminar la visita porque el despacho esta aprobado.'}, status=status.HTTP_400_BAD_REQUEST)        
                else:
                    despacho.peso = despacho.peso - visita.peso
                    despacho.volumen = despacho.volumen - visita.volumen
                    despacho.visitas -= 1
                    despacho.save()        
        self.perform_destroy(visita)
        return Response(status=status.HTTP_204_NO_CONTENT) 

    @action(detail=False, methods=["post"], url_path=r'nuevo',)
    def nuevo(self, request):
        google = Google()
        franjas = RutFranja.objects.all()
        direccion_destinatario = VisitaServicio.limpiar_direccion(request.data.get('destinatario_direccion'))
        data = {
            'numero': request.data.get('numero', None),
            'documento': request.data.get('documento', None),
            'destinatario': request.data.get('destinatario', None),
            'destinatario_direccion': direccion_destinatario,
            'destinatario_telefono': request.data.get('destinatario_telefono', None),
            'destinatario_correo': request.data.get('destinatario_correo', None),
            'unidades': request.data.get('unidades', None),
            'peso': request.data.get('peso', None),
            'volumen': request.data.get('volumen', None),
            'tiempo_servicio': request.data.get('tiempo_servicio', None),
            'ciudad_id': request.data.get('ciudad', None),
            'estado_franja': False,
            'franja': None,
            'resultados': None,
            'latitud': None,
            'longitud': None,
        }
        if direccion_destinatario:                   
            direccion = CtnDireccion.objects.filter(direccion=direccion_destinatario).first()
            if direccion:
                data['estado_decodificado'] = True            
                data['latitud'] = direccion.latitud                        
                data['longitud'] = direccion.longitud
                data['destinatario_direccion_formato'] = direccion.direccion_formato
                data['resultados'] = direccion.resultados
                if direccion.cantidad_resultados > 1:
                    data['estado_decodificado_alerta'] = True                                
            else:
                respuesta = google.decodificar_direccion(data['destinatario_direccion'])
                if respuesta['error'] == False:   
                    data['estado_decodificado'] = True            
                    data['latitud'] = respuesta['latitud']
                    data['longitud'] = respuesta['longitud']
                    data['destinatario_direccion_formato'] = respuesta['direccion_formato']
                    data['resultados'] = respuesta['resultados']
                    if respuesta['cantidad_resultados'] > 1:
                        data['estado_decodificado_alerta'] = True
        if data['estado_decodificado'] == True:
            respuesta = VisitaServicio.ubicar_punto(franjas, data['latitud'], data['longitud'])
            if respuesta['encontrado']:
                data['franja'] = respuesta['franja']['id']
                data['estado_franja'] = True
            else:
                data['estado_franja'] = False
        serializer = RutVisitaSerializador(data=data)
        if serializer.is_valid():
            visita = serializer.save() 
            visitas = RutVisita.objects.filter(estado_despacho = False, estado_decodificado = True)
            VisitaServicio.ubicar(visitas)
            VisitaServicio.ordenar(visitas)    
            return Response({'mensaje': 'Visita creada exitosamente'}, status=status.HTTP_200_OK)    
        else: 
            return Response({'mensaje': 'Error en los datos', 'errores_validador': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)        

    @action(detail=False, methods=["post"], url_path=r'importar-excel',)
    def importar_excel(self, request):
        raw = request.data
        archivo_base64 = raw.get('archivo_base64')
        if archivo_base64:
            archivo_data = base64.b64decode(archivo_base64)
            archivo = BytesIO(archivo_data)
            wb = openpyxl.load_workbook(archivo)
            sheet = wb.active    
            data_modelo = []
            errores = False
            errores_datos = []    
            franjas = RutFranja.objects.all()
            google = Google()
            total_registros = sheet.max_row - 1
            if total_registros <= 1000:
                for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):                
                    if len(row) < 14:
                        return Response({'mensaje':'El archivo no tiene la estructura requerida', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)
                    fecha_texto = str(row[1])
                    try:                                        
                        fecha = datetime.strptime(fecha_texto, '%Y%m%d')
                    except ValueError:
                        fecha = None                    
                    documento = str(row[2])
                    direccion_destinatario = VisitaServicio.limpiar_direccion(row[4])
                    telefono_destinatario = str(row[5])
                    if telefono_destinatario:
                        telefono_destinatario[:50]                    
                    data = {
                        'numero': row[0],
                        'fecha':fecha,
                        'documento': documento[:30],
                        'destinatario': row[3],
                        'destinatario_direccion': direccion_destinatario,
                        'destinatario_telefono': telefono_destinatario,
                        'destinatario_correo': row[6],
                        'unidades': row[7],
                        'peso': row[8],
                        'volumen': row[9],
                        'latitud': None,
                        'longitud': None,
                        'estado_decodificado': False,
                        'tiempo_servicio': row[13],
                        'estado_franja': False,
                        'franja': None,
                        'resultados': None
                    }                 
                    if direccion_destinatario:                   
                        direccion = CtnDireccion.objects.filter(direccion=direccion_destinatario).first()
                        if direccion:
                            data['estado_decodificado'] = True            
                            data['latitud'] = direccion.latitud                        
                            data['longitud'] = direccion.longitud
                            data['destinatario_direccion_formato'] = direccion.direccion_formato
                            data['resultados'] = direccion.resultados
                            if direccion.cantidad_resultados > 1:
                                data['estado_decodificado_alerta'] = True                                
                        else:
                            respuesta = google.decodificar_direccion(data['destinatario_direccion'])
                            if respuesta['error'] == False:   
                                data['estado_decodificado'] = True            
                                data['latitud'] = respuesta['latitud']
                                data['longitud'] = respuesta['longitud']
                                data['destinatario_direccion_formato'] = respuesta['direccion_formato']
                                data['resultados'] = respuesta['resultados']
                                if respuesta['cantidad_resultados'] > 1:
                                    data['estado_decodificado_alerta'] = True
                    if data['estado_decodificado'] == True:
                        respuesta = VisitaServicio.ubicar_punto(franjas, data['latitud'], data['longitud'])
                        if respuesta['encontrado']:
                            data['franja'] = respuesta['franja']['id']
                            data['estado_franja'] = True
                        else:
                            data['estado_franja'] = False
                    serializer = RutVisitaSerializador(data=data)
                    if serializer.is_valid():
                        data_modelo.append(serializer.validated_data)
                    else:
                        errores = True
                        error_dato = {
                            'fila': i,
                            'errores': serializer.errors
                        }
                        errores_datos.append(error_dato)
                if not errores:
                    for detalle in data_modelo:
                        RutVisita.objects.create(**detalle)
                    gc.collect()
                    visitas = RutVisita.objects.filter(estado_despacho = False, estado_decodificado = True)
                    VisitaServicio.ubicar(visitas)
                    VisitaServicio.ordenar(visitas)                                        
                    return Response({'mensaje': 'Se importó el archivo con éxito'}, status=status.HTTP_200_OK)                
                else:
                    gc.collect()                    
                    return Response({'mensaje':'Errores de validacion', 'codigo':1, 'errores_validador': errores_datos}, status=status.HTTP_400_BAD_REQUEST)                                    
            else:
                return Response({'mensaje':'Solo se permiten importar hasta 1000 registros', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'mensaje':'Faltan parametros', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=["post"], url_path=r'importar-complemento',)
    def importar_complemento_action(self, request):
        raw = request.data 
        limite = raw.get('limite', 1)
        guia_desde = raw.get('guia_desde', None)
        guia_hasta = raw.get('guia_hasta', None)
        fecha_desde = raw.get('fecha_desde', None)
        fecha_hasta = raw.get('fecha_hasta', None)        
        pendiente_despacho = raw.get('pendiente_despacho', None)
        codigo_contacto = raw.get('codigo_contacto', None)
        codigo_destino = raw.get('codigo_destino', None)
        codigo_zona = raw.get('codigo_zona', None)
        codigo_despacho = raw.get('codigo_despacho', None)
        respuesta = VisitaServicio.importar_complemento(limite=limite, guia_desde=guia_desde, guia_hasta=guia_hasta, fecha_desde=fecha_desde, fecha_hasta=fecha_hasta, pendiente_despacho=pendiente_despacho, codigo_contacto=codigo_contacto, codigo_destino=codigo_destino, codigo_zona=codigo_zona, codigo_despacho=codigo_despacho, despacho_id=None)
        if respuesta['error'] == False:
            cantidad = respuesta['cantidad']
            visitas = respuesta['visitas_creadas']
            #visitas = RutVisita.objects.filter(estado_despacho = False, estado_decodificado = True)
            VisitaServicio.ubicar(visitas)            
            VisitaServicio.ordenar(visitas)            
            return Response({'mensaje': f'Se importaron {cantidad} guias con exito'}, status=status.HTTP_200_OK)
        else:
            return Response({'mensaje': respuesta['mensaje']}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["post"], url_path=r'decodificar',)
    def decodificar_action(self, request):
        visitas = RutVisita.objects.filter(estado_decodificado = None)[:1]
        if visitas.exists():            
            datosVisitas = []           
            for visita in visitas:  
                if visita.destinatario_direccion:
                    datosVisita = {
                        "cuenta": "1",
                        "modelo": "guia",
                        "canal": 3,
                        "codigo": visita.id,
                        "direccion": visita.destinatario_direccion,
                        "ciudad": visita.ciudad_id,
                        "principal": False,                
                    }    
                    datosVisitas.append(datosVisita)                
            if datosVisitas:
                datos = {
                    "lote": datosVisitas
                }
                #zinc = Zinc()
                #respuesta = zinc.decodificar_direccion_lote(datos)
                '''if respuesta['error'] == False:                     
                    datos = respuesta['datos']
                    visita.estado_decodificado = datos['decodificado']                                                                                                   
                    visita.latitud = datos['latitud']
                    visita.longitud = datos['longitud']                    
                    visita.save()'''
            return Response({'mensaje': 'Proceso exitoso'}, status=status.HTTP_200_OK)
        else:
            return Response({'mensaje': 'No hay guias pendientes por decodificar'}, status=status.HTTP_200_OK) 

    @action(detail=False, methods=["post"], url_path=r'ordenar',)
    def ordenar_action(self, request):      
        raw = request.data
        filtros = raw.get('filtros', [])           
        visitas = RutVisita.objects.filter(estado_despacho=False, estado_decodificado=True)        
        for filtro in filtros:
            operador = filtro.get('operador')
            propiedad = filtro['propiedad']
            valor = filtro['valor1']  
            if operador == 'in' and isinstance(valor, str):                      
                if ',' in valor:
                    valor = [int(v.strip()) for v in valor.split(',')]
                else:
                    valor = [int(valor.strip())]
                filtro['valor1'] = valor                    
            if operador == 'range':
                visitas = visitas.filter(**{f'{propiedad}__{operador}': (filtro['valor1'], filtro['valor2'])})
            elif operador:
                visitas = visitas.filter(**{f'{propiedad}__{operador}': filtro['valor1']})
            else:
                visitas = visitas.filter(**{propiedad: filtro['valor1']})
        if visitas.exists():
            respuesta = VisitaServicio.ordenar(visitas) 
            if respuesta['error'] == True:
                return Response({'mensaje': respuesta['mensaje']}, status=status.HTTP_400_BAD_REQUEST)
        return Response({'mensaje':'visitas ordenadas'}, status=status.HTTP_200_OK)
        
    @action(detail=False, methods=["post"], url_path=r'rutear')
    def rutear(self, request):
        raw = request.data
        filtros = raw.get('filtros')

        decimal_6_places = Decimal('0.000001')
        
        configuracion = GenConfiguracion.objects.filter(id=1).first()
        rutear_franja = getattr(configuracion, 'rut_rutear_franja', False)

        flota_disponible = list(
            RutFlota.objects.filter(vehiculo__estado_asignado=False)
            .select_related('vehiculo')
            .order_by('prioridad') 
        )
        if not flota_disponible:
            return Response({'mensaje': 'No hay vehículos disponibles'}, 
                        status=status.HTTP_400_BAD_REQUEST)

        visitas = RutVisita.objects.filter(
            estado_despacho=False, 
            estado_devolucion=False, 
            estado_novedad=False
        )

        if filtros:
            for filtro in filtros:
                operador = filtro.get('operador')
                propiedad = filtro['propiedad']
                valor = filtro['valor1']  
                if operador == 'in' and isinstance(valor, str):                      
                    if ',' in valor:
                        valor = [int(v.strip()) for v in valor.split(',')]
                    else:
                        valor = [int(valor.strip())]
                    filtro['valor1'] = valor                               
                if operador == 'range':
                    visitas = visitas.filter(**{f'{propiedad}__{operador}': (filtro['valor1'], filtro['valor2'])})
                elif operador:
                    visitas = visitas.filter(**{f'{propiedad}__{operador}': filtro['valor1']})
        visitas = visitas.order_by('orden')
        visitas_pendientes = list(visitas)
        despachos_creados = 0
        def vehiculo_puede_tomar_visita(vehiculo, visita, peso_actual, tiempo_actual, verificar_franja):
            if (peso_actual + visita.peso > vehiculo.capacidad or 
                tiempo_actual + visita.tiempo > vehiculo.tiempo):
                return False
            
            if not verificar_franja:
                return True
                
            return vehiculo.franjas.filter(id=visita.franja_id).exists()

        if rutear_franja:
            for flota_item in flota_disponible:
                vehiculo = flota_item.vehiculo
                peso_total = 0
                tiempo_total = 0
                unidades_total = 0
                despacho = None

                # Procesar solo si el vehículo no está ya asignado
                if vehiculo.estado_asignado:
                    continue

                # Intentar asignar todas las visitas posibles a este vehículo
                for visita in list(visitas_pendientes):
                    if vehiculo_puede_tomar_visita(vehiculo, visita, peso_total, tiempo_total, True):
                        if despacho is None:
                            despacho = RutDespacho.objects.create(
                                fecha=timezone.now(),
                                vehiculo=vehiculo,
                                peso=visita.peso,
                                volumen=visita.volumen,
                                tiempo=visita.tiempo,
                                tiempo_servicio=visita.tiempo_servicio,
                                tiempo_trayecto=visita.tiempo_trayecto,
                                unidades=visita.unidades,
                                visitas=1
                            )
                            vehiculo.estado_asignado = True
                            vehiculo.save()
                            despachos_creados += 1
                        else:
                            despacho.peso += visita.peso
                            despacho.volumen += visita.volumen
                            despacho.tiempo = (Decimal(despacho.tiempo) + Decimal(visita.tiempo)).quantize(
                                decimal_6_places,
                                rounding=ROUND_HALF_UP
                            )
                            despacho.tiempo_servicio = (Decimal(despacho.tiempo_servicio) + Decimal(visita.tiempo_servicio)).quantize(
                                decimal_6_places,
                                rounding=ROUND_HALF_UP
                            )
                            despacho.tiempo_trayecto = (Decimal(despacho.tiempo_trayecto) + Decimal(visita.tiempo_trayecto)).quantize(
                                decimal_6_places,
                                rounding=ROUND_HALF_UP
                            )
                            despacho.unidades += visita.unidades
                            despacho.visitas += 1
                            despacho.save()

                        peso_total += visita.peso
                        tiempo_total += visita.tiempo
                        unidades_total += visita.unidades
                        
                        visita.estado_despacho = True
                        visita.despacho = despacho
                        visita.save()
                        
                        visitas_pendientes.remove(visita)

                if visitas_pendientes:
                    continue
                else:
                    break 

            mensaje = (
                f"Operación completada: {despachos_creados} rutas creadas"
                + (f"Pendientes: {len(visitas_pendientes)} visitas (IDs: {', '.join(str(v.id) for v in visitas_pendientes)})" 
                if visitas_pendientes else "")
            )
            return Response({
                'mensaje': mensaje,
                'status': status.HTTP_200_OK
            })

        else:
            vehiculo_index = 0
            while vehiculo_index < len(flota_disponible) and visitas_pendientes:
                flota_item = flota_disponible[vehiculo_index]
                vehiculo = flota_item.vehiculo
                
                if vehiculo.estado_asignado:
                    vehiculo_index += 1
                    continue
                    
                peso_total = 0
                tiempo_total = 0
                unidades_total = 0
                despacho = None
                visitas_asignadas_este_vehiculo = 0

                # Hacer copia para iterar seguramente mientras removemos
                for visita in list(visitas_pendientes):
                    if vehiculo_puede_tomar_visita(vehiculo, visita, peso_total, tiempo_total, False):
                        if despacho is None:
                            despacho = RutDespacho.objects.create(
                                fecha=timezone.now(),
                                vehiculo=vehiculo,
                                peso=visita.peso,
                                volumen=visita.volumen,
                                tiempo=visita.tiempo,
                                tiempo_servicio=visita.tiempo_servicio,
                                tiempo_trayecto=visita.tiempo_trayecto,
                                unidades=visita.unidades,
                                visitas=1
                            )
                            vehiculo.estado_asignado = True
                            vehiculo.save()
                            despachos_creados += 1
                        else:
                            despacho.peso += visita.peso
                            despacho.volumen += visita.volumen
                            despacho.tiempo = (Decimal(despacho.tiempo) + Decimal(visita.tiempo)).quantize(
                                decimal_6_places,
                                rounding=ROUND_HALF_UP
                            )
                            despacho.tiempo_servicio = (Decimal(despacho.tiempo_servicio) + Decimal(visita.tiempo_servicio)).quantize(
                                decimal_6_places,
                                rounding=ROUND_HALF_UP
                            )
                            despacho.tiempo_trayecto = (Decimal(despacho.tiempo_trayecto) + Decimal(visita.tiempo_trayecto)).quantize(
                                decimal_6_places,
                                rounding=ROUND_HALF_UP
                            )
                            despacho.unidades += visita.unidades
                            despacho.visitas += 1
                            despacho.save()

                        peso_total += visita.peso
                        tiempo_total += visita.tiempo
                        unidades_total += visita.unidades
                        
                        visita.estado_despacho = True
                        visita.despacho = despacho
                        visita.save()
                        
                        visitas_pendientes.remove(visita)
                        visitas_asignadas_este_vehiculo += 1

                vehiculo_index += 1

            mensaje = (
                f"Operación completada: {despachos_creados} rutas creadas"
                + (f"Pendientes: {len(visitas_pendientes)} visitas (IDs: {', '.join(str(v.id) for v in visitas_pendientes)})" 
                if visitas_pendientes else "")
            )

            return Response({
                'mensaje': mensaje,
                'status': status.HTTP_200_OK
            })
        
    @action(detail=False, methods=["post"], url_path=r'ubicar',)
    def ubicar_action(self, request):             
        raw = request.data
        visitas = RutVisita.objects.filter(estado_despacho = False, estado_decodificado = True)
        cantidad = VisitaServicio.ubicar(visitas)          
        return Response({'mensaje': f'Se asignó franja a {cantidad} de visitas'}, status=status.HTTP_200_OK)

    @action(detail=False, methods=["post"], url_path=r'ubicar-punto',)
    def ubicar_punto(self, request):             
        raw = request.data
        latitud = raw.get('latitud')
        longitud = raw.get('longitud')           
        if latitud and longitud:
            franjas = RutFranja.objects.all()
            respuesta = VisitaServicio.ubicar_punto(franjas, latitud, longitud)
            return Response(respuesta, status=status.HTTP_200_OK)                                                    
        else:
            return Response({'mensaje':'Faltan parametros', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)  

    @action(detail=False, methods=["post"], url_path=r'eliminar-todos',)
    def eliminar_todos(self, request):             
        raw = request.data
        estado_decodificado = raw.get('estado_decodificado', None)
        if estado_decodificado == False:
            RutVisita.objects.filter(estado_decodificado=False).delete()
        else:
            RutVisita.objects.filter(estado_despacho=False).delete()
        return Response({'mensaje':'eliminados'}, status=status.HTTP_200_OK) 

    @action(detail=False, methods=["post"], url_path=r'resumen',)
    def resumen(self, request):
        raw = request.data
        filtros = raw.get('filtros')
        visitas = RutVisita.objects.filter(estado_despacho=False, estado_devolucion=False)
        errores = RutVisita.objects.filter(estado_decodificado=False, estado_despacho= False)
        alertas = RutVisita.objects.filter(estado_decodificado_alerta=True, estado_despacho= False)
        if filtros:
            for filtro in filtros:                    
                operador = filtro.get('operador', None)
                valor = filtro['valor1']  
                if operador == 'in' and isinstance(valor, str):                      
                    if ',' in valor:
                        valor = [int(v.strip()) for v in valor.split(',')]
                    else:
                        valor = [int(valor.strip())]
                    filtro['valor1'] = valor                    

                if operador == 'range':
                    visitas = visitas.filter(**{filtro['propiedad']+'__'+operador: (filtro['valor1'], filtro['valor2'])})
                    errores = errores.filter(**{filtro['propiedad']+'__'+operador: (filtro['valor1'], filtro['valor2'])})
                    alertas = alertas.filter(**{filtro['propiedad']+'__'+operador: (filtro['valor1'], filtro['valor2'])})
                else:
                    visitas = visitas.filter(**{filtro['propiedad']+'__'+operador: filtro['valor1']})
                    errores = errores.filter(**{filtro['propiedad']+'__'+operador: filtro['valor1']})
                    alertas = alertas.filter(**{filtro['propiedad']+'__'+operador: filtro['valor1']})


        visitas = visitas.aggregate(
            cantidad=Count('id'), 
            peso=Coalesce(Sum('peso'), 0.0),
            unidades=Coalesce(Sum('unidades'), 0.0),
            tiempo=Coalesce(Sum('tiempo'), Decimal(0.0)),
            tiempo_servicio=Coalesce(Sum('tiempo_servicio'), Decimal(0.0)),
            tiempo_trayecto=Coalesce(Sum('tiempo_trayecto'), Decimal(0.0))
            
        )    
        errores = errores.aggregate(
            cantidad=Count('id'))        
        alertas = alertas.aggregate(
            cantidad=Count('id'))        
        return Response({'resumen': visitas, 'errores': errores, 'alertas':alertas}, status=status.HTTP_200_OK)   

    @action(detail=False, methods=["post"], url_path=r'resumen-pendiente',)
    def resumen_pendiente(self, request):
        raw = request.data
        filtros = raw.get('filtros')
        visitas = RutVisita.objects.filter(estado_despacho=False, estado_devolucion=False)        
        if filtros:
            for filtro in filtros:
                operador = filtro.get('operador', None)
                valor = filtro['valor1']  
                if operador == 'in' and isinstance(valor, str):                      
                    if ',' in valor:
                        valor = [int(v.strip()) for v in valor.split(',')]
                    else:
                        valor = [int(valor.strip())]
                    filtro['valor1'] = valor                 
                
                if operador == 'range':
                    visitas = visitas.filter(**{filtro['propiedad']+'__'+operador: (filtro['valor1'], filtro['valor2'])})
                else:
                    visitas = visitas.filter(**{filtro['propiedad']+'__'+operador: filtro['valor1']})                    
        visitas = visitas.values('franja_codigo').annotate(
            cantidad=Count('id'), 
            peso=Coalesce(Sum('peso'), 0.0),
            unidades=Coalesce(Sum('unidades'), 0.0),
            tiempo=Coalesce(Sum('tiempo'), Decimal(0.0)),
            tiempo_servicio=Coalesce(Sum('tiempo_servicio'), Decimal(0.0)),
            tiempo_trayecto=Coalesce(Sum('tiempo_trayecto'), Decimal(0.0)),
            )    
        return Response({'resumen': visitas}, status=status.HTTP_200_OK) 

    @action(detail=False, methods=["post"], url_path=r'seleccionar-direccion-alternativa',)
    def seleccionar_direccion_alternativa(self, request):             
        raw = request.data
        id = raw.get('id')
        latitud = raw.get('latitud')
        longitud = raw.get('longitud')
        destinatario_direccion_formato = raw.get('destinatario_direccion_formato')
        if id and latitud and longitud and destinatario_direccion_formato:
            try:                
                visita = RutVisita.objects.get(pk=id)
                franjas = RutFranja.objects.all()
                respuesta = VisitaServicio.ubicar_punto(franjas, latitud, longitud)
                if respuesta['encontrado']:                    
                    visita.franja_id = respuesta['franja']['id']
                    visita.estado_franja = True
                else:
                    visita.franja_id = None
                    visita.estado_franja = False 
                visita.latitud = latitud
                visita.longitud = longitud
                visita.destinatario_direccion_formato = destinatario_direccion_formato
                visita.estado_decodificado_alerta = False
                visita.save()               
                return Response({'mensaje': 'Se actualizo la visita'}, status=status.HTTP_200_OK)
            except RutVisita.DoesNotExist:
                return Response({'mensaje':'La visita no existe', 'codigo':15}, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'mensaje':'Faltan parametros', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)    
        
    @action(detail=False, methods=["post"], url_path=r'actualizar-direccion',)
    def actualizar_direccion(self, request):             
        raw = request.data
        id = raw.get('id')
        destinatario_direccion = raw.get('destinatario_direccion')
        numero = raw.get('numero')
        documento = raw.get('documento')
        destinatario = raw.get('destinatario')
        destinatario_telefono = raw.get('destinatario_telefono')
        unidades = raw.get('unidades')
        peso = raw.get('peso')
        volumen = raw.get('volumen')
        if id and destinatario and destinatario_direccion:
            try:  
                google = Google()              
                visita = RutVisita.objects.get(pk=id)                                 
                destinatario_direccion = destinatario_direccion.replace("\t", "").replace("\n", "")
                destinatario_direccion = re.sub(r'[\s\u2000-\u200F\u3000\u31A0]+', ' ', destinatario_direccion).strip()   
                destinatario_direccion = re.sub(r'[\s\u2000-\u200F\u3000\u3164]+', ' ', destinatario_direccion).strip()                 
                destinatario_direccion = re.sub(r'\s+', ' ', destinatario_direccion.strip())                    
                destinatario_direccion = destinatario_direccion[:150]
                if visita.destinatario_direccion != destinatario_direccion:
                    visita.destinatario_direccion = destinatario_direccion
                    visita.estado_decodificado = False
                    visita.estado_decodificado_alerta = False
                    if destinatario_direccion:                   
                        direccion = CtnDireccion.objects.filter(direccion=destinatario_direccion).first()
                        if direccion:
                            visita.estado_decodificado = True            
                            visita.latitud = direccion.latitud                        
                            visita.longitud = direccion.longitud
                            visita.destinatario_direccion_formato = direccion.direccion_formato
                            visita.resultados = direccion.resultados
                            if direccion.cantidad_resultados > 1:
                                visita.estado_decodificado_alerta = True                                
                        else:
                            respuesta = google.decodificar_direccion(destinatario_direccion)
                            if respuesta['error'] == False:   
                                visita.estado_decodificado = True            
                                visita.latitud = respuesta['latitud']
                                visita.longitud = respuesta['longitud']
                                visita.destinatario_direccion_formato = respuesta['direccion_formato']
                                visita.resultados = respuesta['resultados']
                                if respuesta['cantidad_resultados'] > 1:
                                    visita.estado_decodificado_alerta = True                
                    if visita.estado_decodificado == True:
                        franjas = RutFranja.objects.all()
                        respuesta = VisitaServicio.ubicar_punto(franjas, visita.latitud, visita.longitud)
                        if respuesta['encontrado']:
                            visita.franja_id = respuesta['franja']['id']
                            visita.estado_franja = True
                        else:
                            visita.franja_id = None
                            visita.estado_franja = False                                                     
                visita.numero = numero
                visita.documento = documento
                visita.destinatario = destinatario
                visita.destinatario_telefono = destinatario_telefono
                visita.unidades = unidades
                visita.peso = peso
                visita.volumen = volumen
                visita.save()               
                return Response({'mensaje': 'Se actualizo la visita'}, status=status.HTTP_200_OK)
            except RutVisita.DoesNotExist:
                return Response({'mensaje':'La visita no existe', 'codigo':15}, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'mensaje':'Faltan parametros', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)    

    @action(detail=False, methods=["post"], url_path=r'despacho-retirar',)
    def despacho_retirar(self, request):             
        raw = request.data
        id = raw.get('id')
        if id:
            try:                               
                visita = RutVisita.objects.get(pk=id) 
                despacho = RutDespacho.objects.get(pk=visita.despacho_id)
                if despacho.estado_aprobado == True:
                    return Response({'mensaje': 'No se puede retirar la visita porque el despacho esta aprobado.'}, status=status.HTTP_400_BAD_REQUEST)        
                else:
                    despacho.peso = despacho.peso - visita.peso
                    despacho.volumen = despacho.volumen - visita.volumen
                    despacho.visitas -= 1
                    despacho.save()                 
                    visita.despacho = None
                    visita.estado_despacho = False
                    visita.save()               
                    return Response({'mensaje': 'Se retiro la visita del despacho'}, status=status.HTTP_200_OK)
            except RutVisita.DoesNotExist:
                return Response({'mensaje':'La visita no existe', 'codigo':15}, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'mensaje':'Faltan parametros', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)     

    @action(detail=False, methods=["post"], url_path=r'despacho-cambiar',)
    def despacho_cambiar(self, request):             
        raw = request.data
        id = raw.get('id')
        despacho_id = raw.get('despacho_id')
        if id and despacho_id:            
            try:                               
                visita = RutVisita.objects.get(pk=id) 
                if visita.despacho_id != despacho_id:               
                    try:
                        despacho_nuevo = RutDespacho.objects.get(pk=despacho_id)
                        if visita.despacho_id:                                                
                            despacho_actual = RutDespacho.objects.get(pk=visita.despacho_id)
                            despacho_actual.peso = despacho_actual.peso - visita.peso
                            despacho_actual.volumen = despacho_actual.volumen - visita.volumen
                            despacho_actual.visitas -= 1
                            despacho_actual.save()

                            despacho_nuevo.peso = despacho_nuevo.peso + visita.peso
                            despacho_nuevo.volumen = despacho_nuevo.volumen + visita.volumen
                            despacho_nuevo.visitas += 1
                            despacho_nuevo.save()

                            visita.despacho = despacho_nuevo
                            visita.save() 
                            return Response({'mensaje': 'Se cambio la visita de despacho'}, status=status.HTTP_200_OK)                        
                        else:
                            return Response({'mensaje':'La visita no tiene despacho', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)                                                                            
                    except RutDespacho.DoesNotExist:
                        return Response({'mensaje':'El despacho no existe', 'codigo':15}, status=status.HTTP_400_BAD_REQUEST)                
                else:
                    return Response({'mensaje':'La visita ya está en este despacho', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)                
            except RutVisita.DoesNotExist:
                return Response({'mensaje':'La visita no existe', 'codigo':15}, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'mensaje':'Faltan parametros', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["post"], url_path=r'entrega',)
    def entrega_action(self, request):                     
        id = request.POST.get('id')
        imagenes = request.FILES.getlist('imagenes')
        firmas = request.FILES.getlist('firmas')
        fecha_entrega_parametro = request.POST.get('fecha_entrega')
        datos_adicionales = request.POST.get('datos_adicionales')        
        if id and fecha_entrega_parametro:
            try:
                fecha_entrega_nativa = datetime.strptime(fecha_entrega_parametro, '%Y-%m-%d %H:%M')
                fecha_entrega = timezone.make_aware(fecha_entrega_nativa)
                if fecha_entrega > timezone.now():
                    return Response({'mensaje':'La fecha de entrega no puede ser mayor a la fecha actual', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)                             
            except ValueError:
                return Response({'mensaje':'Formato de fecha inválido. Use YYYY-MM-DD HH:MM', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)
            try:
                visita = RutVisita.objects.get(pk=id)                            
            except RutVisita.DoesNotExist:
                #return Response({'mensaje':'La visita no existe', 'codigo':15}, status=status.HTTP_400_BAD_REQUEST)           
                return Response({'mensaje': f'La visita no existe'}, status=status.HTTP_200_OK)
            if visita.estado_entregado == False:  
                if visita.despacho_id:
                    try:
                        despacho = RutDespacho.objects.get(pk=visita.despacho_id)
                    except ValueError:
                        return Response({'mensaje':'El despacho no existe', 'codigo':15}, status=status.HTTP_400_BAD_REQUEST) 
                    with transaction.atomic():                                                                              
                        datos_entrega = UtilidadGeneral.json_texto(datos_adicionales)
                        visita.estado_entregado = True
                        visita.fecha_entrega = fecha_entrega
                        visita.datos_entrega = datos_entrega
                        visita.save()
                        RutDespacho.objects.filter(pk=visita.despacho_id).update(
                            visitas_entregadas=F('visitas_entregadas') + 1
                        )                                 
                        backblaze = Backblaze()
                        tenant = request.tenant.schema_name                    
                        if imagenes:                        
                            for imagen in imagenes:       
                                #file_content = imagen.read()                                  
                                file_content = Imagen.comprimir_imagen_jpg(imagen, calidad=20, max_width=1920)
                                nombre_archivo = f'{id}.jpg'                                                   
                                id_almacenamiento, tamano, tipo, uuid, url = backblaze.subir_data(file_content, tenant, nombre_archivo)
                                archivo = GenArchivo()
                                archivo.archivo_tipo_id = 2
                                archivo.almacenamiento_id = id_almacenamiento
                                archivo.nombre = nombre_archivo
                                archivo.tipo = tipo
                                archivo.tamano = tamano
                                archivo.uuid = uuid
                                archivo.codigo = id
                                archivo.modelo = "RutVisita"
                                archivo.url = url
                                archivo.save()
                        if firmas:
                            for firma in firmas:       
                                # No comprimir porque daña el png
                                file_content = firma.read()                                                                                           
                                nombre_archivo = f'{id}.png'                                                   
                                id_almacenamiento, tamano, tipo, uuid, url = backblaze.subir_data(file_content, tenant, nombre_archivo)
                                archivo = GenArchivo()
                                archivo.archivo_tipo_id = 3
                                archivo.almacenamiento_id = id_almacenamiento
                                archivo.nombre = nombre_archivo
                                archivo.tipo = tipo
                                archivo.tamano = tamano
                                archivo.uuid = uuid
                                archivo.codigo = id
                                archivo.modelo = "RutVisita"
                                archivo.url = url
                                archivo.save()                        
                        configuracion = GenConfiguracion.objects.filter(pk=1).values('rut_sincronizar_complemento')[0]
                        if configuracion['rut_sincronizar_complemento']:
                            imagenes_b64 = []
                            if imagenes:                        
                                for imagen in imagenes:    
                                    imagen.seek(0)   
                                    file_content = imagen.read()   
                                    base64_encoded = base64.b64encode(file_content).decode('utf-8')                                                    
                                    imagenes_b64.append({
                                        'base64': base64_encoded,
                                    })     
                            firmas_b64 = []
                            if firmas:
                                for firma in firmas:
                                    firma.seek(0)       
                                    file_content = firma.read()   
                                    base64_encoded = base64.b64encode(file_content).decode('utf-8')                                                    
                                    firmas_b64.append({
                                        'base64': base64_encoded,
                                    })                                                                                                                                                                                                     
                            VisitaServicio.entrega_complemento(visita, imagenes_b64, firmas_b64, datos_entrega)
                        return Response({'mensaje': f'Entrega con exito'}, status=status.HTTP_200_OK)
                else:
                    #return Response({'mensaje':'La visita no tiene despacho', 'codigo':15}, status=status.HTTP_400_BAD_REQUEST)                  
                    return Response({'mensaje': f'La visita no tiene despacho'}, status=status.HTTP_200_OK)    
            else:
                return Response({'mensaje': f'La visita ya fue entregada con anterioridad'}, status=status.HTTP_200_OK)
                #return Response({'mensaje':'La visita ya fue entregada con anterioridad', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST) 
        else:
            return Response({'mensaje':'Faltan parametros', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["post"], url_path=r'consulta-documento',)
    def consulta_documento(self, request):             
        raw = request.data
        despacho_id = raw.get('despacho_id')    
        estado_despacho = raw.get('estado_despacho', None)
        numero = raw.get('numero')
        if numero:
            visita = RutVisita.objects.filter(numero=numero)   
            if despacho_id:
                visita = visita.filter(despacho_id=despacho_id)  
            if estado_despacho == True:  
                visita = visita.filter(estado_despacho=True)
            if estado_despacho == False:  
                visita = visita.filter(estado_despacho=False)
            visita = visita.first()
            if visita:
                return Response({'id': visita.id}, status=status.HTTP_200_OK)
            else:
                return Response({'mensaje':f'La visita con numero {numero} no existe', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)                                                              
        else:
            return Response({'mensaje':'Faltan parametros', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)    

    @action(detail=False, methods=["post"], url_path=r'liberar',)
    def liberar(self, request):             
        raw = request.data
        id = raw.get('id')        
        if id:
            try:
                visita = RutVisita.objects.get(pk=id)                            
            except RutVisita.DoesNotExist:
                return Response({'mensaje':'La visita no existe', 'codigo':15}, status=status.HTTP_400_BAD_REQUEST)            
            if visita.estado_entregado == False and visita.estado_despacho == True:                
                despacho = RutDespacho.objects.get(pk=visita.despacho_id)                
                despacho.visitas_liberadas += 1
                despacho.peso = despacho.peso - visita.peso
                despacho.volumen = despacho.volumen - visita.volumen
                despacho.tiempo = despacho.tiempo - visita.tiempo
                despacho.tiempo_servicio = despacho.tiempo_servicio - visita.tiempo_servicio
                despacho.tiempo_trayecto = despacho.tiempo_trayecto - visita.tiempo_trayecto
                despacho.visitas = despacho.visitas - 1    
                if visita.estado_novedad:
                    despacho.visitas_novedad = despacho.visitas_novedad - 1              
                despacho.save()  
                visita.estado_despacho = False 
                visita.despacho = None           
                visita.save()                        
                return Response({'mensaje': f'Se libero con exito'}, status=status.HTTP_200_OK)
            else:
                return Response({'mensaje':'La visita ya fue entregada o no esta despachada', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)    
        else:
            return Response({'mensaje':'Faltan parametros', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)        
            
    @action(detail=False, methods=["post"], url_path=r'entrega-complemento',)
    def entrega_complemento_action(self, request):   
        backblaze = Backblaze()
        visitas = RutVisita.objects.filter(estado_entregado=True, estado_entregado_complemento=False)                
        for visita in visitas:
            imagenes_b64 = []
            archivos = GenArchivo.objects.filter(modelo='RutVisita', codigo=visita.id, archivo_tipo_id=2)
            for archivo in archivos:
                contenido = backblaze.descargar_bytes(archivo.almacenamiento_id)
                if contenido is not None:
                    contenido_base64 = base64.b64encode(contenido).decode('utf-8')                    
                    imagenes_b64.append({
                        'comprimido': True,
                        'base64': contenido_base64,
                    })   
            firmas_b64 = []
            archivos = GenArchivo.objects.filter(modelo='RutVisita', codigo=visita.id, archivo_tipo_id=3)
            for archivo in archivos:
                contenido = backblaze.descargar_bytes(archivo.almacenamiento_id)
                if contenido is not None:
                    contenido_base64 = base64.b64encode(contenido).decode('utf-8')                    
                    firmas_b64.append({
                        'base64': contenido_base64,
                    })                              
            VisitaServicio.entrega_complemento(visita, imagenes_b64, firmas_b64, visita.datos_entrega)
        return Response({'mensaje': f'Entrega complemento {visitas.count()}'}, status=status.HTTP_200_OK)
    