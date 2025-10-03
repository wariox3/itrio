from rest_framework import viewsets, permissions, status
from rest_framework.response import Response
from rest_framework.filters import OrderingFilter
from django_filters.rest_framework import DjangoFilterBackend
from general.models.documento_detalle import GenDocumentoDetalle
from general.models.documento_impuesto import GenDocumentoImpuesto
from general.models.documento import GenDocumento
from general.models.item import GenItem
from general.models.item_impuesto import GenItemImpuesto
from general.models.impuesto import GenImpuesto
from inventario.models.almacen import InvAlmacen
from contabilidad.models.grupo import ConGrupo
from general.serializers.documento_detalle import GenDocumentoDetalleListaDetalleCuentaSerializador, GenDocumentoDetalleSerializador, GenDocumentoDetalleInformeVentaSerializador, GenDocumentoDetalleInformeInventarioSerializador, GenDocumentoDetalleNominaSerializador, GenDocumentoDetalleNominaExcelSerializador, GenDocumentoDetalleCreditoPagoSerializador, GenDocumentoDetalleListaAgregarSerializador
from general.serializers.documento_impuesto import GenDocumentoImpuestoSerializador
from general.filters.documento_detalle import DocumentoDetalleFilter
from utilidades.excel_exportar import ExcelExportar
from rest_framework.decorators import action
import base64
from io import BytesIO
import openpyxl

class DocumentoDetalleViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, OrderingFilter] 
    filterset_class = DocumentoDetalleFilter
    queryset = GenDocumentoDetalle.objects.all()
    serializadores = {
        'informe_venta': GenDocumentoDetalleInformeVentaSerializador,
        'informe_inventario': GenDocumentoDetalleInformeInventarioSerializador,
        'informe_nomina_detalle': GenDocumentoDetalleNominaExcelSerializador,
        'nomina': GenDocumentoDetalleNominaSerializador,
        'lista_detalle_cuenta': GenDocumentoDetalleListaDetalleCuentaSerializador,
        'lista_detalle_item': GenDocumentoDetalleListaDetalleCuentaSerializador,
        'credito_pago' : GenDocumentoDetalleCreditoPagoSerializador,
        'lista_agregar' : GenDocumentoDetalleListaAgregarSerializador
    }
    
    def get_serializer_class(self):
        serializador_parametro = self.request.query_params.get('serializador', None)
        if not serializador_parametro or serializador_parametro not in self.serializadores:
            return GenDocumentoDetalleSerializador
        return self.serializadores[serializador_parametro]    

    def get_queryset(self):
        queryset = super().get_queryset()
        serializer_class = self.get_serializer_class()        
        select_related = getattr(serializer_class.Meta, 'select_related_fields', [])
        if select_related:
            queryset = queryset.select_related(*select_related)        
        campos = serializer_class.Meta.fields        
        if campos and campos != '__all__':
            queryset = queryset.only(*campos)
        return queryset            

    def list(self, request, *args, **kwargs):
        if request.query_params.get('excel'):
            queryset = self.filter_queryset(self.get_queryset())
            serializer = self.get_serializer(queryset, many=True)
            titulo = 'Informe documentos detalles'
            nombre_hoja = "documentos_detalles"
            nombre_archivo = "documentos_detalles.xlsx"
            if request.query_params.get('excel_masivo'):
                exporter = ExcelExportar(serializer.data, nombre_hoja, nombre_archivo)
                return exporter.exportar() 
            elif request.query_params.get('excel_informe'): 
                serializador_parametro = self.request.query_params.get('serializador', None)                
                if serializador_parametro == 'informe_venta':
                    titulo = 'Ventas por ítem' 
                    nombre_archivo = "ventas_por_ítem.xlsx"  
                    nombre_hoja = 'ventas_por_ítem'    
                if serializador_parametro == 'informe_nomina_detalle':
                    titulo = 'Nomina detalle' 
                    nombre_archivo = "nóminas_detalles.xlsx"  
                    nombre_hoja = 'nóminas_detalles'   
                exporter = ExcelExportar(serializer.data, nombre_hoja, nombre_archivo, titulo)
                return exporter.exportar_informe()                    
            else:
                exporter = ExcelExportar(serializer.data, nombre_hoja, nombre_archivo)
                return exporter.exportar_estilo()            
        return super().list(request, *args, **kwargs)

    def create(self, request):
        raw = request.data
        documentoDetalleSerializador = GenDocumentoDetalleSerializador(data=raw)
        if documentoDetalleSerializador.is_valid():
            documentoDetalleSerializador.save()            
            return Response({'documento': documentoDetalleSerializador.data}, status=status.HTTP_200_OK)
        return Response({'mensaje':'Errores de validacion', 'codigo':14, 'validaciones': documentoDetalleSerializador.errors}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["post"], url_path=r'importar_detalle_venta',)
    def importar_detalle_venta(self, request):
        raw = request.data
        documento_id = raw.get('documento_id')
        archivo_base64 = raw.get('archivo_base64')
        if documento_id and archivo_base64:
            try:
                documento = GenDocumento.objects.get(pk=documento_id)
            except GenDocumento.DoesNotExist:
                return Response({'mensaje':'El documento no existe', 'codigo':15}, status=status.HTTP_400_BAD_REQUEST)
            try:
                archivo_data = base64.b64decode(archivo_base64)
                archivo = BytesIO(archivo_data)
                wb = openpyxl.load_workbook(archivo)
                sheet = wb.active         
            except Exception as e:     
                return Response({'mensaje':'Error procesando el archivo', 'codigo':15}, status=status.HTTP_400_BAD_REQUEST)  
            
            data_documento_detalle = []
            errores = False
            errores_datos = []
            registros_importados = 0
            if documento.documento_tipo_id in [1]:
                for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    if len(row) == 4:
                        data = {
                            'item': row[0],
                            'almacen':row[1],
                            'cantidad': row[2] if row[2] is not None else 0,
                            'precio': row[3] if row[3] is not None else 0,              
                        }                    
                        if not data['item']:
                            error_dato = {
                                'fila': i,
                                'errores': {
                                    'item': ['Debe digitar el código del item']
                                }
                            }
                            errores_datos.append(error_dato)
                            errores = True
                        else:
                            item = GenItem.objects.filter(id=data['item']).first()
                            if item is None:
                                error_dato = {
                                    'fila': i,
                                    'errores': {
                                        'item': [f'El item {data["item"]} no existe']
                                    }
                                }
                                errores_datos.append(error_dato)
                                errores = True
                            else:
                                data['item_id'] = item.id

                        if not data['almacen']:
                            error_dato = {
                                'fila': i,
                                'errores': {
                                    'almacen': 'Debe digitar el código del almacen',
                                }
                            }
                            errores_datos.append(error_dato)
                            errores = True  
                        else:
                            almacen = InvAlmacen.objects.filter(id=data['almacen']).first()
                            if almacen is None:
                                errores = True
                                error_dato = {
                                    'fila': i,
                                    'errores': {
                                        'almacen': [f'El almacen con código {data["almacen"]} no existe']
                                    }
                                }
                                errores_datos.append(error_dato)
                            else:
                                data['almacen_id'] = almacen.id                                 
                        data_documento_detalle.append(data) 
                    else:
                        error_dato = {
                            'fila': i,
                            'errores': {
                                'general': ['La linea no tiene 4 columnas'],
                            }
                        }
                        errores_datos.append(error_dato)
                        errores = True
            if errores == False:
                for detalle in data_documento_detalle:
                    documento_detalle = GenDocumentoDetalle.objects.create(
                        documento=documento,
                        item_id=detalle['item_id'],
                        almacen_id=detalle['almacen_id'],
                        cantidad=detalle['cantidad'],
                        precio=detalle['precio']
                    )

                    item_impuestos = GenItemImpuesto.objects.filter(item_id=detalle['item_id'])

                    for item_impuesto in item_impuestos:
                        impuesto = GenImpuesto.objects.filter(
                            id=item_impuesto.impuesto_id,
                            venta=True
                        ).first()
                        
                        if impuesto:

                            base = documento_detalle.cantidad * documento_detalle.precio
                            total_impuesto = base * impuesto.porcentaje
                            total_operado = total_impuesto * impuesto.operacion

                            GenDocumentoImpuesto.objects.create(
                                documento_detalle=documento_detalle,
                                impuesto_id=impuesto.id,
                                base= base,
                                porcentaje=impuesto.porcentaje,
                                total=total_impuesto,
                                total_operado=total_operado
                            )
                registros_importados += 1

                return Response({'registros_importados': registros_importados}, status=status.HTTP_200_OK)
            else:
                return Response({'mensaje':'Errores de validacion', 'codigo':1, 'errores_validador': errores_datos}, status=status.HTTP_400_BAD_REQUEST)         
        else:
            return Response({'mensaje':'Faltan parametros', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)
        
    @action(detail=False, methods=["post"], url_path=r'importar_detalle_compra',)
    def importar_detalle_compra(self, request):
        raw = request.data
        documento_id = raw.get('documento_id')
        archivo_base64 = raw.get('archivo_base64')
        if documento_id and archivo_base64:
            try:
                documento = GenDocumento.objects.get(pk=documento_id)
            except GenDocumento.DoesNotExist:
                return Response({'mensaje':'El documento no existe', 'codigo':15}, status=status.HTTP_400_BAD_REQUEST)
            try:
                archivo_data = base64.b64decode(archivo_base64)
                archivo = BytesIO(archivo_data)
                wb = openpyxl.load_workbook(archivo)
                sheet = wb.active         
            except Exception as e:     
                return Response({'mensaje':'Error procesando el archivo', 'codigo':15}, status=status.HTTP_400_BAD_REQUEST)  
            
            data_documento_detalle = []
            errores = False
            errores_datos = []
            registros_importados = 0
            if documento.documento_tipo_id in [5]:
                for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    if len(row) == 5:
                        data = {
                            'item': row[0],
                            'almacen':row[1],
                            'grupo':row[2],
                            'cantidad': row[3] if row[3] is not None else 0,
                            'precio': row[4] if row[4] is not None else 0,              
                        }                    
                        if not data['item']:
                            errores = True
                            error_dato = {
                                'fila': i,
                                'errores': {
                                    'item': ['Debe digitar el código del item']
                                }
                            }
                            errores_datos.append(error_dato)
                        else:
                            item = GenItem.objects.filter(id=data['item']).first()
                            if item is None:
                                error_dato = {
                                    'fila': i,
                                    'errores': {
                                        'item': [f'El item {data["item"]} no existe']
                                    }
                                }
                                errores_datos.append(error_dato)
                            else:
                                data['item_id'] = item.id

                        if not data['almacen']:
                            error_dato = {
                                'fila': i,
                                'errores': {
                                    'almacen': 'Debe digitar el código del almacen',
                                }
                            }
                            errores_datos.append(error_dato)
                            errores = True  
                        else:
                            almacen = InvAlmacen.objects.filter(id=data['almacen']).first()
                            if almacen is None:
                                errores = True
                                error_dato = {
                                    'fila': i,
                                    'errores': {
                                        'almacen': [f'El almacen con código {data["almacen"]} no existe']
                                    }
                                }
                                errores_datos.append(error_dato)
                            else:
                                data['almacen_id'] = almacen.id
                                
                        if data['grupo']:
                            grupo = ConGrupo.objects.filter(id=data['grupo']).first()
                            if grupo is None:
                                errores = True
                                error_dato = {
                                    'fila': i,
                                    'errores': {
                                        'grupo': [f'El grupo con código {data["grupo"]} no existe']
                                    }
                                }
                                errores_datos.append(error_dato)
                            else:
                                data['grupo_id'] = grupo.id

                        data_documento_detalle.append(data) 
                    else:
                        error_dato = {
                            'fila': i,
                            'errores': {
                                'general': ['La linea no tiene 5 columnas'],
                            }
                        }
                        errores_datos.append(error_dato)
                        errores = True
            if errores == False:
                for detalle in data_documento_detalle:
                    documento_detalle = GenDocumentoDetalle.objects.create(
                        documento=documento,
                        item_id=detalle['item_id'],
                        grupo_id=detalle.get('grupo_id'),
                        almacen_id=detalle['almacen_id'],
                        cantidad=detalle['cantidad'],
                        precio=detalle['precio']
                    )
                    item_impuestos = GenItemImpuesto.objects.filter(item_id=detalle['item_id'])

                    for item_impuesto in item_impuestos:
                        impuesto = GenImpuesto.objects.filter(
                            id=item_impuesto.impuesto_id,
                            compra=True
                        ).first()
                        
                        if impuesto:

                            base = documento_detalle.cantidad * documento_detalle.precio
                            total_impuesto = base * impuesto.porcentaje
                            total_operado = total_impuesto * impuesto.operacion

                            GenDocumentoImpuesto.objects.create(
                                documento_detalle=documento_detalle,
                                impuesto_id=impuesto.id,
                                base= base,
                                porcentaje=impuesto.porcentaje,
                                total=total_impuesto,
                                total_operado=total_operado
                            )
                registros_importados += 1
                return Response({'registros_importados': registros_importados}, status=status.HTTP_200_OK)
            else:
                return Response({'mensaje':'Errores de validacion', 'codigo':1, 'errores_validador': errores_datos}, status=status.HTTP_400_BAD_REQUEST)      
        else:
            return Response({'mensaje':'Faltan parametros', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)
        
    @action(detail=False, methods=["post"], url_path=r'eliminar-todos',)
    def eliminar_todos_action(self, request):
        raw = request.data
        documento_id = raw.get('documento_id')
        if documento_id:
            documento_detalles = GenDocumentoDetalle.objects.filter(documento_id=documento_id)
            if documento_detalles.exists():
                documento_detalles.delete()
            return Response({'mensaje':'Todos los detalles han sido eliminados', 'codigo':0}, status=status.HTTP_200_OK)
        else:
            return Response({'mensaje':'Faltan parametros', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)        
                