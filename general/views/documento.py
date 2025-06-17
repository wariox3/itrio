from rest_framework import viewsets, permissions, status
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework.filters import OrderingFilter
from django_filters.rest_framework import DjangoFilterBackend
from general.models.documento import GenDocumento
from general.models.documento_detalle import GenDocumentoDetalle
from general.models.documento_impuesto import GenDocumentoImpuesto
from general.models.documento_tipo import GenDocumentoTipo
from general.models.documento_pago import GenDocumentoPago
from general.models.empresa import GenEmpresa
from general.models.contacto import GenContacto
from general.models.resolucion import GenResolucion
from general.models.configuracion import GenConfiguracion
from humano.models.contrato import HumContrato
from humano.models.concepto_cuenta import HumConceptoCuenta
from humano.models.configuracion_provision import HumConfiguracionProvision
from humano.models.configuracion_aporte import HumConfiguracionAporte
from contabilidad.models.cuenta import ConCuenta
from contabilidad.models.periodo import ConPeriodo
from contabilidad.models.movimiento import ConMovimiento
from contabilidad.models.activo import ConActivo
from general.models.item import GenItem
from inventario.models.existencia import InvExistencia
from inventario.models.almacen import InvAlmacen
from general.serializers.documento import GenDocumentoSerializador, GenDocumentoListaSerializador, GenDocumentoListaNominaSerializador, GenDocumentoInformeSerializador, GenDocumentoRetrieveSerializador
from general.serializers.documento_detalle import GenDocumentoDetalleSerializador
from general.serializers.documento_impuesto import GenDocumentoImpuestoSerializador
from general.serializers.documento_pago import GenDocumentoPagoSerializador
from general.serializers.contacto import GenContactoSerializador
from contabilidad.serializers.movimiento import ConMovimientoSerializador
from general.formatos.factura import FormatoFactura
from general.formatos.cuenta_cobro import FormatoCuentaCobro
from general.formatos.documento_soporte import FormatoDocumentoSoporte
from general.formatos.nomina import FormatoNomina
from general.formatos.recibo_caja import FormatoRecibo
from general.formatos.egreso import FormatoEgreso
from general.formatos.compra import FormatoCompra
from general.formatos.factura_pos import FormatoFacturaPOS
from general.filters.documento import DocumentoFilter
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from django.db.models import Sum, F, Count
from django.db.models.functions import TruncDay
from django.utils import timezone
from django.db.models.functions import Coalesce, Cast
from django.db.models import Sum, Q, DecimalField, CharField
from django.db import transaction
from utilidades.wolframio import Wolframio
from utilidades.zinc import Zinc
from utilidades.utilidades import Utilidades
from utilidades.exportar_excel import ExportarExcel
from decimal import Decimal
from datetime import datetime, timedelta, date
from io import BytesIO
import base64
import openpyxl
import calendar
import gc
import zipfile
import xml.etree.ElementTree as ET
import io
import re

def transformar_decimal(obj):
    if isinstance(obj, dict):
        for key, value in obj.items():
            obj[key] = transformar_decimal(value)
    elif isinstance(obj, list):
        for index, value in enumerate(obj):
            obj[index] = transformar_decimal(value)
    elif isinstance(obj, Decimal):
        return str(obj)
    return obj

class DocumentoViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_class = DocumentoFilter 
    queryset = GenDocumento.objects.all()   
    serializadores = {
        'lista': GenDocumentoListaSerializador,
        'lista_nomina': GenDocumentoListaNominaSerializador,
        'informe': GenDocumentoInformeSerializador,
        'nomina': GenDocumentoSerializador,
    }

    def get_serializer_class(self):
        serializador_parametro = self.request.query_params.get('serializador', None)
        if not serializador_parametro or serializador_parametro not in self.serializadores:
            return GenDocumentoSerializador
        return self.serializadores[serializador_parametro]

    def get_queryset(self):
        queryset = super().get_queryset()
        serializer_class = self.get_serializer_class()        
        select_related = getattr(serializer_class.Meta, 'select_related_fields', [])
        if select_related:
            queryset = queryset.select_related(*select_related)        
        campos = serializer_class.Meta.fields        
        if campos and campos != '__all__':
            queryset = queryset.only(*campos) 
        return queryset 

    def list(self, request, *args, **kwargs):
        if request.query_params.get('excel'):
            queryset = self.filter_queryset(self.get_queryset())
            serializer = self.get_serializer(queryset, many=True)
            exporter = ExportarExcel(serializer.data, sheet_name="documentos", filename="documentos.xlsx")
            return exporter.export()
        return super().list(request, *args, **kwargs)

    def retrieve(self, request, pk=None):
        queryset = GenDocumento.objects.all()
        documento = get_object_or_404(queryset, pk=pk)
        documentoSerializador = GenDocumentoRetrieveSerializador(documento)
        documentoDetalles = GenDocumentoDetalle.objects.filter(documento=pk).order_by('id')
        documentoDetallesSerializador = GenDocumentoDetalleSerializador(documentoDetalles, many=True)
        detalles = documentoDetallesSerializador.data
        for detalle in detalles:
            documentoImpuestos = GenDocumentoImpuesto.objects.filter(documento_detalle=detalle['id'])
            documentoImpuestosSerializador = GenDocumentoImpuestoSerializador(documentoImpuestos, many=True)
            detalle['impuestos'] = documentoImpuestosSerializador.data
        documentoRespuesta = documentoSerializador.data
        documentoRespuesta['detalles'] = detalles
        
        documentoPagos = GenDocumentoPago.objects.filter(documento=pk)
        documentoPagosSerializador = GenDocumentoPagoSerializador(documentoPagos, many=True)
        pagos = documentoPagosSerializador.data
        documentoRespuesta['pagos'] = pagos        
        
        return Response({'documento':documentoRespuesta}, status=status.HTTP_200_OK)
        
    def destroy(self, request, *args, **kwargs):        
        instance = self.get_object()
        if instance.estado_aprobado:
                return Response({'mensaje': 'No se puede eliminar un documento aprobado.'}, status=status.HTTP_400_BAD_REQUEST)
        if instance.documento_tipo_id == 15:
            nominas = GenDocumento.objects.filter(documento_referencia_id=instance.id)
            nominas.update(documento_referencia_id=None)
        self.perform_destroy(instance)
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=["post"], url_path=r'nuevo',)
    def nuevo(self, request):
        raw = request.data
        documentoSerializador = GenDocumentoSerializador(data=raw)
        if documentoSerializador.is_valid():  
            with transaction.atomic():          
                documento_tipo = documentoSerializador.validated_data['documento_tipo']
                documentoSerializador.validated_data['fecha_contable'] = documentoSerializador.validated_data['fecha']
                resolucion = documento_tipo.resolucion        
                documento = documentoSerializador.save(resolucion=resolucion)                        
                documentoRespuesta = documentoSerializador.data 
                detalles = raw.get('detalles')            
                pagos = raw.get('pagos')
                if detalles is not None:
                    for detalle in detalles:                
                        detalle['documento'] = documento.id                    
                        if detalle['tipo_registro'] == "I":
                            item = GenItem.objects.get(pk=detalle['item'])
                            if item:
                                if item.inventario:
                                    detalle['operacion_inventario'] = documento_tipo.operacion_inventario
                                    detalle['cantidad_operada'] = detalle['cantidad'] * documento_tipo.operacion_inventario
                        detalleSerializador = GenDocumentoDetalleSerializador(data=detalle)
                        if detalleSerializador.is_valid():                        
                            documentoDetalle = detalleSerializador.save() 
                            impuestos = detalle.get('impuestos')
                            if impuestos is not None:
                                for impuesto in impuestos:
                                    datosDocumentoImpuesto = {
                                        "documento_detalle":documentoDetalle.id, 
                                        "impuesto":impuesto['impuesto'],
                                        "base":impuesto['base'],
                                        "porcentaje":impuesto['porcentaje'],
                                        "total":impuesto['total'],
                                        "total_operado":impuesto['total_operado'],
                                        "porcentaje_base":impuesto['porcentaje_base']
                                    }
                                    documentoImpuestoSerializador = GenDocumentoImpuestoSerializador(data=datosDocumentoImpuesto)
                                    if documentoImpuestoSerializador.is_valid():
                                        documentoImpuestoSerializador.save()
                                    else:
                                        return Response({'mensaje':'Errores de validacion detalle impuesto', 'codigo':14, 'validaciones': documentoImpuestoSerializador.errors}, status=status.HTTP_400_BAD_REQUEST)                                                
                        else:
                            return Response({'mensaje':'Errores de validacion detalle', 'codigo':14, 'validaciones': detalleSerializador.errors}, status=status.HTTP_400_BAD_REQUEST)            
                    documentoDetalles = GenDocumentoDetalle.objects.filter(documento=documento.id)
                    documentoDetallesSerializador = GenDocumentoDetalleSerializador(documentoDetalles, many=True)
                    detalles = documentoDetallesSerializador.data
                    for detalle in detalles:
                        documentoImpuestos = GenDocumentoImpuesto.objects.filter(documento_detalle=detalle['id'])
                        documentoImpuestosSerializador = GenDocumentoImpuestoSerializador(documentoImpuestos, many=True)
                        detalle['impuestos'] = documentoImpuestosSerializador.data
                    documentoRespuesta['detalles'] = detalles   
                if pagos is not None:
                    for pago in pagos:                
                        pago['documento'] = documento.id
                        pagoSerializador = GenDocumentoPagoSerializador(data=pago)
                        if pagoSerializador.is_valid():
                            documentoPago = pagoSerializador.save() 
                        else:
                            return Response({'mensaje':'Errores de validacion pago', 'codigo':14, 'validaciones': pagoSerializador.errors}, status=status.HTTP_400_BAD_REQUEST)            
                    documentoRespuesta['pagos'] = pagos                                         
                return Response({'documento': documentoRespuesta}, status=status.HTTP_200_OK)
        return Response({'mensaje':'Errores de validacion', 'codigo':14, 'validaciones': documentoSerializador.errors}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["post"], url_path=r'actualizar',)
    def actualizar(self, request):
        raw = request.data        
        pk = raw.get('id')
        saltar_aprobado = raw.get('saltar_aprobado', False)
        resolucion_actualizar = raw.get('resolucion')
        try:
            with transaction.atomic():
                documento = GenDocumento.objects.get(pk=pk)        
                if documento.estado_aprobado == False or saltar_aprobado == True:
                    documentoSerializador = GenDocumentoSerializador(documento, data=raw, partial=True)
                    if documentoSerializador.is_valid():
                        documentoSerializador.validated_data['fecha_contable'] = documentoSerializador.validated_data['fecha']
                        documentoSerializador.save()
                        if resolucion_actualizar:
                            resolucion = GenResolucion.objects.get(pk=resolucion_actualizar)
                            documento = documentoSerializador.save(resolucion=resolucion)    
                        detalles = raw.get('detalles')
                        if detalles is not None:
                            for detalle in detalles:                
                                if detalle.get('id'):
                                    documentoDetalle = GenDocumentoDetalle.objects.get(pk=detalle['id'])
                                    if detalle['tipo_registro'] == "I":
                                        item = GenItem.objects.get(pk=detalle['item'])
                                        if item.inventario:
                                            detalle['operacion_inventario'] = documento.documento_tipo.operacion_inventario
                                            detalle['cantidad_operada'] = detalle['cantidad'] * documento.documento_tipo.operacion_inventario                              
                                    detalleSerializador = GenDocumentoDetalleSerializador(documentoDetalle, data=detalle, partial=True)    
                                else:
                                    detalle['documento'] = documento.id
                                    if detalle['tipo_registro'] == "I":
                                        item = GenItem.objects.get(pk=detalle['item'])
                                        if item.inventario:
                                            detalle['operacion_inventario'] = documento.documento_tipo.operacion_inventario
                                            detalle['cantidad_operada'] = detalle['cantidad'] * documento.documento_tipo.operacion_inventario
                                    detalleSerializador = GenDocumentoDetalleSerializador(data=detalle)
                                if detalleSerializador.is_valid():
                                    documentoDetalle = detalleSerializador.save() 
                                    impuestos = detalle.get('impuestos')
                                    if impuestos is not None:
                                        for impuesto in impuestos:
                                            if impuesto.get('id'):
                                                documentoImpuesto = GenDocumentoImpuesto.objects.get(pk=impuesto['id'])
                                                documentoImpuestoSerializador = GenDocumentoImpuestoSerializador(documentoImpuesto, data=impuesto, partial=True)    
                                            else:        
                                                impuesto['documento_detalle'] = documentoDetalle.id                                     
                                                documentoImpuestoSerializador = GenDocumentoImpuestoSerializador(data=impuesto)                            
                                            if documentoImpuestoSerializador.is_valid():
                                                documentoImpuestoSerializador.save()
                                            else:
                                                return Response({'mensaje':'Errores de validacion detalle impuesto', 'codigo':14, 'validaciones': documentoImpuestoSerializador.errors}, status=status.HTTP_400_BAD_REQUEST)                                                                        
                                    impuestosEliminados = detalle.get('impuestos_eliminados')
                                    if impuestosEliminados is not None:
                                        for documentoImpuesto in impuestosEliminados:                                
                                            documentoImpuesto = GenDocumentoImpuesto.objects.get(pk=documentoImpuesto)
                                            documentoImpuesto.delete()                         
                                else:
                                    return Response({'mensaje':'Errores de validacion detalle', 'codigo':14, 'validaciones': detalleSerializador.errors}, status=status.HTTP_400_BAD_REQUEST)            
                        detallesEliminados = raw.get('detalles_eliminados')
                        if detallesEliminados is not None:
                            for detalle in detallesEliminados:                                
                                documentoDetalle = GenDocumentoDetalle.objects.get(pk=detalle)
                                documentoDetalle.delete()
                        documentoDetalles = GenDocumentoDetalle.objects.filter(documento=pk)
                        documentoDetallesSerializador = GenDocumentoDetalleSerializador(documentoDetalles, many=True)
                        detalles = documentoDetallesSerializador.data
                        for detalle in detalles:
                            documentoImpuestos = GenDocumentoImpuesto.objects.filter(documento_detalle=detalle['id'])
                            documentoImpuestosSerializador = GenDocumentoImpuestoSerializador(documentoImpuestos, many=True)
                            detalle['impuestos'] = documentoImpuestosSerializador.data
                        documentoRespuesta = documentoSerializador.data
                        documentoRespuesta['detalles'] = detalles

                        pagos = raw.get('pagos')
                        if pagos is not None:
                            for pago in pagos:                
                                if pago.get('id'):
                                    documentoPago = GenDocumentoPago.objects.get(pk=pago['id'])
                                    documentoPagoSerializador = GenDocumentoPagoSerializador(documentoPago, data=pago, partial=True)    
                                else:
                                    pago['documento'] = documento.id
                                    documentoPagoSerializador = GenDocumentoPagoSerializador(data=pago)

                                if documentoPagoSerializador.is_valid():
                                    documentoPago = documentoPagoSerializador.save()                         
                                else:
                                    return Response({'mensaje':'Errores de validacion pago', 'codigo':14, 'validaciones': documentoPagoSerializador.errors}, status=status.HTTP_400_BAD_REQUEST)            
                        pagosEliminados = raw.get('pagos_eliminados')
                        if pagosEliminados is not None:
                            for pago in pagosEliminados:                                
                                documentoPago = GenDocumentoPago.objects.get(pk=pago)
                                documentoPago.delete()
                        documentoPagos = GenDocumentoPago.objects.filter(documento=pk)
                        documentoPagosSerializador = GenDocumentoPagoSerializador(documentoPagos, many=True)
                        pagos = documentoPagosSerializador.data
                        documentoRespuesta['pagos'] = pagos   
                        return Response({'documento': documentoRespuesta}, status=status.HTTP_200_OK)                    
                    return Response({'mensaje':'Errores de validacion', 'codigo':14, 'validaciones': documentoSerializador.errors}, status=status.HTTP_400_BAD_REQUEST)
                else:
                    return Response({'mensaje':'Los documentos aprobados no se pueden modificar', 'codigo':14}, status=status.HTTP_400_BAD_REQUEST)
        except GenDocumento.DoesNotExist:
            return Response({'mensaje': 'El documento no existe', 'codigo': 15}, status=status.HTTP_400_BAD_REQUEST)
        
    @action(detail=False, methods=["post"], url_path=r'eliminar',)
    def eliminar(self, request):
        try:
            raw = request.data
            documentos = raw.get('documentos')
            if documentos:  
                for documento in documentos:
                    documentoEliminar = GenDocumento.objects.get(pk=documento)  
                    if documentoEliminar:
                        if documentoEliminar.estado_aprobado == False or documentoEliminar.documento_tipo_id in (18,19):
                            if not documentoEliminar.documentos_detalles_documento_rel.exists():
                                if documentoEliminar.documento_tipo_id == 15:
                                    nominas = GenDocumento.objects.filter(documento_referencia_id=documentoEliminar.id)
                                    nominas.update(documento_referencia_id=None)                                
                                if not documentoEliminar.documentos_detalles_documento_afectado_rel.exists():
                                    documentoEliminar.delete()   
                                else:
                                    return Response({'mensaje':'El documento con id ' + str(documentoEliminar.id) + ' esta afectado con algunos detalles', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)                                                                        
                            else:                            
                                return Response({'mensaje':'El documento con id ' + str(documentoEliminar.id) + ' no se puede eliminar por que tiene detalles', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)                                    
                        else:
                            return Response({'mensaje':'El documento con id ' + str(documentoEliminar.id) + ' no se puede eliminar por que se encuentra aprobado', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)
                return Response({'mensaje':'Registros eliminados'}, status=status.HTTP_200_OK)
            return Response({'mensaje':'Faltan parametros', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)
        except GenDocumento.DoesNotExist:
            return Response({'mensaje':'El documento no existe', 'codigo':15}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["post"], url_path=r'aprobar',)
    def aprobar(self, request):        
        raw = request.data
        id = raw.get('id')
        if id:
            try:
                with transaction.atomic():
                    documento = GenDocumento.objects.get(pk=id)
                    consecutivo = 0
                    documento_tipo = GenDocumentoTipo.objects.get(id=documento.documento_tipo_id)
                    if documento.numero is None:
                        consecutivo = documento_tipo.consecutivo
                    else: 
                        consecutivo = documento.numero
                    documento_detalles = GenDocumentoDetalle.objects.filter(documento_id=id)
                    respuesta = self.validacion_aprobar(documento, documento_detalles, consecutivo)
                    if respuesta['error'] == False:                         
                        if documento.numero is None:
                            documento.numero = documento_tipo.consecutivo
                            documento_tipo.consecutivo += 1
                            documento_tipo.save()                
                        documento.estado_aprobado = True
                        if documento.documento_tipo.documento_clase_id in (100,101,102,104,105,300,301,302,303,304):
                            if documento.documento_referencia:
                                documento.pendiente = 0
                            else:
                                documento.pendiente = documento.total - documento.afectado
                            
                        # Compra, Documento soporte
                        if documento.documento_tipo_id in [5,11]:
                            if documento.forma_pago:
                                documento.cuenta_id = documento.forma_pago.cuenta_id

                        # Procesar detalles
                        for documento_detalle in documento_detalles:
                            if documento_detalle.documento_afectado:
                                if documento.documento_tipo.documento_clase_id in (200,400):
                                    documento_afectado = documento_detalle.documento_afectado                        
                                    documento_afectado.afectado += documento_detalle.precio
                                    documento_afectado.pendiente = documento_afectado.total - documento_afectado.afectado
                                    documento_afectado.save(update_fields=['afectado', 'pendiente'])     
                                    
                            if documento_detalle.operacion_inventario != 0:
                                existencia = InvExistencia.objects.filter(almacen_id=documento_detalle.almacen_id, item_id=documento_detalle.item_id).first()
                                if existencia:
                                    existencia.existencia += documento_detalle.cantidad_operada
                                    existencia.disponible += documento_detalle.cantidad_operada
                                    existencia.save(update_fields=['existencia', 'disponible'])
                                else:
                                    existencia = InvExistencia.objects.create(almacen_id=documento_detalle.almacen_id, item_id=documento_detalle.item_id, existencia=documento_detalle.cantidad_operada, disponible=documento_detalle.cantidad_operada)
                                item = GenItem.objects.get(id=documento_detalle.item_id)
                                item.existencia += documento_detalle.cantidad_operada
                                item.disponible += documento_detalle.cantidad_operada
                                item.save(update_fields=['existencia', 'disponible'])

                            if documento_detalle.tipo_registro == 'D':
                                activo = ConActivo.objects.get(id=documento_detalle.activo_id)
                                if activo:
                                    depreciacion_acumulada = activo.depreciacion_acumulada + documento_detalle.precio
                                    depreciacion_saldo = activo.depreciacion_saldo - documento_detalle.precio
                                    activo.depreciacion_acumulada = depreciacion_acumulada
                                    activo.depreciacion_saldo = depreciacion_saldo
                                    activo.save()
                        
                        # Nota credito
                        if documento.documento_referencia:
                            documento_referencia = documento.documento_referencia
                            documento_referencia.afectado += documento.total
                            documento_referencia.pendiente = documento_referencia.total - documento_referencia.afectado
                            documento_referencia.save(update_fields=['afectado', 'pendiente'])  

                        documento.save()
                        if documento.documento_tipo.electronico:
                            configuracion = GenConfiguracion.objects.filter(pk=1).values('gen_emitir_automaticamente')[0] 
                            if configuracion['gen_emitir_automaticamente']:
                                self.emitir(id)

                        return Response({'estado_aprobado': True}, status=status.HTTP_200_OK)
                    else:
                        return Response({'mensaje':respuesta['mensaje'], 'codigo':1}, status=status.HTTP_400_BAD_REQUEST) 
            except GenDocumento.DoesNotExist:
                return Response({'mensaje':'El documento no existe', 'codigo':15}, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'mensaje':'Faltan parametros', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["post"], url_path=r'desaprobar',)
    def desaprobar(self, request):        
        raw = request.data
        id = raw.get('id')
        if id:
            try:
                with transaction.atomic():
                    documento = GenDocumento.objects.get(pk=id)
                    if documento.estado_aprobado == True and documento.estado_anulado == False and documento.estado_contabilizado == False and documento.estado_electronico_enviado == False:                                  
                        if documento.documento_tipo.documento_clase_id in (100, 101, 200,300, 303,400,500,501,601,603):                    
                            respuesta = self.validacion_desaprobar(documento)
                            if respuesta['error'] == False:
                                documento.estado_aprobado = False
                                if documento.documento_tipo.documento_clase_id in (100,101,102,104,300,301,302,303):
                                    documento.pendiente = 0                           

                                documento_detalles = GenDocumentoDetalle.objects.filter(documento_id=id)
                                for documento_detalle in documento_detalles:
                                    if documento_detalle.documento_afectado:
                                        if documento.documento_tipo.documento_clase_id in (200,400):
                                            documento_afectado = documento_detalle.documento_afectado                        
                                            documento_afectado.afectado -= documento_detalle.precio
                                            documento_afectado.pendiente = documento_afectado.total - documento_afectado.afectado
                                            documento_afectado.save(update_fields=['afectado', 'pendiente'])
                                                                                                    
                                    if documento_detalle.operacion_inventario != 0:
                                        existencia = InvExistencia.objects.filter(almacen_id=documento_detalle.almacen_id, item_id=documento_detalle.item_id).first()
                                        if existencia:
                                            existencia.existencia -= documento_detalle.cantidad_operada
                                            existencia.disponible -= documento_detalle.cantidad_operada
                                            existencia.save(update_fields=['existencia', 'disponible'])
                                        item = GenItem.objects.get(id=documento_detalle.item_id)
                                        item.existencia -= documento_detalle.cantidad_operada
                                        item.disponible -= documento_detalle.cantidad_operada
                                        item.save(update_fields=['existencia', 'disponible'])  

                                # Nota credito
                                if documento.documento_referencia:
                                    documento_referencia = documento.documento_referencia
                                    documento_referencia.afectado -= documento.total
                                    documento_referencia.pendiente = documento_referencia.total - documento_referencia.afectado
                                    documento_referencia.save(update_fields=['afectado', 'pendiente'])                                                                               
                                documento.save()
                                return Response({'estado_aprobado': False}, status=status.HTTP_200_OK)
                            else:
                                return Response({'mensaje':respuesta['mensaje'], 'codigo':1}, status=status.HTTP_400_BAD_REQUEST) 
                        else:
                            return Response({'mensaje':'El documento no permite desaprobacion', 'codigo':15}, status=status.HTTP_400_BAD_REQUEST)                      
                    else:
                        return Response({'mensaje':'El documento debe estar aprobado, sin anular, sin contabilizar, sin emitir electricamente para permitir la accion', 'codigo':15}, status=status.HTTP_400_BAD_REQUEST)                                                                
            except GenDocumento.DoesNotExist:
                return Response({'mensaje':'El documento no existe', 'codigo':15}, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'mensaje':'Faltan parametros', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["post"], url_path=r'contabilizar',)
    def contabilizar(self, request):        
        raw = request.data
        ids = raw.get('ids')
        if ids:
            cantidad = 0
            for id in ids:
                try:
                    documento = GenDocumento.objects.get(pk=id)                            
                except GenDocumento.DoesNotExist:
                    return Response({'mensaje':'El documento no existe', 'codigo':15}, status=status.HTTP_400_BAD_REQUEST)
                            
                if documento.estado_aprobado == False:   
                    return Response({'mensaje':'El documento debe estar aprobado', 'codigo':15}, status=status.HTTP_400_BAD_REQUEST)
                                
                if documento.estado_contabilizado == True:
                    return Response({'mensaje':'El documento ya esta contabilizado', 'codigo':15}, status=status.HTTP_400_BAD_REQUEST)
                
                try:
                    periodo_id = documento.fecha_contable.strftime("%Y%m")
                    if documento.documento_tipo_id == 25:
                        periodo_id = documento.fecha.strftime("%Y") + '13'
                    periodo = ConPeriodo.objects.get(pk=periodo_id)
                    if periodo.estado_bloqueado == True:
                        return Response({'mensaje': f'El periodo {periodo_id} esta bloqueado y no es posible contabilizar el documento', 'codigo':15}, status=status.HTTP_400_BAD_REQUEST)
                except ConPeriodo.DoesNotExist:
                    return Response({'mensaje':f'El periodo contable {periodo_id} no existe', 'codigo':15}, status=status.HTTP_400_BAD_REQUEST)
                
                movimientos_validos = []
                if documento.estado_anulado == False:
                    movimientos_validos = []    
                    if documento.documento_tipo_id == 13:
                        comprobante_id = documento.comprobante_id
                    else:
                        comprobante_id = documento.documento_tipo.comprobante_id                    
                    data_general = {
                        'documento': id,
                        'periodo': periodo_id,
                        'numero': documento.numero,
                        'fecha': documento.fecha_contable,
                        'comprobante': comprobante_id
                    }
                            
                    if documento.documento_tipo.cobrar:
                        data = data_general.copy()                                                                        
                        data['cuenta'] = documento.documento_tipo.cuenta_cobrar_id
                        #Se actualiza la cuenta en el documento para cuando se haga el recibo/pago quede a esta cuenta
                        documento.cuenta_id = documento.documento_tipo.cuenta_cobrar_id
                        data['contacto'] = documento.contacto_id        
                        if documento.documento_tipo_id in [1,3, 17, 24]:
                            data['naturaleza'] = 'D'
                            data['debito'] = documento.total
                        else: 
                            data['naturaleza'] = 'C'
                            data['credito'] = documento.total
                        data['detalle'] = 'CLIENTE'
                        if documento.documento_tipo.cuenta_cobrar:
                            if documento.documento_tipo.cuenta_cobrar.exige_grupo:
                                if documento.sede:
                                    data['grupo'] = documento.sede.grupo_id                        
                        movimiento_serializador = ConMovimientoSerializador(data=data)
                        if movimiento_serializador.is_valid():                
                            movimientos_validos.append(movimiento_serializador)
                        else:
                            return Response({'validaciones': movimiento_serializador.errors, 
                                                'mensaje': 'Cuenta por cobrar'}, status=status.HTTP_400_BAD_REQUEST)
                
                    if documento.documento_tipo.pagar:
                        data = data_general.copy()
                        #Se actualiza la cuenta en el documento para cuando se haga el egreso quede a esta cuenta                                                
                        data['cuenta'] = documento.documento_tipo.cuenta_pagar_id
                        documento.cuenta_id = documento.documento_tipo.cuenta_pagar_id                         
                        if documento.forma_pago:
                            data['cuenta'] = documento.forma_pago.cuenta_id                                                   
                            documento.cuenta_id = documento.forma_pago.cuenta_id 
                        if documento.documento_tipo_id == 22:
                            clase = documento.orden_compra
                            configuracion_aporte = HumConfiguracionAporte.objects.filter(tipo=clase).first()
                            if configuracion_aporte:
                                data['cuenta'] = configuracion_aporte.cuenta_id                                                   
                                documento.cuenta_id = configuracion_aporte.cuenta_id
                        data['contacto'] = documento.contacto_id        
                        if documento.documento_tipo_id in [6, 12]:
                            data['naturaleza'] = 'D'
                            data['debito'] = documento.total
                        else:
                            data['naturaleza'] = 'C'
                            data['credito'] = documento.total
                        if documento.documento_tipo_id == 22:
                            data['credito'] = documento.subtotal
                        data['detalle'] = 'PROVEEDOR'
                        movimiento_serializador = ConMovimientoSerializador(data=data)
                        if movimiento_serializador.is_valid():
                            movimientos_validos.append(movimiento_serializador)
                        else:
                            return Response({'validaciones': movimiento_serializador.errors, 
                                                'mensaje': 'Cuenta por pagar'}, status=status.HTTP_400_BAD_REQUEST)
                                                
                    documento_detalles = GenDocumentoDetalle.objects.filter(documento_id=id)
                    for documento_detalle in documento_detalles:
                        if documento_detalle.documento_afectado:
                            if documento_detalle.documento_afectado.documento_tipo.cobrar:
                                data = data_general.copy()                            
                                data['cuenta'] = documento_detalle.documento_afectado.cuenta_id
                                data['contacto'] = documento_detalle.contacto_id     
                                if documento_detalle.documento_afectado.documento_tipo_id in [2]:
                                    data['naturaleza'] = 'D'
                                    data['debito'] = documento_detalle.precio
                                else:
                                    data['naturaleza'] = 'C'
                                    data['credito'] = documento_detalle.precio
                                data['detalle'] = 'CLIENTE'
                                movimiento_serializador = ConMovimientoSerializador(data=data)
                                if movimiento_serializador.is_valid():
                                    movimientos_validos.append(movimiento_serializador)
                                else:
                                    return Response({'validaciones': movimiento_serializador.errors, 
                                                'mensaje': f'Detalle {documento_detalle.id}: cuenta por cobrar documento referencia'}, status=status.HTTP_400_BAD_REQUEST)
                                
                            if documento_detalle.documento_afectado.documento_tipo.pagar:
                                data = data_general.copy()                            
                                data['cuenta'] = documento_detalle.documento_afectado.cuenta_id
                                data['contacto'] = documento_detalle.contacto_id        
                                if documento_detalle.documento_afectado.documento_tipo_id in [6, 12]:
                                    data['naturaleza'] = 'C'
                                    data['credito'] = documento_detalle.precio
                                else:
                                    data['naturaleza'] = 'D'
                                    data['debito'] = documento_detalle.precio
                                data['detalle'] = 'PROVEEDOR'
                                movimiento_serializador = ConMovimientoSerializador(data=data)
                                if movimiento_serializador.is_valid():
                                    movimientos_validos.append(movimiento_serializador)
                                    documento_detalle.cuenta_id = data['cuenta']
                                    documento_detalle.naturaleza = data['naturaleza']
                                    documento_detalle.save()
                                else:
                                    return Response({'validaciones': movimiento_serializador.errors, 
                                                'mensaje': f'Detalle {documento_detalle.id}: cuenta por pagar documento referencia'}, status=status.HTTP_400_BAD_REQUEST)                                                                
                        else:
                            # Item
                            if documento_detalle.tipo_registro == 'I':                                                                
                                if documento.documento_tipo.venta:
                                    data = data_general.copy()                            
                                    data['cuenta'] = documento_detalle.item.cuenta_venta_id
                                    data['contacto'] = documento.contacto_id                                
                                    if documento.documento_tipo_id in [1,3, 17, 24]:
                                        data['naturaleza'] = 'C'
                                        data['credito'] = documento_detalle.subtotal
                                    else:
                                        data['naturaleza'] = 'D'
                                        data['debito'] = documento_detalle.subtotal
                                    data['detalle'] = 'VENTA'
                                    if documento_detalle.item.cuenta_venta:
                                        if documento_detalle.item.cuenta_venta.exige_grupo:
                                            if documento.sede:
                                                data['grupo'] = documento.sede.grupo_id
                                    movimiento_serializador = ConMovimientoSerializador(data=data)
                                    if movimiento_serializador.is_valid():
                                        movimientos_validos.append(movimiento_serializador)
                                    else:
                                        return Response({'validaciones': movimiento_serializador.errors, 
                                                'mensaje': 'Item cuenta de venta'}, status=status.HTTP_400_BAD_REQUEST) 
                                    
                                if documento.documento_tipo.compra:
                                    data = data_general.copy()                            
                                    data['cuenta'] = documento_detalle.item.cuenta_compra_id
                                    data['contacto'] = documento.contacto_id     
                                    if documento.documento_tipo_id in [6, 12]:
                                        data['naturaleza'] = 'C'
                                        data['credito'] = documento_detalle.subtotal
                                    else:
                                        data['naturaleza'] = 'D'
                                        data['debito'] = documento_detalle.subtotal   
                                    data['detalle'] = 'ITEM COMPRA'
                                    if documento_detalle.item.cuenta_compra:
                                        if documento_detalle.item.cuenta_compra.exige_grupo:
                                            data['grupo'] = documento_detalle.grupo_id
                                    movimiento_serializador = ConMovimientoSerializador(data=data)
                                    if movimiento_serializador.is_valid():
                                        movimientos_validos.append(movimiento_serializador)
                                    else:
                                        return Response({'validaciones': movimiento_serializador.errors, 
                                                'mensaje': 'Item cuenta compra'}, status=status.HTTP_400_BAD_REQUEST)  

                            # Cuenta                                                                            
                            if documento_detalle.tipo_registro == 'C':                                    
                                data = data_general.copy()                            
                                data['cuenta'] = documento_detalle.cuenta_id
                                if documento_detalle.cuenta:
                                    if documento_detalle.cuenta.exige_contacto:
                                        data['contacto'] = documento_detalle.contacto_id
                                    else:
                                        data['contacto'] = None
                                data['naturaleza'] = documento_detalle.naturaleza
                                data['base'] = documento_detalle.base
                                data['detalle'] = documento_detalle.detalle
                                if documento_detalle.naturaleza == 'D':
                                    data['debito'] = documento_detalle.precio
                                if documento_detalle.naturaleza == 'C':
                                    data['credito'] = documento_detalle.precio
                                if documento_detalle.cuenta:
                                    if documento_detalle.cuenta.exige_grupo:
                                        data['grupo'] = documento_detalle.grupo_id 
                                if documento.documento_tipo_id == 25:
                                    data['cierre'] = True

                                movimiento_serializador = ConMovimientoSerializador(data=data)
                                if movimiento_serializador.is_valid():
                                    movimientos_validos.append(movimiento_serializador)
                                else:
                                    return Response({'validaciones': movimiento_serializador.errors, 'mensaje': f'Detalle {documento_detalle.id}'}, status=status.HTTP_400_BAD_REQUEST) 
                    
                            # Nomina
                            if documento_detalle.tipo_registro == 'N':
                                tipo_costo_id = documento.contrato.tipo_costo_id
                                if documento_detalle.operacion != 0:
                                    data = data_general.copy()                                       
                                    concepto_cuenta = HumConceptoCuenta.objects.filter(concepto_id=documento_detalle.concepto_id, tipo_costo_id=tipo_costo_id).first()                                                       
                                    if concepto_cuenta:
                                        data['cuenta'] = concepto_cuenta.cuenta_id                                                                        
                                        if concepto_cuenta.cuenta.exige_grupo:
                                            data['grupo'] = documento.contrato.grupo_id                                                                        
                                    data['contacto'] = documento_detalle.contacto_id        
                                    data['naturaleza'] = documento_detalle.naturaleza
                                    data['base'] = documento_detalle.base                                    
                                    data['detalle'] = documento_detalle.concepto.nombre
                                    if documento_detalle.concepto.concepto_tipo_id in [5,6]:                                        
                                        data['detalle'] = f'{documento.contacto.numero_identificacion}-{documento.contacto.nombre_corto}'
                                    if documento_detalle.operacion == 1:
                                        data['naturaleza'] = 'D'
                                        data['debito'] = documento_detalle.pago
                                    if documento_detalle.operacion == -1:
                                        data['naturaleza'] = 'C'                                    
                                        data['credito'] = documento_detalle.pago
                                    
                                    movimiento_serializador = ConMovimientoSerializador(data=data)
                                    if movimiento_serializador.is_valid():
                                        movimientos_validos.append(movimiento_serializador)
                                    else:
                                        return Response({'validaciones': movimiento_serializador.errors, 'mensaje': f'Documento detalle {documento_detalle.id} nomina'}, status=status.HTTP_400_BAD_REQUEST) 
                            
                            # Seguridad social
                            if documento_detalle.tipo_registro == 'S':                                
                                tipo_costo_id = documento_detalle.contrato.tipo_costo_id
                                data = data_general.copy()
                                clase = documento_detalle.detalle
                                configuracion_provision = HumConfiguracionProvision.objects.filter(tipo=clase, tipo_costo_id=tipo_costo_id).first()                                                       
                                if configuracion_provision:
                                    data['cuenta'] = configuracion_provision.cuenta_debito_id                                                                        
                                    if configuracion_provision.cuenta_debito:
                                        if configuracion_provision.cuenta_debito.exige_grupo:
                                            data['grupo'] = documento_detalle.contrato.grupo_id                                                                                                                                                          
                                data['contacto'] = documento.contacto_id
                                data['naturaleza'] = 'D'
                                data['debito'] = documento_detalle.precio
                                data['detalle'] = f'SS {clase} COSTO'
                                movimiento_serializador = ConMovimientoSerializador(data=data)
                                if movimiento_serializador.is_valid():
                                    movimientos_validos.append(movimiento_serializador)
                                else:
                                    return Response({'validaciones': movimiento_serializador.errors, 
                                                        'mensaje': 'Seguridad social pension'}, status=status.HTTP_400_BAD_REQUEST)                                                                                                                             

                            # Depreciacion
                            if documento_detalle.tipo_registro == 'D':                                    
                                data = data_general.copy()                            
                                data['cuenta'] = documento_detalle.activo.cuenta_depreciacion_id
                                data['contacto'] = documento_detalle.contacto_id                                        
                                data['base'] = documento_detalle.base
                                data['naturaleza'] = 'C'
                                data['credito'] = documento_detalle.precio                                
                                if documento_detalle.activo.cuenta_depreciacion:
                                    if documento_detalle.activo.cuenta_depreciacion.exige_grupo:
                                        data['grupo'] = documento.grupo_contabilidad_id    
                                data['detalle'] = 'Depreciacion acumulada'                         
                                movimiento_serializador = ConMovimientoSerializador(data=data)
                                if movimiento_serializador.is_valid():
                                    movimientos_validos.append(movimiento_serializador)
                                else:
                                    return Response({'validaciones': movimiento_serializador.errors, 'mensaje': f'Detalle {documento_detalle.id}'}, status=status.HTTP_400_BAD_REQUEST) 

                                data = data_general.copy()                            
                                data['cuenta'] = documento_detalle.activo.cuenta_gasto_id
                                data['contacto'] = documento_detalle.contacto_id                                        
                                data['base'] = documento_detalle.base
                                data['naturaleza'] = 'D'
                                data['debito'] = documento_detalle.precio                                
                                if documento_detalle.activo.cuenta_gasto:
                                    if documento_detalle.activo.cuenta_gasto.exige_grupo:
                                        data['grupo'] = documento.grupo_contabilidad_id   
                                data['detalle'] = 'Gasto depreciacion'                                                                     
                                movimiento_serializador = ConMovimientoSerializador(data=data)
                                if movimiento_serializador.is_valid():
                                    movimientos_validos.append(movimiento_serializador)
                                else:
                                    return Response({'validaciones': movimiento_serializador.errors, 'mensaje': f'Detalle {documento_detalle.id}'}, status=status.HTTP_400_BAD_REQUEST) 

                    documento_impuestos = GenDocumentoImpuesto.objects.filter(
                        documento_detalle__documento_id=id
                    ).values(
                        'impuesto_id',
                        'impuesto__cuenta_id',
                        'impuesto__compra',
                        'impuesto__venta',
                        'impuesto__operacion',
                        'impuesto__nombre',
                        'documento_detalle_id',
                        'documento_detalle__documento__documento_tipo_id'
                    ).annotate(
                        total=Coalesce(Sum('total'), 0, output_field=DecimalField()),
                        base=Coalesce(Sum('base'), 0, output_field=DecimalField()),
                    )
                    for documento_impuesto in documento_impuestos:
                        data = data_general.copy()                            
                        data['cuenta'] = documento_impuesto['impuesto__cuenta_id']
                        data['contacto'] = documento.contacto_id        
                        if documento_impuesto['impuesto__venta']:
                            if documento_impuesto['impuesto__operacion'] == 1:
                                if documento_impuesto['documento_detalle__documento__documento_tipo_id'] in [1,3,17,24]:
                                    data['naturaleza'] = 'C'
                                    data['credito'] = documento_impuesto['total']
                                else:
                                    data['naturaleza'] = 'D'
                                    data['debito'] = documento_impuesto['total']

                            if documento_impuesto['impuesto__operacion'] == -1:
                                if documento_impuesto['documento_detalle__documento__documento_tipo_id'] in [1,3,17,24]:
                                    data['naturaleza'] = 'D'
                                    data['debito'] = documento_impuesto['total']                    
                                else:
                                    data['naturaleza'] = 'C'
                                    data['credito'] = documento_impuesto['total']                    
                        
                        if documento_impuesto['impuesto__compra']:
                            if documento_impuesto['impuesto__operacion'] == 1:
                                if documento_impuesto['documento_detalle__documento__documento_tipo_id'] in [6,12]:
                                    data['naturaleza'] = 'C'
                                    data['credito'] = documento_impuesto['total']
                                else:
                                    data['naturaleza'] = 'D'
                                    data['debito'] = documento_impuesto['total']
                            if documento_impuesto['impuesto__operacion'] == -1:
                                if documento_impuesto['documento_detalle__documento__documento_tipo_id'] in [6,12]:
                                    data['naturaleza'] = 'D'
                                    data['debito'] = documento_impuesto['total']
                                else:
                                    data['naturaleza'] = 'C'
                                    data['credito'] = documento_impuesto['total']
                        data['base'] = documento_impuesto['base']
                        data['detalle'] = 'IMPUESTO'
                        movimiento_serializador = ConMovimientoSerializador(data=data)
                        if movimiento_serializador.is_valid():
                            movimientos_validos.append(movimiento_serializador)
                        else:
                            return Response({'validaciones': movimiento_serializador.errors, 
                                                'mensaje': f'En el detalle Id {documento_impuesto["documento_detalle_id"]}, el impuesto {documento_impuesto["impuesto__nombre"]} no tiene cuenta'}, status=status.HTTP_400_BAD_REQUEST)   
                                                    
                    # Pago y Egreso
                    if documento.documento_tipo_id in [4, 8]:
                        if documento.documento_tipo_id == 4:
                            data = data_general.copy()                            
                            data['cuenta'] = documento.cuenta_banco.cuenta_id                            
                            data['naturaleza'] = 'D'
                            data['debito'] = documento.total
                            data['detalle'] = 'CUENTA BANCO'
                            movimiento_serializador = ConMovimientoSerializador(data=data)
                            if movimiento_serializador.is_valid():
                                movimientos_validos.append(movimiento_serializador)
                            else:
                                return Response({'validaciones': movimiento_serializador.errors, 
                                                    'mensaje': 'Pago / Egreso cuenta banco'}, status=status.HTTP_400_BAD_REQUEST) 
                            
                        if documento.documento_tipo_id == 8:
                            data = data_general.copy()                            
                            data['cuenta'] = documento.cuenta_banco.cuenta_id                            
                            data['naturaleza'] = 'C'
                            data['credito'] = documento.total
                            data['detalle'] = 'CUENTA BANCO'
                            movimiento_serializador = ConMovimientoSerializador(data=data)
                            if movimiento_serializador.is_valid():
                                movimientos_validos.append(movimiento_serializador)
                            else:
                                return Response({'validaciones': movimiento_serializador.errors, 
                                                    'mensaje': 'Pago / Egreso cuenta banco'}, status=status.HTTP_400_BAD_REQUEST) 

                    # Nomina - Provisiones
                    if documento.documento_tipo_id == 14:
                        tipo_costo_id = documento.contrato.tipo_costo_id
                        if documento.provision_cesantia > 0:
                            data = data_general.copy()
                            configuracion_provision = HumConfiguracionProvision.objects.filter(tipo='CESANTIA', tipo_costo_id=tipo_costo_id).first()                                                       
                            if configuracion_provision:
                                data['cuenta'] = configuracion_provision.cuenta_debito_id                                                                        
                                if configuracion_provision.cuenta_debito:
                                    if configuracion_provision.cuenta_debito.exige_grupo:
                                        data['grupo'] = documento.contrato.grupo_id                                                                                                                                                          
                            data['contacto'] = documento.contacto_id
                            data['naturaleza'] = 'D'
                            data['debito'] = documento.provision_cesantia
                            data['detalle'] = 'PROVISION CESANTIA DEB'
                            movimiento_serializador = ConMovimientoSerializador(data=data)
                            if movimiento_serializador.is_valid():
                                movimientos_validos.append(movimiento_serializador)
                            else:
                                return Response({'validaciones': movimiento_serializador.errors, 
                                                    'mensaje': 'Provision cesantia'}, status=status.HTTP_400_BAD_REQUEST)
                            
                            data = data_general.copy()
                            if configuracion_provision:
                                data['cuenta'] = configuracion_provision.cuenta_credito_id                                                                        
                                if configuracion_provision.cuenta_credito:
                                    if configuracion_provision.cuenta_credito.exige_grupo:
                                        data['grupo'] = documento.contrato.grupo_id                                                                                                                                                          
                            data['contacto'] = documento.contacto_id
                            data['naturaleza'] = 'C'
                            data['credito'] = documento.provision_cesantia
                            data['detalle'] = 'PROVISION CESANTIA CRE'
                            movimiento_serializador = ConMovimientoSerializador(data=data)
                            if movimiento_serializador.is_valid():
                                movimientos_validos.append(movimiento_serializador)
                            else:
                                return Response({'validaciones': movimiento_serializador.errors, 
                                                    'mensaje': 'Provision cesantia'}, status=status.HTTP_400_BAD_REQUEST)                             
                            
                        if documento.provision_interes > 0:                            
                            data = data_general.copy()
                            configuracion_provision = HumConfiguracionProvision.objects.filter(tipo='INTERES', tipo_costo_id=tipo_costo_id).first()                                                       
                            if configuracion_provision:
                                data['cuenta'] = configuracion_provision.cuenta_debito_id                                                                        
                                if configuracion_provision.cuenta_debito:
                                    if configuracion_provision.cuenta_debito.exige_grupo:
                                        data['grupo'] = documento.contrato.grupo_id                                                                                                                                                          
                            data['contacto'] = documento.contacto_id
                            data['naturaleza'] = 'D'
                            data['debito'] = documento.provision_interes
                            data['detalle'] = 'PROVISION INTERES DEB'
                            movimiento_serializador = ConMovimientoSerializador(data=data)
                            if movimiento_serializador.is_valid():
                                movimientos_validos.append(movimiento_serializador)
                            else:
                                return Response({'validaciones': movimiento_serializador.errors, 
                                                    'mensaje': 'Provision interes'}, status=status.HTTP_400_BAD_REQUEST)                            
                            data = data_general.copy()
                            if configuracion_provision:
                                data['cuenta'] = configuracion_provision.cuenta_credito_id                                                                        
                                if configuracion_provision.cuenta_credito:
                                    if configuracion_provision.cuenta_credito.exige_grupo:
                                        data['grupo'] = documento.contrato.grupo_id                                                                                                                                                          
                            data['contacto'] = documento.contacto_id
                            data['naturaleza'] = 'C'
                            data['credito'] = documento.provision_interes
                            data['detalle'] = 'PROVISION INTERES CRE'
                            movimiento_serializador = ConMovimientoSerializador(data=data)
                            if movimiento_serializador.is_valid():
                                movimientos_validos.append(movimiento_serializador)
                            else:
                                return Response({'validaciones': movimiento_serializador.errors, 
                                                    'mensaje': 'Provision interes'}, status=status.HTTP_400_BAD_REQUEST)
                            
                        if documento.provision_prima > 0:
                            data = data_general.copy()
                            configuracion_provision = HumConfiguracionProvision.objects.filter(tipo='PRIMA', tipo_costo_id=tipo_costo_id).first()                                                       
                            if configuracion_provision:
                                data['cuenta'] = configuracion_provision.cuenta_debito_id                                                                        
                                if configuracion_provision.cuenta_debito:
                                    if configuracion_provision.cuenta_debito.exige_grupo:
                                        data['grupo'] = documento.contrato.grupo_id                                                                                                                                                          
                            data['contacto'] = documento.contacto_id
                            data['naturaleza'] = 'D'
                            data['debito'] = documento.provision_prima
                            data['detalle'] = 'PROVISION PRIMA DEB'
                            movimiento_serializador = ConMovimientoSerializador(data=data)
                            if movimiento_serializador.is_valid():
                                movimientos_validos.append(movimiento_serializador)
                            else:
                                return Response({'validaciones': movimiento_serializador.errors, 
                                                    'mensaje': 'Provision prima'}, status=status.HTTP_400_BAD_REQUEST) 
                            
                            data = data_general.copy()
                            if configuracion_provision:
                                data['cuenta'] = configuracion_provision.cuenta_credito_id                                                                        
                                if configuracion_provision.cuenta_credito:
                                    if configuracion_provision.cuenta_credito.exige_grupo:
                                        data['grupo'] = documento.contrato.grupo_id                                                                                                                                                          
                            data['contacto'] = documento.contacto_id
                            data['naturaleza'] = 'C'
                            data['credito'] = documento.provision_prima
                            data['detalle'] = 'PROVISION PRIMA CRE'
                            movimiento_serializador = ConMovimientoSerializador(data=data)
                            if movimiento_serializador.is_valid():
                                movimientos_validos.append(movimiento_serializador)
                            else:
                                return Response({'validaciones': movimiento_serializador.errors, 
                                                    'mensaje': 'Provision prima'}, status=status.HTTP_400_BAD_REQUEST)

                        if documento.provision_vacacion > 0:
                            data = data_general.copy()
                            configuracion_provision = HumConfiguracionProvision.objects.filter(tipo='VACACION', tipo_costo_id=tipo_costo_id).first()                                                       
                            if configuracion_provision:
                                data['cuenta'] = configuracion_provision.cuenta_debito_id                                                                        
                                if configuracion_provision.cuenta_debito:
                                    if configuracion_provision.cuenta_debito.exige_grupo:
                                        data['grupo'] = documento.contrato.grupo_id                                                                                                                                                          
                            data['contacto'] = documento.contacto_id
                            data['naturaleza'] = 'D'
                            data['debito'] = documento.provision_vacacion
                            data['detalle'] = 'PROVISION VACACION DEB'
                            movimiento_serializador = ConMovimientoSerializador(data=data)
                            if movimiento_serializador.is_valid():
                                movimientos_validos.append(movimiento_serializador)
                            else:
                                return Response({'validaciones': movimiento_serializador.errors, 
                                                    'mensaje': 'Provision vacacion'}, status=status.HTTP_400_BAD_REQUEST)      

                            data = data_general.copy()
                            if configuracion_provision:
                                data['cuenta'] = configuracion_provision.cuenta_credito_id                                                                        
                                if configuracion_provision.cuenta_credito:
                                    if configuracion_provision.cuenta_credito.exige_grupo:
                                        data['grupo'] = documento.contrato.grupo_id                                                                                                                                                          
                            data['contacto'] = documento.contacto_id
                            data['naturaleza'] = 'C'
                            data['credito'] = documento.provision_vacacion
                            data['detalle'] = 'PROVISION VACACION CRE'
                            movimiento_serializador = ConMovimientoSerializador(data=data)
                            if movimiento_serializador.is_valid():
                                movimientos_validos.append(movimiento_serializador)
                            else:
                                return Response({'validaciones': movimiento_serializador.errors, 
                                                    'mensaje': 'Provision vacacion'}, status=status.HTTP_400_BAD_REQUEST)                                                                                                                             
                
                with transaction.atomic():
                    for serializador in movimientos_validos:
                        serializador.save()
                    documento.estado_contabilizado = True
                    documento.save()  
                    cantidad += 1                          
            return Response({'mensaje': f'{cantidad} documentos contabilizados'}, status=status.HTTP_200_OK)        
        else:
            return Response({'mensaje':'Faltan parametros', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST) 

    @action(detail=False, methods=["post"], url_path=r'descontabilizar',)
    def descontabilizar(self, request):        
        raw = request.data
        ids = raw.get('ids')
        if ids:            
            cantidad = 0
            for id in ids:
                try:
                    documento = GenDocumento.objects.get(pk=id)                            
                    if documento.estado_contabilizado:
                        periodo_id = documento.fecha_contable.strftime("%Y%m")
                        periodo = ConPeriodo.objects.get(pk=periodo_id)
                        if periodo.estado_bloqueado == False:
                            movimientos = ConMovimiento.objects.filter(documento_id=id).delete()
                            documento.estado_contabilizado = False
                            documento.save()      
                            cantidad += 1   
                        else:
                            return Response({'mensaje':f'El periodo esta bloqueado y no se puede descontabilizar el documento', 'codigo':15}, status=status.HTTP_400_BAD_REQUEST)                                                                                 
                except GenDocumento.DoesNotExist:
                    return Response({'mensaje':f'El documento id {id} no existe', 'codigo':15}, status=status.HTTP_400_BAD_REQUEST)                        
            return Response({'mensaje': f'{cantidad} documentos descontabilizados'}, status=status.HTTP_200_OK)
        else:
            return Response({'mensaje':'Faltan parametros', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["post"], url_path=r'anular',)
    def anular(self, request):
        try:
            raw = request.data
            id = raw.get('id')
            if id:
                documento = GenDocumento.objects.get(pk=id)                
                if documento.estado_anulado == False and documento.estado_aprobado == True:    
                    if documento.estado_electronico_enviado == False:
                        if documento.afectado <= 0:
                            respuesta = self.validacion_anular(documento)
                            if respuesta['error'] == False:
                                documento_detalles = GenDocumentoDetalle.objects.filter(documento_id=id)                                          
                                for documento_detalle in documento_detalles:                                               
                                    if documento_detalle.documento_afectado_id:                                    
                                        documento_afectado = GenDocumento.objects.get(pk=documento_detalle.documento_afectado_id)
                                        if documento.documento_tipo.documento_clase_id in (200,400):                                                                    
                                            documento_afectado.afectado -= documento_detalle.pago
                                            documento_afectado.pendiente = documento_afectado.total - documento_afectado.afectado
                                            documento_afectado.save(update_fields=['afectado', 'pendiente'])                                                                    

                                    if documento_detalle.operacion_inventario != 0:
                                        existencia = InvExistencia.objects.filter(almacen_id=documento_detalle.almacen_id, item_id=documento_detalle.item_id).first()
                                        if existencia:
                                            existencia.existencia -= documento_detalle.cantidad_operada
                                            existencia.disponible -= documento_detalle.cantidad_operada
                                            existencia.save(update_fields=['existencia', 'disponible'])   
                                        item = GenItem.objects.get(id=documento_detalle.item_id)
                                        item.existencia -= documento_detalle.cantidad_operada
                                        item.disponible -= documento_detalle.cantidad_operada
                                        item.save(update_fields=['existencia', 'disponible'])                                                                               

                                    documento_detalle.cantidad = 0
                                    documento_detalle.cantidad_operada = 0
                                    documento_detalle.operacion_inventario = 0
                                    documento_detalle.precio = 0
                                    documento_detalle.porcentaje_descuento = 0
                                    documento_detalle.descuento = 0
                                    documento_detalle.subtotal = 0
                                    documento_detalle.impuesto = 0
                                    documento_detalle.impuesto_operado = 0
                                    documento_detalle.impuesto_retencion = 0
                                    documento_detalle.base_impuesto = 0
                                    documento_detalle.total = 0
                                    documento_detalle.total_bruto = 0
                                    documento_detalle.pago = 0
                                    documento_detalle.save(update_fields=['cantidad', 'precio', 'porcentaje_descuento', 'descuento', 'subtotal', 
                                                                        'impuesto', 'impuesto_operado', 'impuesto_retencion', 'base_impuesto', 
                                                                        'total', 'total_bruto', 'pago'])
                                    documento_impuestos = GenDocumentoImpuesto.objects.filter(documento_detalle_id=documento_detalle.id)
                                    for documento_impuesto in documento_impuestos:
                                        documento_impuesto.base = 0
                                        documento_impuesto.total = 0 
                                        documento_impuesto.total_operado = 0 
                                        documento_impuesto.save(update_fields=['base','total', 'total_operado'])

                                documento.estado_anulado = True  
                                documento.subtotal = 0
                                documento.total = 0
                                documento.total_bruto = 0
                                documento.base_impuesto = 0
                                documento.descuento = 0
                                documento.impuesto = 0
                                documento.impuesto_operado = 0
                                documento.impuesto_retencion = 0
                                documento.pendiente = 0
                                documento.save(update_fields=['estado_anulado', 'subtotal', 'total', 'total_bruto', 'pendiente', 'base_impuesto', 'descuento', 'impuesto', 'impuesto_operado', 'impuesto_retencion'])
                                return Response({'estado_anulado': True}, status=status.HTTP_200_OK)
                            else:
                                return Response({'mensaje':respuesta['mensaje'], 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)
                        else:
                            return Response({'mensaje':'El documento esta afectado, no se puede anular', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)                        
                    else:
                        return Response({'mensaje':'Los documentos electronicos enviados a la DIAN no se pueden anular', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)
                else:
                    return Response({'mensaje':'El documento no esta aprobado o ya esta anulado', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)
            else:
                return Response({'mensaje':'Faltan parametros', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)
        except GenDocumento.DoesNotExist:
            return Response({'mensaje':'El documento no existe', 'codigo':15}, status=status.HTTP_400_BAD_REQUEST)
   
    @action(detail=False, methods=["post"], url_path=r'imprimir',)
    def imprimir(self, request):
        raw = request.data
        id = raw.get('documento_id')
        if id:
            try:
                pdf = None
                documento = GenDocumento.objects.get(pk=id)
                if documento.documento_tipo_id in (1,2,3):                                        
                    formato = FormatoFactura()
                    pdf = formato.generar_pdf(id)                                          
                    nombres_archivo = {
                        100: f"factura_{documento.numero}.pdf" if documento.numero else "Factura.pdf",
                        101: f"notaCredito{documento.numero}.pdf" if documento.numero else "NotaCredito.pdf",
                        102: f"notaDebito{documento.numero}.pdf" if documento.numero else "NotaDebito.pdf"
                    }
                    nombre_archivo = nombres_archivo.get(documento.documento_tipo.documento_clase.id)                
                
                if documento.documento_tipo_id in (11, 12):
                    formato = FormatoDocumentoSoporte()
                    pdf = formato.generar_pdf(id)              
                    nombre_archivo = f"doc_soporte_{id}.pdf"
                # Nomina
                if documento.documento_tipo_id == 14:
                    formato = FormatoNomina()
                    pdf = formato.generar_pdf(id)              
                    nombre_archivo = f"nomina_{id}.pdf"  

                if documento.documento_tipo_id == 17:
                    formato = FormatoCuentaCobro()
                    pdf = formato.generar_pdf(id)                                        
                    nombre_archivo = f"cuentaCobro{documento.numero}.pdf" if documento.numero else f"cuentaCobro.pdf"

                if documento.documento_tipo_id == 4:
                    formato = FormatoRecibo()
                    pdf = formato.generar_pdf(id)                                        
                    nombre_archivo = f"reciboCaja{documento.numero}.pdf" if documento.numero else f"reciboCaja.pdf"

                if documento.documento_tipo_id == 8:
                    formato = FormatoEgreso()
                    pdf = formato.generar_pdf(id)                                        
                    nombre_archivo = f"egreso{documento.numero}.pdf" if documento.numero else f"egreso.pdf"

                if documento.documento_tipo_id == 5:
                    formato = FormatoCompra()
                    pdf = formato.generar_pdf(id)                                        
                    nombre_archivo = f"compra{documento.numero}.pdf" if documento.numero else f"compra.pdf"

                if documento.documento_tipo_id == 24 or documento.documento_tipo_id == 27:                                        
                    formato = FormatoFacturaPOS()
                    pdf = formato.generar_pdf(id)                                          
                    nombre_archivo = f"factura{documento.numero}.pdf" if documento.numero else f"factura.pdf"       

                if pdf:
                    response = HttpResponse(pdf, content_type='application/pdf')
                    response['Access-Control-Expose-Headers'] = 'Content-Disposition'
                    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
                    return response
                else:
                    return Response({'mensaje':'El documento no tiene un formato', 'codigo':15}, status=status.HTTP_400_BAD_REQUEST)
            except GenDocumento.DoesNotExist:
                return Response({'mensaje':'El documento no existe', 'codigo':15}, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'mensaje':'Faltan parametros', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)            
    
    @action(detail=False, methods=["post"], url_path=r'emitir',)
    def emitir_action(self, request):
        try:
            raw = request.data
            id = raw.get('documento_id')
            if id:
                respuesta = self.emitir(id)
                if respuesta['error'] == False:
                    return Response({'mensaje': 'Documento emitido correctamente'}, status=status.HTTP_200_OK)
                else:
                    return Response({'mensaje': respuesta['mensaje'], 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST)                    
            else:
                return Response({'mensaje': 'Faltan parámetros', 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST)                
        except GenDocumento.DoesNotExist:
            return Response({'mensaje': 'El documento no existe', 'codigo': 15}, status=status.HTTP_400_BAD_REQUEST)
   
    @action(detail=False, methods=["post"], url_path=r'electronico_descartar',)
    def electronico_descartar(self, request):
        try:
            raw = request.data
            id = raw.get('id')
            if id:
                documento = GenDocumento.objects.get(pk=id)
                if documento.estado_aprobado == True:
                    if documento.estado_electronico_enviado == False:
                        documento.estado_electronico_descartado = True
                        documento.save()
                        return Response({'mensaje': 'Documento descartado correctamente'}, status=status.HTTP_200_OK)
                    else:       
                        if documento.documento_tipo_id == 5:
                            if documento.evento_documento == 'PE':
                                documento.estado_electronico_descartado = True
                                documento.save()
                                return Response({'mensaje': 'Documento descartado correctamente'}, status=status.HTTP_200_OK)
                            else:
                                return Response({'mensaje': 'El documento de compra solo se puede descartar si no ha enviado eventos', 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST)    
                        else:
                            return Response({'mensaje': 'El documento ya se encuentra enviado electronicamente', 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST)
                else:
                    return Response({'mensaje': 'El documento no se puede emitir ya que no está aprobado', 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST)
            else:
                return Response({'mensaje': 'Faltan parámetros', 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST)                
        except GenDocumento.DoesNotExist:
            return Response({'mensaje': 'El documento no existe', 'codigo': 15}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["post"], url_path=r'emitir-evento',)
    def emitir_evento(self, request):
        try:
            raw = request.data
            id = raw.get('id')
            evento_id = raw.get('evento_id')
            nombre = raw.get('nombre')
            apellido = raw.get('apellido')
            identificacion = raw.get('identificacion')
            numero_identifiacion = raw.get('numero_identificacion')
            cargo = raw.get('cargo')
            area = raw.get('area')
            if id and evento_id and nombre and apellido and identificacion and numero_identifiacion and cargo and area:
                documento = GenDocumento.objects.get(pk=id)
                if documento.estado_aprobado == True:
                    empresa = GenEmpresa.objects.get(pk=1)
                    if empresa.rededoc_id:
                        if documento.estado_electronico == True: 
                            if documento.numero:
                                if evento_id in [30, 32, 33]:                                                                                                            
                                    datos_evento = {
                                        "cuentaId": empresa.rededoc_id,
                                        "documentoId" : documento.electronico_id,                                            
                                        "eventoTipoId" : evento_id,
                                        "evento" : {                                                
                                            "evento_tipo_id" : evento_id,
                                            "nombre" : nombre,
                                            "apellido" : apellido,
                                            "identificacion" :identificacion,
                                            "numero_identificacion" : numero_identifiacion,
                                            "cargo" : cargo,
                                            "area" : area,
                                            "valor_aceptado":"",
                                            "concepto_reclamo" : "",
                                        }
                                    }                                                                                                
                                    wolframio = Wolframio()
                                    respuesta = wolframio.emitir_evento(datos_evento)
                                    if respuesta['error'] == False: 
                                        if evento_id == 30:
                                            documento.evento_documento = 'RC'
                                        if evento_id == 32:
                                            documento.evento_recepcion = 'RC'                                            
                                        if evento_id == 33:
                                            documento.evento_aceptacion = 'AC'
                                            documento.estado_electronico_evento = True
                                        documento.save()
                                        return Response({'mensaje': 'Evento emitido correctamente'}, status=status.HTTP_200_OK)                                    
                                    else:
                                        return Response({'mensaje': respuesta['mensaje'], 'codigo': 15}, status=status.HTTP_400_BAD_REQUEST)                                    
                                else:
                                    return Response({'mensaje': 'Solo se admiten eventos 30, 32, 33', 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST)                                                        
                            else:
                                return Response({'mensaje': 'La factura no cuenta con un número', 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST)
                        else:
                            return Response({'mensaje': "El documento debe estar emitido", 'codigo': 15}, status=status.HTTP_400_BAD_REQUEST)                        
                    else:  
                        return Response({'mensaje': 'La empresa no se encuentra activada para emitir documentos electronicos', 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST)
                else:
                    return Response({'mensaje': 'El documento no está aprobado', 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST)
            else:
                return Response({'mensaje': 'Faltan parámetros', 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST)                
        except GenDocumento.DoesNotExist:
            return Response({'mensaje': 'El documento no existe', 'codigo': 15}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["post"], url_path=r'electronico_respuesta_emitir', permission_classes=[permissions.AllowAny])
    def electronico_respuesta_emitir(self, request):
        raw = request.data
        documento_id = raw.get('documento_id')
        qr = raw.get('qr')
        cue = raw.get('cue')
        fecha_validacion = raw.get('fecha_validacion')
        if documento_id and qr and cue and fecha_validacion:
            try:
                documento = GenDocumento.objects.get(pk=documento_id)
                if documento.estado_electronico_enviado:
                    if documento.estado_electronico == False:
                        fecha_validacion_obj = datetime.strptime(fecha_validacion, '%Y-%m-%d %H:%M:%S')
                        documento.qr = qr
                        documento.cue = cue
                        documento.fecha_validacion = fecha_validacion_obj
                        documento.estado_electronico = True
                        documento.save()                
                        self.notificar(documento_id)
                        return Response({'respuesta':True}, status=status.HTTP_200_OK)
                    else:                        
                        return Response({'respuesta':True}, status=status.HTTP_200_OK)    
                else:
                    return Response({'mensaje':'No se puede entregar una respuesta porque el documento no se ha enviado'}, status=status.HTTP_400_BAD_REQUEST)
            except GenDocumento.DoesNotExist:
                return Response({'mensaje': 'El documento no existe', 'codigo': 15}, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'mensaje': 'Faltan parámetros', 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=["post"], url_path=r'notificar')
    def electronico_notificar(self, request):
        raw = request.data
        documento_id = raw.get('documento_id')
        if documento_id:
            respuesta = self.notificar(documento_id)
            if respuesta['error'] == False:
                return Response({'mensaje': 'notificado'}, status=status.HTTP_200_OK)
            else:
                return Response({'mensaje': respuesta['mensaje'], 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST)      
        else:
            return Response({'mensaje': 'Faltan parámetros', 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST)  

    @action(detail=False, methods=["post"], url_path=r'renotificar')
    def electronico_renotificar(self, request):
        raw = request.data
        documento_id = raw.get('documento_id')        
        if documento_id:
            try:
                documento = GenDocumento.objects.get(id=documento_id)
                if documento.estado_electronico_notificado == True:                    
                    if documento.electronico_id:                        
                        formatoFactura = FormatoFactura()
                        pdf = formatoFactura.generar_pdf(documento_id)   
                        pdf_base64 = "data:application/pdf;base64," + base64.b64encode(pdf).decode('utf-8')                                                    
                        wolframio = Wolframio()
                        respuesta = wolframio.renotificar(documento.electronico_id, documento.contacto.correo, pdf_base64)
                        if respuesta['error'] == False: 
                            return Response({'mensaje': 'Documento re-notificado con éxito'}, status=status.HTTP_200_OK)           
                        else:
                            return Response({'mensaje': f"{respuesta['mensaje']}", 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST)                    
                    else:
                        return Response({'mensaje': 'El documento esta marcado como notificado pero no tiene electronico_id', 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST)    
                else:                      
                    return Response({'mensaje': 'El documento nunca ha sido notificado', 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST)
            except GenDocumento.DoesNotExist:
                return Response({'mensaje': 'El documento no existe', 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST)     
        else:
            return Response({'mensaje': 'Faltan parámetros', 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST) 

    @action(detail=False, methods=["post"], url_path=r'electronico_log')
    def electronico_log(self, request):
        raw = request.data
        documento_id = raw.get('documento_id')
        if documento_id:
            try:
                documento = GenDocumento.objects.get(id=documento_id)
                if documento.estado_electronico_notificado == True:                    
                    if documento.electronico_id:
                        empresa = GenEmpresa.objects.get(pk=1)
                        if empresa.rededoc_id:                       
                            zinc = Zinc()                        
                            respuesta = zinc.log_envio(empresa.rededoc_id, documento.electronico_id)
                            if respuesta['error'] == False: 
                                return Response({'log': respuesta['log']}, status=status.HTTP_200_OK)
                            else:
                                return Response({'mensaje': f"{respuesta['mensaje']}", 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST)                                                                                                                    
                        else:
                            return Response({'mensaje': 'La empresa no esta activa en facturacion electronica', 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST)
                    else:
                        return Response({'mensaje': 'El documento esta marcado como notificado pero no tiene electronico_id', 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST)    
                else:                      
                    return Response({'mensaje': 'El documento nunca ha sido notificado', 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST)
            except GenDocumento.DoesNotExist:
                return Response({'mensaje': 'El documento no existe', 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST)     
        else:
            return Response({'mensaje': 'Faltan parámetros', 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST) 

    @action(detail=False, methods=["post"], url_path=r'evento-dian')
    def evento_dian(self, request):
        raw = request.data
        id = raw.get('id')
        if id:
            try:
                documento = GenDocumento.objects.get(id=id)
                if documento.estado_electronico == True:                    
                    if documento.electronico_id:
                        empresa = GenEmpresa.objects.get(pk=1)
                        if empresa.rededoc_id:                       
                            wolframio = Wolframio()
                            respuesta = wolframio.eventos(documento.electronico_id)                            
                            if respuesta['error'] == False: 
                                return Response({'eventos': respuesta['eventos']}, status=status.HTTP_200_OK)
                            else:
                                return Response({'mensaje': f"{respuesta['mensaje']}", 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST)                                                                                                                    
                        else:
                            return Response({'mensaje': 'La empresa no esta activa en facturacion electronica', 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST)
                    else:
                        return Response({'mensaje': 'El documento esta emitido pero no tiene electronico_id', 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST)    
                else:                      
                    return Response({'mensaje': 'El documento no esta emitido a la DIAN', 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST)
            except GenDocumento.DoesNotExist:
                return Response({'mensaje': 'El documento no existe', 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST)     
        else:
            return Response({'mensaje': 'Faltan parámetros', 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["post"], url_path=r'proceso_corregir_pendiente')
    def proceso_corregir_pendiente(self, request):
        query = f'''
            select
                d.id,
                d.afectado,
                coalesce((
                select SUM(precio) from gen_documento_detalle dd 
                left join gen_documento d1 on dd.documento_id = d1.id
                left join gen_documento_tipo dt1 on d1.documento_tipo_id = dt1.id
                where dd.documento_afectado_id = d.id and d1.estado_aprobado = true and (dt1.documento_clase_id = 200 or dt1.documento_clase_id = 400)),0) as afectado_detalle
            from
                gen_documento d
            left join gen_documento_tipo dt on d.documento_tipo_id = dt.id
            where d.estado_aprobado = true and (dt.cobrar = true or dt.pagar = true)
        '''
        actualizados = 0
        documentos_actualizar = []
        documentos_query = GenDocumento.objects.raw(query)
        with transaction.atomic():
            for documento_query in documentos_query:
                if documento_query.afectado != documento_query.afectado_detalle:
                    documento = GenDocumento.objects.get(id=documento_query.id)            
                    documento.afectado = documento_query.afectado_detalle
                    documento.pendiente = documento.total - documento_query.afectado_detalle            
                    documentos_actualizar.append(documento)
                    actualizados += 1
            GenDocumento.objects.bulk_update(documentos_actualizar, ['afectado', 'pendiente'])                
        return Response({'mensaje':f'Se corrigieron {actualizados} documentos'}, status=status.HTTP_200_OK)

    @action(detail=False, methods=["post"], url_path=r'proceso_corregir_provision')
    def proceso_corregir_provision(self, request):
        raw = request.data
        id = raw.get('id')
        if id:
            try:
                documento = GenDocumento.objects.get(id=id)
                if documento.documento_tipo_id == 14:
                    with transaction.atomic():
                        provision_cesantia = documento.base_prestacion * Decimal(0.0833)
                        provision_interes = provision_cesantia * Decimal(0.12)
                        provision_prima = documento.base_prestacion * Decimal(0.0833)
                        provision_vacacion = documento.base_prestacion_vacacion * Decimal(0.0417)
                        documento.provision_cesantia = round(provision_cesantia)
                        documento.provision_interes = round(provision_interes)
                        documento.provision_prima = round(provision_prima)
                        documento.provision_vacacion = round(provision_vacacion)
                        documento.save()
                    return Response({'mensaje':f'Se corrigieron las provisiones'}, status=status.HTTP_200_OK)                
                else:
                    return Response({'mensaje': 'Solo se pueden corregir documentos nomina', 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST)
            except GenDocumento.DoesNotExist:
                return Response({'mensaje':'El documento no existe', 'codigo':15}, status=status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'mensaje': 'Faltan parámetros', 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST)
                   
    @action(detail=False, methods=["get"], url_path=r'resumen-cobrar',)
    def resumen_cobrar(self, request):      
        fecha_actual = timezone.localtime(timezone.now()).date()
        resumen = GenDocumento.objects.filter(documento_tipo__cobrar=True, estado_aprobado=True, pendiente__gt=0
                                                ).aggregate(
                                                    cantidad=Count('id'), 
                                                    saldo_pendiente=Coalesce(Sum('pendiente'), 0, output_field=DecimalField()))        
        vigente = GenDocumento.objects.filter(documento_tipo__cobrar=True, estado_aprobado=True, pendiente__gt=0, fecha_vence__gte=fecha_actual
                                                ).aggregate(
                                                    cantidad=Count('id'), 
                                                    saldo_pendiente=Coalesce(Sum('pendiente'), 0, output_field=DecimalField()))
        vencido = GenDocumento.objects.filter(documento_tipo__cobrar=True, estado_aprobado=True, pendiente__gt=0, fecha_vence__lt=fecha_actual
                                                ).aggregate(
                                                    cantidad=Count('id'), 
                                                    saldo_pendiente=Coalesce(Sum('pendiente'), 0, output_field=DecimalField()))    
        #print("Zona horaria de Django:", timezone.get_current_timezone_name())  # Ej: "America/Bogota"
        #print("Fecha/Hora actual con zona:", timezone.localtime(timezone.now()))  # Fecha LOCAL con hora
        #print("Fecha actual (date()):", timezone.now().date())  # Solo fecha (puede ser ambiguo)                    
        return Response({'resumen':resumen, 'vigente': vigente, 'vencido': vencido}, status=status.HTTP_200_OK)

    @action(detail=False, methods=["get"], url_path=r'resumen-pagar',)
    def resumen_pagar(self, request):      
        fecha_actual = timezone.now().date()
        resumen = GenDocumento.objects.filter(documento_tipo__pagar=True, estado_aprobado=True, pendiente__gt=0
                                                ).aggregate(
                                                    cantidad=Count('id'), 
                                                    saldo_pendiente=Coalesce(Sum('pendiente'), 0, output_field=DecimalField()))
        vigente = GenDocumento.objects.filter(documento_tipo__pagar=True, estado_aprobado=True, pendiente__gt=0, fecha_vence__gte=fecha_actual
                                                ).aggregate(
                                                    cantidad=Count('id'), 
                                                    saldo_pendiente=Coalesce(Sum('pendiente'), 0, output_field=DecimalField()))
        vencido = GenDocumento.objects.filter(documento_tipo__pagar=True, estado_aprobado=True, pendiente__gt=0, fecha_vence__lt=fecha_actual
                                                ).aggregate(
                                                    cantidad=Count('id'), 
                                                    saldo_pendiente=Coalesce(Sum('pendiente'), 0, output_field=DecimalField()))                
        return Response({'resumen': resumen, 'vigente':vigente, 'vencido': vencido}, status=status.HTTP_200_OK)

    @action(detail=False, methods=["get"], url_path=r'resumen-venta-dia',)
    def resumen_venta_dia(self, request):      
        fecha_actual = timezone.now().date()
        # Obtener el primer y último día del mes actual
        primer_dia_mes = fecha_actual.replace(day=1)
        ultimo_dia_mes = (primer_dia_mes + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        dias_del_mes = [primer_dia_mes + timedelta(days=i) for i in range((ultimo_dia_mes - primer_dia_mes).days + 1)]
        ventas_por_dia = GenDocumento.objects.filter(fecha__year=fecha_actual.year, fecha__month=fecha_actual.month
                                                  ).annotate(dia=TruncDay('fecha')
                                                             ).values('dia'
                                                                      ).annotate(total=Sum('total')).order_by('dia')
        ventas_dict = {venta['dia']: venta['total'] for venta in ventas_por_dia}
        venta_dia = [{'dia': dia, 'total': ventas_dict.get(dia, 0)} for dia in dias_del_mes]        
        return Response({'resumen': venta_dia}, status=status.HTTP_200_OK)

    @action(detail=False, methods=["post"], url_path=r'importar',)
    def importar(self, request):
        raw = request.data        
        archivo_base64 = raw.get('archivo_base64')   
        documento_clase_id = raw.get('documento_tipo_id')     
        if archivo_base64 and documento_clase_id:
            try:
                archivo_data = base64.b64decode(archivo_base64)
                archivo = BytesIO(archivo_data)
                wb = openpyxl.load_workbook(archivo)
                sheet = wb.active    
            except Exception as e:     
                return Response({f'mensaje':'Error procesando el archivo, valide que es un archivo de excel .xlsx', 'codigo':15}, status=status.HTTP_400_BAD_REQUEST)  
            cuentas_map = {c.codigo: c.id for c in ConCuenta.objects.all()}
            data_modelo = []
            errores = False
            errores_datos = []
            documento_tipo_id = None
            if documento_clase_id == 201:
                documento_tipo_id = 18
            if documento_clase_id == 401:
                documento_tipo_id = 19
            registros_importados = 0            
            for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):     
                pendiente = 0
                if documento_clase_id in (201,401):
                    pendiente = row[4]
                fecha = row[1].date()
                fecha_vence = row[2].date()
                cuenta_codigo = str(row[5])
                cuenta_id = cuentas_map.get(cuenta_codigo)
                data = {
                    'numero': row[0],                    
                    'fecha':fecha,
                    'fecha_contable': fecha,
                    'fecha_vence':fecha_vence,
                    'contacto':row[3],
                    'total':row[4],
                    'cuenta':cuenta_id,
                    'pendiente': pendiente,
                    'documento_tipo': documento_tipo_id,
                    'estado_aprobado': True,
                    'empresa': 1
                }         
                if cuenta_codigo:   
                    if not cuenta_id:
                        errores = True
                        errores_datos.append({
                            'fila': i,
                            'errores': {'cuenta': [f'Cuenta "{cuenta_codigo}" no encontrada.']}
                        })                       
                else:
                    errores = True
                    errores_datos.append({
                        'fila': i,
                        'errores': {'cuenta': [f'La cuenta es requerida']}
                    })                       
                if errores == False:
                    serializer = GenDocumentoSerializador(data=data)
                    if serializer.is_valid():
                        data_modelo.append(serializer.validated_data)
                        registros_importados += 1
                    else:
                        errores = True
                        error_dato = {
                            'fila': i,
                            'errores': serializer.errors
                        }                                    
                        errores_datos.append(error_dato)                    
            if not errores:
                for detalle in data_modelo:
                    GenDocumento.objects.create(**detalle)
                gc.collect()
                return Response({'registros_importados': registros_importados}, status=status.HTTP_200_OK)
            else:
                gc.collect()                    
                return Response({'mensaje':'Errores de validacion', 'codigo':1, 'errores_validador': errores_datos}, status=status.HTTP_400_BAD_REQUEST)   
        else:
            return Response({'mensaje':'Faltan parametros', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["post"], url_path=r'importar-detalle-cuenta',)
    def importar_detalle_cuenta_action(self, request):
        raw = request.data
        documento_id = raw.get('documento_id')
        archivo_base64 = raw.get('archivo_base64')
        if documento_id and archivo_base64:
            try:
                documento = GenDocumento.objects.get(pk=documento_id)
            except GenDocumento.DoesNotExist:
                return Response({'mensaje':'El documento no existe', 'codigo':15}, status=status.HTTP_400_BAD_REQUEST)
            try:
                archivo_data = base64.b64decode(archivo_base64)
                archivo = BytesIO(archivo_data)
                wb = openpyxl.load_workbook(archivo)
                sheet = wb.active         
            except Exception as e:     
                return Response({'mensaje':'Error procesando el archivo', 'codigo':15}, status=status.HTTP_400_BAD_REQUEST)  
            
            data_documento_detalle = []
            errores = False
            errores_datos = []
            registros_importados = 0
            if documento.documento_tipo_id == 13:
                for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    if len(row) == 5:
                        data = {
                            'cuenta': row[0],
                            'numero_identificacion':row[1],
                            'debito': row[2] if row[2] is not None else 0,
                            'credito': row[3] if row[3] is not None else 0,
                            'base': row[4] if row[4] is not None else 0,
                            'descripcion': row[5]                    
                        }                    
                        if not data['cuenta']:
                            error_dato = {
                                'fila': i,
                                'Mensaje': 'Debe digitar la cuenta'
                            }
                            errores_datos.append(error_dato)
                            errores = True
                        else:
                            cuenta = ConCuenta.objects.filter(codigo=data['cuenta']).first()
                            if cuenta is None:
                                error_dato = {
                                    'fila': i,
                                    'Mensaje': f'La cuenta {data["cuenta"]} no existe'
                                }
                                errores_datos.append(error_dato)
                                errores = True
                            else:
                                data['cuenta_id'] = cuenta.id

                        if not data['numero_identificacion']:
                            error_dato = {
                                'fila': i,
                                'Mensaje': 'Debe digitar el numero de identificacion'
                            }
                            errores_datos.append(error_dato)
                            errores = True  
                        else:
                            contacto = GenContacto.objects.filter(numero_identificacion=data['numero_identificacion']).first()
                            if contacto is None:
                                error_dato = {
                                    'fila': i,
                                    'Mensaje': f'El contacto con numero identificacion {data["numero_identificacion"]} no existe'
                                }
                                errores_datos.append(error_dato)
                                errores = True
                            else:
                                data['contacto_id'] = contacto.id                                 
                        
                        if data['debito'] == 0 and data['credito'] == 0:
                                error_dato = {
                                    'fila': i,
                                    'Mensaje': f'Los debitos y creditos estan en cero'
                                }
                                errores_datos.append(error_dato)
                                errores = True
                        else:
                            if data['debito'] != 0 and data['credito'] != 0:
                                error_dato = {
                                    'fila': i,
                                    'Mensaje': f'Los debitos y creditos tienen valores'
                                }
                                errores_datos.append(error_dato)
                                errores = True
                            else:
                                total = data['debito']
                                naturaleza = 'D'
                                if data['credito'] > 0:
                                    total = data['credito']
                                    naturaleza = 'C'
                                data['naturaleza'] = naturaleza
                                data['total'] = total
                        data_documento_detalle.append(data) 
                    else:
                        error_dato = {
                            'fila': i,
                            'Mensaje': f'La linea no tiene 5 columnas'
                        }
                        errores_datos.append(error_dato)
                        errores = True
            if errores == False:
                for detalle in data_documento_detalle:
                    GenDocumentoDetalle.objects.create(
                        documento=documento,
                        cuenta_id=detalle['cuenta_id'],
                        contacto_id=detalle['contacto_id'],
                        total=detalle['total'],
                        base_impuesto=detalle['base'],
                        naturaleza=detalle['naturaleza'],
                        detalle=detalle['descripcion']
                    )
                    registros_importados += 1
                return Response({'registros_importados': registros_importados}, status=status.HTTP_200_OK)
            else:
                return Response({'errores': True, 'errores_datos': errores_datos}, status=status.HTTP_400_BAD_REQUEST)       
        else:
            return Response({'mensaje':'Faltan parametros', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["post"], url_path=r'importar-detalle-inventario',)
    def importar_detalle_inventario(self, request):
        raw = request.data
        documento_id = raw.get('documento_id')
        archivo_base64 = raw.get('archivo_base64')
        if documento_id and archivo_base64:
            try:
                documento = GenDocumento.objects.get(pk=documento_id)
            except GenDocumento.DoesNotExist:
                return Response({'mensaje':'El documento no existe', 'codigo':15}, status=status.HTTP_400_BAD_REQUEST)
            try:
                archivo_data = base64.b64decode(archivo_base64)
                archivo = BytesIO(archivo_data)
                wb = openpyxl.load_workbook(archivo)
                sheet = wb.active         
            except Exception as e:     
                return Response({'mensaje':'Error procesando el archivo', 'codigo':15}, status=status.HTTP_400_BAD_REQUEST)  
            
            data_documento_detalle = []
            errores = False
            errores_datos = []
            registros_importados = 0
            if documento.documento_tipo_id in [9, 10]:
                for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    if len(row) == 4:
                        data = {
                            'item': row[0],
                            'almacen':row[1],
                            'cantidad': row[2] if row[2] is not None else 0,
                            'precio': row[3] if row[3] is not None else 0,              
                        }                    
                        if not data['item']:
                            error_dato = {
                                'fila': i,
                                'Mensaje': 'Debe digitar el código del item'
                            }
                            errores_datos.append(error_dato)
                            errores = True
                        else:
                            item = GenItem.objects.filter(id=data['item']).first()
                            if item is None:
                                error_dato = {
                                    'fila': i,
                                    'Mensaje': f'El item {data["item"]} no existe'
                                }
                                errores_datos.append(error_dato)
                                errores = True
                            else:
                                data['item_id'] = item.id

                        if not data['almacen']:
                            error_dato = {
                                'fila': i,
                                'Mensaje': 'Debe digitar el código del almacen'
                            }
                            errores_datos.append(error_dato)
                            errores = True  
                        else:
                            almacen = InvAlmacen.objects.filter(id=data['almacen']).first()
                            if almacen is None:
                                error_dato = {
                                    'fila': i,
                                    'Mensaje': f'El almacen con código {data["almacen"]} no existe'
                                }
                                errores_datos.append(error_dato)
                                errores = True
                            else:
                                data['almacen_id'] = almacen.id                                 
                        data_documento_detalle.append(data) 
                    else:
                        error_dato = {
                            'fila': i,
                            'Mensaje': f'La linea no tiene 4 columnas'
                        }
                        errores_datos.append(error_dato)
                        errores = True
            if errores == False:
                for detalle in data_documento_detalle:
                    GenDocumentoDetalle.objects.create(
                        documento=documento,
                        item_id=detalle['item_id'],
                        almacen_id=detalle['almacen_id'],
                        cantidad=detalle['cantidad'],
                        precio=detalle['precio']
                    )
                    registros_importados += 1
                return Response({'registros_importados': registros_importados}, status=status.HTTP_200_OK)
            else:
                return Response({'errores': True, 'errores_datos': errores_datos}, status=status.HTTP_400_BAD_REQUEST)       
        else:
            return Response({'mensaje':'Faltan parametros', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["post"], url_path=r'generar-nomina-electronica',)
    def generar_nomina_electronica(self, request):      
        raw = request.data
        anio = int(raw.get('anio'))
        mes = int(raw.get('mes'))
        if anio and mes:
            with transaction.atomic():                    
                documento_tipo = GenDocumentoTipo.objects.get(id=15)
                fecha_desde = date(anio, mes, 1)            
                ultimo_dia_mes = calendar.monthrange(anio, mes)[1]
                fecha_hasta = date(anio, mes, ultimo_dia_mes)
                nominas_mes = GenDocumento.objects.filter(fecha__gte=fecha_desde, 
                                                        fecha__lte=fecha_hasta, 
                                                        documento_tipo_id=14,
                                                        documento_referencia_id = None,
                                                        estado_aprobado = True)\
                                                .values('contacto_id', 'contrato_id')\
                                                .annotate(
                                                    base_cotizacion=Sum('base_cotizacion'),
                                                    base_prestacion=Sum('base_prestacion'),
                                                    deduccion=Sum('deduccion'),
                                                    devengado=Sum('devengado'),
                                                    total=Sum('total'))           
                for nomina_mes in nominas_mes:
                    contrato = HumContrato.objects.get(pk=nomina_mes['contrato_id'])
                    data = {
                        'empresa': 1,
                        'numero': documento_tipo.consecutivo,
                        'documento_tipo': 15,
                        'fecha': fecha_desde,
                        'fecha_contable': fecha_desde,                    
                        'fecha_hasta': fecha_hasta,
                        'contacto': nomina_mes['contacto_id'],
                        'contrato': nomina_mes['contrato_id'],
                        'grupo': contrato.grupo_id,
                        'salario': contrato.salario,
                        'base_cotizacion': nomina_mes['base_cotizacion'],
                        'base_prestacion': nomina_mes['base_prestacion'],
                        'deduccion': nomina_mes['deduccion'],
                        'devengado': nomina_mes['devengado'],
                        'total': nomina_mes['total'],
                        'estado_aprobado': True
                    }
                    documento_tipo.consecutivo += 1
                    documento_tipo.save()
                    documento_serializador = GenDocumentoSerializador(data=data)
                    if documento_serializador.is_valid():
                        documento = documento_serializador.save()
                        nominas = GenDocumento.objects.filter(fecha__gte=fecha_desde, 
                                                            fecha__lte=fecha_hasta, 
                                                            documento_tipo_id=14,
                                                            contacto_id = nomina_mes['contacto_id'],
                                                            contrato_id = nomina_mes['contrato_id'],
                                                            documento_referencia_id = None,
                                                            estado_aprobado = True) 
                        for nomina in nominas:
                            nomina.documento_referencia = documento
                            nomina.save() 
                    else:
                        return Response({'validaciones':documento_serializador.errors}, status=status.HTTP_400_BAD_REQUEST)                 
            return Response({'resumen': 1}, status=status.HTTP_200_OK)
        else:
            return Response({'mensaje': 'Faltan parámetros', 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST)           

    @action(detail=False, methods=["post"], url_path=r'generar-masivo')
    def generar_masivo(self, request):      
        raw = request.data
        documento_tipo_id = raw.get('documento_tipo_id', 1)
        documento_ids = raw.get('ids', []) 
        generar_todos = raw.get('generar_todos', False)
        
        if documento_tipo_id:
            try:
                documento_tipo = GenDocumentoTipo.objects.get(pk=documento_tipo_id)
            except GenDocumentoTipo.DoesNotExist:
                return Response({'mensaje': 'El tipo de documento no existe', 'codigo': 2}, status=status.HTTP_400_BAD_REQUEST)
            
            if generar_todos:
                documentos = GenDocumento.objects.filter(documento_tipo_id=16)
            else:
                if not documento_ids:
                    return Response({'mensaje': 'Debe proporcionar los IDs de documentos o habilitar "generar_todos"', 'codigo': 3}, status=status.HTTP_400_BAD_REQUEST)
                documentos = GenDocumento.objects.filter(id__in=documento_ids, documento_tipo_id=16)
            
            if not documentos.exists():
                return Response({'mensaje': 'No se encontraron documentos para procesar', 'codigo': 4}, status=status.HTTP_400_BAD_REQUEST)
            
            # Procesar documentos
            for documento in documentos:
                dias_plazo = documento.plazo_pago.dias
                documento_data = documento.__dict__.copy()
                documento_data.pop('id', None)           
                documento_data.pop('_state', None)
                documento_data['documento_tipo_id'] = documento_tipo_id
                documento_data['resolucion_id'] = documento_tipo.resolucion_id
                documento_data['fecha'] = timezone.now()
                documento_data['fecha_contable'] = timezone.now()
                documento_data['fecha_vence'] = timezone.now() + timedelta(days=dias_plazo)
                
                documento_nuevo = GenDocumento(**documento_data)
                documento_nuevo.save()
                
                documento_detalles = GenDocumentoDetalle.objects.filter(documento_id=documento.id)
                for documento_detalle in documento_detalles:
                    documento_detalle_data = documento_detalle.__dict__.copy()
                    documento_detalle_data.pop('id', None)
                    documento_detalle_data.pop('_state', None)
                    documento_detalle_data['documento_id'] = documento_nuevo.id
                    
                    documento_detalle_nuevo = GenDocumentoDetalle(**documento_detalle_data)
                    documento_detalle_nuevo.save()
                    
                    documento_impuestos = GenDocumentoImpuesto.objects.filter(documento_detalle_id=documento_detalle.id)
                    for documento_impuesto in documento_impuestos:
                        documento_impuesto_data = documento_impuesto.__dict__.copy()
                        documento_impuesto_data.pop('id', None)
                        documento_impuesto_data.pop('_state', None)
                        documento_impuesto_data['documento_detalle_id'] = documento_detalle_nuevo.id
                        
                        documento_impuesto_nuevo = GenDocumentoImpuesto(**documento_impuesto_data)
                        documento_impuesto_nuevo.save()
            
            return Response({'mensaje': 'Proceso exitoso'}, status=status.HTTP_200_OK)
        else:
            return Response({'mensaje': 'Faltan parámetros', 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["post"], url_path=r'cargar-activo')
    def cargar_activo(self, request):      
        raw = request.data
        id = raw.get('id')            
        if id:
            try:
                documento = GenDocumento.objects.get(pk=id)
            except GenDocumentoTipo.DoesNotExist:
                return Response({'mensaje': 'El documento no existe', 'codigo': 2}, status=status.HTTP_400_BAD_REQUEST)                                
            fecha = documento.fecha
            fecha_desde = fecha.replace(day=1)
            ultimo_dia = calendar.monthrange(fecha.year, fecha.month)[1]
            fecha_hasta = datetime(fecha.year, fecha.month, ultimo_dia).date()                       
            activos = ConActivo.objects.filter(
                Q(fecha_baja__gte=fecha_desde) | Q(fecha_baja__isnull=True),
                fecha_activacion__lte=fecha_hasta)
            total = 0
            for activo in activos:
                if activo.depreciacion_saldo > 0:
                    dias = 30
                    depreciar = activo.depreciacion_periodo
                    if activo.fecha_activacion >= fecha_desde and activo.fecha_activacion <= fecha_hasta:
                        diferencia = fecha_hasta - activo.fecha_activacion
                        dias = diferencia.days                        
                        if fecha.month == 2:                        
                            if ultimo_dia == 28:
                                dias += 2
                            if ultimo_dia == 29:
                                dias += 1
                    if activo.fecha_baja:
                        if activo.fecha_baja >= fecha_desde and activo.fecha_baja <= fecha_hasta:                        
                            diferencia = fecha_hasta - activo.fecha_baja
                            dias -= (diferencia.days + 1)
                    if dias > 30:
                        dias = 30                    
                    if dias != 30:
                        depreciacion_dia = activo.depreciacion_periodo / 30
                        depreciar = round(depreciacion_dia * dias)
                    if activo.depreciacion_saldo < depreciar:
                        depreciar = activo.depreciacion_saldo
                    total += depreciar
                    depreciacion_acumulada = activo.depreciacion_acumulada + depreciar
                    saldo = activo.depreciacion_saldo - depreciar
                    data = {
                        'tipo_registro': 'D',
                        'documento': documento.id,
                        'contacto': documento.contacto_id,
                        'activo': activo.id,
                        'precio': depreciar,
                        'dias': dias,
                        'grupo': documento.grupo_id
                    }
                    documento_detalle_serializador = GenDocumentoDetalleSerializador(data=data)
                    if documento_detalle_serializador.is_valid():
                        documento_detalle_serializador.save()
                    else:
                        return Response({'validaciones':documento_detalle_serializador.errors}, status=status.HTTP_400_BAD_REQUEST)                                                                                                  
            documento.total = total
            documento.save()
            return Response({'mensaje': 'Proceso exitoso'}, status=status.HTTP_200_OK)
        else:
            return Response({'mensaje': 'Faltan parámetros', 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST)
        
    @action(detail=False, methods=["post"], url_path=r'cargar-cierre')
    def cargar_cierre(self, request):      
        raw = request.data
        id = raw.get('id') 
        cuenta_desde_codigo = raw.get('cuenta_desde_codigo')
        cuenta_hasta_codigo = raw.get('cuenta_hasta_codigo') 
        cuenta_cierre_id = raw.get('cuenta_cierre_id')                  
        if id and cuenta_desde_codigo and cuenta_hasta_codigo and cuenta_cierre_id:
            try:
                documento = GenDocumento.objects.get(pk=id)
            except GenDocumentoTipo.DoesNotExist:
                return Response({'mensaje': 'El documento no existe', 'codigo': 2}, status=status.HTTP_400_BAD_REQUEST)                                
            try:
                cuenta = ConCuenta.objects.get(pk=cuenta_cierre_id)
            except ConCuenta.DoesNotExist:
                return Response({'mensaje': 'La cuenta no existe', 'codigo': 2}, status=status.HTTP_400_BAD_REQUEST)             
            movimientos = ConMovimiento.objects.filter(
                    cuenta__codigo__gte=cuenta_desde_codigo,
                    cuenta__codigo__lte=cuenta_hasta_codigo  
                ).values(
                    'cuenta_id',
                    'contacto_id'
                ).annotate(
                    debito=Sum('debito'),
                    credito=Sum('credito')                
                ).order_by('cuenta_id', 'contacto_id')              
            for movimiento in movimientos:
                saldo_final = movimiento['debito'] - movimiento['credito']
                if saldo_final < 0 or saldo_final > 0:
                    data = {
                        'tipo_registro': 'C',
                        'documento': documento.id,
                        'cuenta': movimiento['cuenta_id'],
                        'contacto': movimiento['contacto_id'],
                        'grupo': documento.grupo_contabilidad_id,
                        'detalle': 'CIERRE AÑO'
                    }
                    if saldo_final < 0:
                        data['naturaleza'] = 'D'
                        data['precio'] = saldo_final * -1
                    else:
                        data['naturaleza'] = 'C'
                        data['precio'] = saldo_final
                    documento_detalle_serializador = GenDocumentoDetalleSerializador(data=data)
                    if documento_detalle_serializador.is_valid():
                        documento_detalle_serializador.save()
                    else:
                        return Response({'validaciones':documento_detalle_serializador.errors}, status=status.HTTP_400_BAD_REQUEST)
                    
                    data = {
                        'tipo_registro': 'C',
                        'documento': documento.id,
                        'cuenta': cuenta_cierre_id,
                        'contacto': documento.contacto_id,
                        'grupo': documento.grupo_contabilidad_id,
                        'detalle': 'CIERRE AÑO'
                    }
                    if saldo_final < 0:
                        data['naturaleza'] = 'C'
                        data['precio'] = saldo_final * -1
                    else:
                        data['naturaleza'] = 'D'
                        data['precio'] = saldo_final
                    documento_detalle_serializador = GenDocumentoDetalleSerializador(data=data)
                    if documento_detalle_serializador.is_valid():
                        documento_detalle_serializador.save()
                    else:
                        return Response({'validaciones':documento_detalle_serializador.errors}, status=status.HTTP_400_BAD_REQUEST)                
                                     
            return Response({'mensaje': 'Proceso exitoso'}, status=status.HTTP_200_OK)
        else:
            return Response({'mensaje': 'Faltan parámetros', 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST)        

    @action(detail=False, methods=["post"], url_path=r'plano-banco',)
    def plano_banco(self, request):        
        raw = request.data
        id = raw.get('id', None)
        # 1 Bancolombia sap
        # 2 Bancolombia pap
        tipo_plano = raw.get('tipo_plano')
        if id and tipo_plano:
            if tipo_plano == 1 or tipo_plano == 2:
                try:                    
                    documento = GenDocumento.objects.get(pk=id)
                    if documento.cuenta_banco:                        
                        documento_detalles = GenDocumentoDetalle.objects.filter(documento_id = id)
                        for documento_detalle in documento_detalles:
                            if documento_detalle.contacto == None:
                                return Response({'mensaje':'Uno de los detalles no tiene contacto', 'codigo':15}, status=status.HTTP_400_BAD_REQUEST)                      
                        if tipo_plano == 1:                            
                            numero_registros = 0
                            total_pagar = 0
                            for documento_detalle in documento_detalles:
                                numero_registros += 1
                                total_pagar += documento_detalle.pago
                            empresa = GenEmpresa.objects.get(pk=1)
                            lineas = []       
                            nit = Utilidades.rellenar(empresa.numero_identificacion, 10, "0", "I")
                            nombre = Utilidades.eliminar_caracteres_especiales(empresa.nombre_corto).upper()
                            nombre = Utilidades.rellenar(nombre, 16, "0", "D")
                            fecha = datetime.now().strftime('%y%m%d')
                            numero_registros = Utilidades.rellenar(str(numero_registros), 6, "0", "I")
                            total_pagar = int(total_pagar)
                            total_pagar = Utilidades.rellenar(str(total_pagar), 24, "0", "I")
                            numero_cuenta = documento.cuenta_banco.numero_cuenta
                            cuenta_banco_clase = "S"
                            if documento.cuenta_banco.cuenta_banco_clase:
                                if documento.cuenta_banco.cuenta_banco_clase_id == 2:
                                    cuenta_banco_clase = "D"
                            lineas.append(f"1{nit}{nombre}225PAGONOMINA{fecha}A{fecha}{numero_registros}{total_pagar}{numero_cuenta}{cuenta_banco_clase}\n")                    
                            for documento_detalle in documento_detalles:
                                numero_identificacion = Utilidades.rellenar(documento_detalle.contacto.numero_identificacion, 15, "0", "I")
                                nombre = Utilidades.rellenar(documento_detalle.contacto.nombre_corto, 18, "0", "D")
                                codigo_banco = "1007"
                                if documento_detalle.contacto.banco:
                                    if documento_detalle.contacto.banco_id == 1:
                                        codigo_banco = "1007"
                                codigo_banco_bancolombia = Utilidades.rellenar(codigo_banco, 9, "0", "I")
                                numero_cuenta = Utilidades.rellenar(documento_detalle.contacto.numero_cuenta, 17, "0", "I")
                                cuenta_banco_clase = "S37"
                                if documento_detalle.contacto.cuenta_banco_clase:
                                    if documento_detalle.contacto.cuenta_banco_clase_id == 2:
                                        cuenta_banco_clase = "D37"         
                                pago = int(documento_detalle.pago)
                                pago = Utilidades.rellenar(str(pago), 10, "0", "I")     
                                concepto = Utilidades.rellenar("NOMINA", 9, " ", "D")
                                numero = ""
                                if documento_detalle.documento_afectado:
                                    numero = documento_detalle.documento_afectado.numero
                                referencia = Utilidades.rellenar(numero, 13, " ", "D")
                                lineas.append(f"6{numero_identificacion}{nombre}{codigo_banco_bancolombia}{numero_cuenta}{cuenta_banco_clase}{pago}{concepto}{referencia}\n")
                            
                            contenido_archivo = "".join(lineas)
                            fecha_nombre_archivo = datetime.now().strftime('%Y%m%d%H%M%S')
                            nombre_archivo = f"planoBancolombiaSAP{fecha_nombre_archivo}.txt"
                            response = HttpResponse(contenido_archivo, content_type='text/csv')
                            response['Access-Control-Expose-Headers'] = 'Content-Disposition'
                            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'                    
                            return response   
                        if tipo_plano == 2:                           
                            numero_registros = 0
                            total_pagar = 0
                            for documento_detalle in documento_detalles:
                                numero_registros += 1
                                total_pagar += documento_detalle.pago
                            empresa = GenEmpresa.objects.get(pk=1)
                            lineas = []       
                            nit = Utilidades.rellenar(empresa.numero_identificacion, 15, "0", "I")                            
                            fecha = datetime.now().strftime('%Y%m%d')
                            numero_registros = Utilidades.rellenar(str(numero_registros), 6, "0", "I")
                            total_pagar = int(total_pagar)
                            total_pagar = Utilidades.rellenar(str(total_pagar), 32, "0", "I")
                            numero_cuenta = documento.cuenta_banco.numero_cuenta
                            cuenta_banco_clase = "S"
                            if documento.cuenta_banco.cuenta_banco_clase:
                                if documento.cuenta_banco.cuenta_banco_clase_id == 2:
                                    cuenta_banco_clase = "D"
                            lineas.append(f"1{nit}I               225NOMINA    {fecha}AA{fecha}{numero_registros}{total_pagar}00{numero_cuenta}{cuenta_banco_clase}\n")                        
                            
                            for documento_detalle in documento_detalles:
                                numero_identificacion = Utilidades.rellenar(documento_detalle.contacto.numero_identificacion, 15, " ", "D")
                                nombre = Utilidades.rellenar(documento_detalle.contacto.nombre_corto, 30, " ", "D")                            
                                codigo_banco = "1007"
                                if documento_detalle.contacto.banco:
                                    if documento_detalle.contacto.banco_id == 1:
                                        codigo_banco = "1007"
                                codigo_banco_bancolombia = Utilidades.rellenar(codigo_banco, 9, "0", "I")
                                numero_cuenta = Utilidades.rellenar(documento_detalle.contacto.numero_cuenta, 17, " ", "D")
                                cuenta_banco_clase = "S37"
                                if documento_detalle.contacto.cuenta_banco_clase:
                                    if documento_detalle.contacto.cuenta_banco_clase_id == 2:
                                        cuenta_banco_clase = "D37"         
                                pago = int(documento_detalle.pago)
                                pago = Utilidades.rellenar(str(pago), 15, "0", "I")                                     
                                numero = ""
                                if documento_detalle.documento_afectado:
                                    numero = documento_detalle.documento_afectado.numero
                                referencia = Utilidades.rellenar(numero, 21, " ", "D")
                                identificacion = "1" # 1-Cedula 2-Cedula extranjeria 3-NIT 4-Tarjeta identidad 5-Pasaporte
                                if documento_detalle.contacto.identificacion_id == 5:
                                    identificacion="2"
                                if documento_detalle.contacto.identificacion_id == 6:
                                    identificacion="3"                                    
                                if documento_detalle.contacto.identificacion_id == 2:
                                    identificacion="4"
                                if documento_detalle.contacto.identificacion_id == 7:
                                    identificacion="5"     
                                espacios = Utilidades.rellenar("", 137, " ", "D")                               
                                lineas.append(f"6{numero_identificacion}{nombre}{codigo_banco_bancolombia}{numero_cuenta}{cuenta_banco_clase}{pago}00{fecha}{referencia}{identificacion}00000{espacios}\n")                            
                            contenido_archivo = "".join(lineas)
                            fecha_nombre_archivo = datetime.now().strftime('%Y%m%d%H%M%S')
                            nombre_archivo = f"planoBancolombiaPAB{fecha_nombre_archivo}.txt"
                            response = HttpResponse(contenido_archivo, content_type='text/csv')
                            response['Access-Control-Expose-Headers'] = 'Content-Disposition'
                            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'                    
                            return response                          
                    else:
                        return Response({'mensaje':'El documento no tiene cuenta de banco asociada', 'codigo':15}, status=status.HTTP_400_BAD_REQUEST)              
                except GenDocumento.DoesNotExist:
                    return Response({'mensaje':'El documento no existe', 'codigo':15}, status=status.HTTP_400_BAD_REQUEST)
            else:
                return Response({'mensaje':'El tipo de plano solo puede ser 1 o 2', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)                
        else:
            return Response({'mensaje':'Faltan parametros', 'codigo':1}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["post"], url_path=r'exportar-zip',)
    def exportar_zip(self, request):        
        raw = request.data
        return Response({'mensaje': 'Proceso exitoso'}, status=status.HTTP_200_OK)

    @action(detail=False, methods=["post"], url_path=r'importar-zip-dian')
    def importar_compra_zip(self, request):
        raw = request.data
        zip_base64 = raw.get('archivo_base64', None)
        
        if zip_base64:
            # Definir namespaces del XML
            namespaces = {
                'cbc': 'urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2',
                'cac': 'urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2',
            }
            
            # Decodificar el Base64
            zip_data = base64.b64decode(zip_base64)
            zip_buffer = io.BytesIO(zip_data)
            
            with zipfile.ZipFile(zip_buffer, 'r') as zip_ref:
                xml_files = [f for f in zip_ref.namelist() if f.lower().endswith('.xml')]
                
                if xml_files:
                    
                    if len(xml_files) <= 1:
                    
                        xml_filename = xml_files[0]
                        with zip_ref.open(xml_filename) as xml_file:
                            xml_content = xml_file.read()
                            root = ET.fromstring(xml_content)
                            
                            prefix = ''
                            company_id = ''
                            document_uuid = ''
                            document_id = ''
                            issue_date = ''
                            due_date = ''
                            note = ''
                            
                            description = root.find('.//cac:Attachment/cac:ExternalReference/cbc:Description', namespaces=namespaces)
                            if description is not None and description.text:
                                cdata_content = description.text.strip()
                                if cdata_content.startswith('<![CDATA['):
                                    cdata_content = cdata_content[9:-3]
                                inner_root = ET.fromstring(cdata_content)
                                inner_namespaces = {
                                    'sts': 'dian:gov:co:facturaelectronica:Structures-2-1',
                                    'cbc': 'urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2',
                                    'cac': 'urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2',
                                }
                                
                                registration_name = root.findtext('.//cac:SenderParty/cac:PartyTaxScheme/cbc:RegistrationName', namespaces=namespaces)
                                prefix = inner_root.findtext('.//sts:Prefix', namespaces=inner_namespaces)
                                company_id = inner_root.findtext('.//cbc:CompanyID', namespaces=inner_namespaces)
                                document_uuid = inner_root.findtext('.//cbc:UUID', namespaces=inner_namespaces)
                                document_id = inner_root.findtext('.//cbc:ID', namespaces=inner_namespaces)
                                issue_date = inner_root.findtext('.//cbc:IssueDate', namespaces=inner_namespaces)
                                due_date = inner_root.findtext('.//cbc:PaymentDueDate', namespaces=inner_namespaces)
                                note = inner_root.findtext('.//cbc:Note', namespaces=inner_namespaces)
                                document_id_numerico = re.sub(r'\D', '', document_id) if document_id else ''

                            contacto = GenContacto.objects.filter(numero_identificacion = company_id).first()
                            contactoSerializador = GenContactoSerializador(contacto)
                            contacto = contactoSerializador.data

                            # Después de obtener los datos y antes de construir la respuesta
                            referencia_numero = document_id_numerico if document_id_numerico else ''
                            referencia_prefijo = prefix.strip() if prefix else ''

                            # Si hay prefijo y número, y el número comienza con el prefijo
                            if referencia_prefijo and referencia_numero and referencia_numero.startswith(referencia_prefijo):
                                # Quitar el prefijo del número
                                referencia_numero = referencia_numero[len(referencia_prefijo):]
                            datos = {
                                'contacto': contacto,
                                'referencia_numero': referencia_numero,
                                'referencia_cue': document_uuid.strip() if document_uuid else '',
                                'referencia_prefijo': referencia_prefijo,
                                'fecha': issue_date.strip() if issue_date else '',
                                'comentario': note if note else '',
                            }

                            return Response(datos, status=status.HTTP_200_OK)
                    else:
                        return Response({'mensaje': 'El ZIP contiene múltiples archivos XML', 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST)
                else:
                    return Response({'mensaje': 'El ZIP no contiene archivos XML', 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST)
        else :
            return Response({'mensaje': 'Faltan parametros', 'codigo': 1}, status=status.HTTP_400_BAD_REQUEST)

    def emitir(self, id):
        documento = GenDocumento.objects.get(pk=id)
        if documento.estado_aprobado == True:
            empresa = GenEmpresa.objects.get(pk=1)
            if empresa.rededoc_id:
                if documento.estado_electronico_enviado == False: 
                    if documento.numero: 
                        if documento.documento_tipo.electronico:
                            if documento.contacto:
                                # Factura y Documento soporte
                                if documento.documento_tipo_id in [1,2,3,11,12,24]:
                                    if documento.resolucion: 
                                        #Las facturas y documento soporte toman prefijo de la resolucion
                                        prefijo = documento.resolucion.prefijo
                                        documento_referencia_id = None
                                        forma_pago = 2
                                        if documento.plazo_pago_id == 1:
                                            forma_pago = 1                                            
                                        if documento.documento_tipo.documento_clase_id == 101:
                                            prefijo = "NC"
                                        if documento.documento_tipo.documento_clase_id == 102:
                                            prefijo = "ND" 
                                        if documento.documento_tipo.documento_clase_id == 304:
                                            prefijo = "DSAJ"
                                            if documento.documento_referencia:
                                                documento_referencia_id = f"DS{documento.documento_referencia}"
                                        correo = documento.contacto.correo_facturacion_electronica if documento.contacto.correo_facturacion_electronica else documento.contacto.correo
                                        responsable = 0
                                        if documento.contacto.regimen_id in (1, 5):
                                            responsable = 1
                                        datos_factura = {
                                            "cuentaId": empresa.rededoc_id,
                                            "documentoClaseId" : documento.documento_tipo_id,
                                            "documentoClienteId": documento.id,
                                            "documento" : {
                                                "ambiente" : 1,
                                                "prefijo" : prefijo,
                                                "numero" : documento.numero,
                                                "fecha" : str(documento.fecha),
                                                "hora" : str("12:00:00-05:00"),
                                                "fecha_vence" : str(documento.fecha_vence),
                                                "tipo_operacion" : str(10),
                                                "moneda" : "COP",
                                                "orden_compra": documento.orden_compra,
                                                "remision": documento.remision,
                                                "resolucion" : str(documento.resolucion.numero),
                                                "forma_pago" : forma_pago,
                                                "cantidad_detalles" :1,
                                                "subtotal" : "{0:.2f}".format(documento.subtotal),
                                                "subtotal_mas_impuestos" : "{0:.2f}".format(documento.subtotal + documento.impuesto),
                                                "base" : "{0:.2f}".format(documento.base_impuesto),
                                                "total_impuestos" : "{0:.2f}".format(documento.impuesto),
                                                "total_descuentos" : "{0:.2f}".format(documento.descuento),
                                                "total_cargos" : "{0:.2f}".format(0),
                                                "total_anticipos" : "{0:.2f}".format(0),
                                                "total_documento" : "{0:.2f}".format(documento.total_bruto),
                                                "total_iva" : "{0:.2f}".format(0),
                                                "total_consumo" : "{0:.2f}".format(0),
                                                "total_ica" : "{0:.2f}".format(0),
                                                "documento_referencia": documento_referencia_id,
                                                "adquiriente" : {
                                                    "identificacion" : documento.contacto.identificacion.codigo,
                                                    "numero_identificacion" : documento.contacto.numero_identificacion,
                                                    "digito_verificacion" : documento.contacto.digito_verificacion,
                                                    "razon_social" : documento.contacto.nombre_corto,
                                                    "pais" : "CO",
                                                    "ciudad" : documento.contacto.ciudad.nombre,
                                                    "departamento" : documento.contacto.ciudad.estado.nombre,
                                                    "direccion" : documento.contacto.direccion,
                                                    "obligaciones" : "0-99",
                                                    "nombres" : documento.contacto.nombre1,
                                                    "apellidos" : documento.contacto.apellido1,
                                                    "correo" : correo,
                                                    "telefono" : documento.contacto.telefono,
                                                    "tipo_organizacion_juridica" : documento.contacto.tipo_persona.id,
                                                    "regimen_tributario" : documento.contacto.regimen.codigo_interface,
                                                    "codigo_postal" : documento.contacto.ciudad.codigo_postal,
                                                    "responsable" : responsable,
                                                    "nacional" : 1
                                                },
                                            }
                                        }
                                        arr_medio_pago = []

                                        arr_medio_pago.append({
                                            "medio_pago": 31,
                                            "descripcion": "",
                                        })

                                        datos_factura['documento']['medios_pago'] = arr_medio_pago

                                        arr_item = []
                                        cantidad_items = 0                                        
                                        documentoDetalles = GenDocumentoDetalle.objects.filter(documento=id)
                                        for documentoDetalle in documentoDetalles:                                                                                    
                                            documento_impuestos = GenDocumentoImpuesto.objects.filter(
                                                documento_detalle_id=documentoDetalle.id,
                                                impuesto__impuesto_tipo_id=1
                                            ).values(
                                                tipo_impuesto=F('impuesto__impuesto_tipo_id'),
                                                porcentual=Cast(F('impuesto__porcentaje'), output_field=CharField())
                                            ).annotate(
                                                total=Cast(Coalesce(Sum('total'), 0, output_field=DecimalField()), output_field=CharField())
                                            )
                                            documento_impuestos = list(documento_impuestos)
                                            cantidad_items += 1
                                            arr_item.append({
                                                "consecutivo": cantidad_items,
                                                "codigo": documentoDetalle.item_id,
                                                "descripcion" : documentoDetalle.item.nombre,
                                                "marca" : "",
                                                "modelo" : "",
                                                "observacion" : "",
                                                "cantidad" : str(documentoDetalle.cantidad),
                                                "cantidad_empque": str(documentoDetalle.cantidad),
                                                "obserquio" : str(0),
                                                "precio_unitario" : "{0:.2f}".format(documentoDetalle.precio),
                                                "precio_referencia" : "{0:.2f}".format(documentoDetalle.precio),
                                                "valor" : "{0:.2f}".format(documentoDetalle.precio),
                                                "total_descuentos" : "{0:.2f}".format(documentoDetalle.descuento),
                                                "total_cargos" : "{0:.2f}".format(0),
                                                "total_impuestos" : "{0:.2f}".format(documentoDetalle.impuesto),
                                                "base" : "{0:.2f}".format(documentoDetalle.base_impuesto),
                                                "subtotal" : "{0:.2f}".format(documentoDetalle.subtotal),
                                                "impuestos" : documento_impuestos
                                            })
                                        datos_factura['documento']['cantidad_detalles'] = cantidad_items                                        
                                        documento_impuestos = GenDocumentoImpuesto.objects.filter(
                                            documento_detalle__documento_id=id,
                                            impuesto__impuesto_tipo_id=1
                                        ).values(
                                            tipo_impuesto=F('impuesto__impuesto_tipo_id'),
                                            porcentual=Cast(F('impuesto__porcentaje'), output_field=CharField())                                                                                                
                                        ).annotate(                                            
                                            total=Cast(Coalesce(Sum('total'), 0, output_field=DecimalField()), output_field=CharField()),
                                            base=Cast(Coalesce(Sum('base'), 0, output_field=DecimalField()), output_field=CharField())
                                        )
                                        documento_impuestos = list(documento_impuestos)
                                        datos_factura['documento']['detalles'] = arr_item
                                        datos_factura['doc_cantidad_item'] = cantidad_items
                                        datos_factura['documento']['impuestos'] = documento_impuestos
                                        wolframio = Wolframio()
                                        respuesta = wolframio.emitir(datos_factura)
                                        if respuesta['error'] == False: 
                                            documento.estado_electronico_enviado = True
                                            documento.electronico_id = respuesta['id']
                                            documento.save()                                        
                                        else:               
                                            return {'error':True, 'mensaje':respuesta['mensaje']}
                                        #return Response({'datos': datos_factura, 'documento_impuestos':documento_impuestos}, status=status.HTTP_200_OK)
                                    else:
                                        return {'error':True, 'mensaje':'La factura no cuenta con una resolución asociada'}                                    
                                    
                                # Compra
                                if documento.documento_tipo_id == 5:                                    
                                    if documento.referencia_prefijo and documento.referencia_numero and documento.referencia_cue:
                                        datos_compra = {
                                            "cuentaId": empresa.rededoc_id,
                                            "documentoClaseId" : documento.documento_tipo_id,
                                            "documentoClienteId": documento.id,
                                            "documento" : {                                                
                                                "prefijo" : documento.referencia_prefijo,
                                                "numero" : documento.referencia_numero,
                                                "cue" : documento.referencia_cue,
                                                "fecha" : str(documento.fecha),
                                                "hora" : str("12:00:00-05:00"),
                                                "fecha_vence" : str(documento.fecha_vence),                                                                                                
                                                "total_documento" : str(documento.total_bruto), 
                                                "proveedor" : {                                                        
                                                    "identificacion" : documento.contacto.identificacion.codigo,
                                                    "numero_identificacion" : documento.contacto.numero_identificacion,
                                                    "digito_verificacion" : documento.contacto.digito_verificacion,
                                                    "razon_social" : documento.contacto.nombre_corto,   
                                                    "correo" : documento.contacto.correo,                                                     
                                                }
                                            }
                                        }
                                                                                                    
                                        wolframio = Wolframio()
                                        respuesta = wolframio.emitir(datos_compra)
                                        if respuesta['error'] == False: 
                                            documento.estado_electronico_enviado = True
                                            documento.estado_electronico = True
                                            documento.electronico_id = respuesta['id']
                                            documento.save()                                        
                                        else:                                        
                                            return {'error':True, 'mensaje':respuesta['mensaje']}
                                        #return Response({'datos': datos_factura}, status=status.HTTP_200_OK)
                                    else:                                    
                                        return {'error':True, 'mensaje':'Los documentos de compra deben tener prefijo, numero y cue'}

                                # Nomina    
                                if documento.documento_tipo_id == 15:                                
                                    respuesta = DocumentoViewSet.validacion_nomina_emitir(documento)
                                    if respuesta['error'] == False:
                                        prefijo = "NE"                                
                                        datos = {
                                            "cuentaId": empresa.rededoc_id,
                                            "documentoClaseId" : documento.documento_tipo_id,
                                            "documentoClienteId": documento.id,
                                            "documento" : {
                                                "ambiente" : 1,
                                                "prefijo" : prefijo,
                                                "numero" : documento.numero,
                                                "fecha" : str(documento.fecha),
                                                "hora" : str("12:00:00-05:00"),
                                                "devengado":str(documento.devengado),
                                                "deduccion":str(documento.deduccion),
                                                "total_documento" : str(documento.total),                                                                                                                                                                                                    
                                                "fecha_desde":str(documento.fecha),
                                                "fecha_hasta":str(documento.fecha_hasta),                                                                                                                    
                                                "empleado" : {
                                                    "codigo": documento.contacto_id,
                                                    "identificacion" : documento.contacto.identificacion.codigo,
                                                    "numero_identificacion" : documento.contacto.numero_identificacion,
                                                    "digito_verificacion" : documento.contacto.digito_verificacion,
                                                    "razon_social" : documento.contacto.nombre_corto,
                                                    "pais" : "CO",
                                                    "ciudad" : documento.contacto.ciudad.nombre,
                                                    "ciudad_codigo_postal" : documento.contacto.ciudad.codigo_postal,
                                                    "departamento" : documento.contacto.ciudad.estado.codigo,
                                                    "direccion" : documento.contacto.direccion,                                                
                                                    "nombre1" : documento.contacto.nombre1,
                                                    "nombre2" : documento.contacto.nombre2,
                                                    "apellido1" : documento.contacto.apellido1,
                                                    "apellido2" : documento.contacto.apellido2,
                                                    "correo" : documento.contacto.correo,
                                                    "telefono" : documento.contacto.telefono,
                                                    "tipo_organizacion_juridica" : documento.contacto.tipo_persona.id,
                                                    "regimen_tributario" : documento.contacto.regimen.codigo_interface,                                                
                                                    "cuenta":documento.contacto.numero_cuenta,
                                                    "banco":documento.contacto.banco.nombre, 
                                                    "responsable" : 1,
                                                    "nacional" : 1
                                                },
                                                "contrato" : {
                                                    "fecha_desde":str(documento.contrato.fecha_desde),
                                                    "salario":str(documento.salario),
                                                    "salario_intergral":documento.contrato.salario_integral,
                                                    "tipo_cotizante":documento.contrato.tipo_cotizante.codigo,
                                                    "subtipo_cotizante":documento.contrato.subtipo_cotizante.codigo,
                                                    "contrato_tipo":documento.contrato.contrato_tipo_id,                                            
                                                    "ciudad_labora" : documento.contrato.ciudad_labora.nombre,
                                                    "ciudad_labora_codigo_postal" : documento.contrato.ciudad_labora.codigo,
                                                    "estado_labora_codigo_postal":documento.contrato.ciudad_labora.estado.codigo,
                                                },
                                            }
                                        }           
                                        
                                        detalle = {
                                            'valido': True,
                                            'Devengados': {
                                                'Basico': {
                                                    'DiasTrabajados': 0,
                                                    'SueldoTrabajado': 0
                                                },
                                                "Transporte": {
                                                    "AuxilioTransporte": 0,
                                                    "ViaticoManuAlojS": 0,
                                                    "ViaticoManuAlojNS": 0
                                                },
                                                "OtrosConceptos": [],
                                                "Bonificaciones": [],
                                                "Auxilios": [],
                                                "HorasExtras": {
                                                    "HED": [],
                                                    "HEN": [],
                                                    "HRN": [],
                                                    "HEDDF": [],
                                                    "HRDDF": [],
                                                    "HENDF": [],
                                                    "HRNDF": [],
                                                },
                                                "Compensaciones": [],
                                                "HuelgasLegales": [],
                                                "BonoEPCTVs": [],
                                                "Comisiones": [],
                                                "PagosTerceros": [],
                                                "Anticipos": [],
                                                "Vacaciones": {
                                                    "VacacionesComunes": [],
                                                    "VacacionesCompensadas": []
                                                },
                                                "Primas": {
                                                    'Cantidad':0,
                                                    'Pago':0,
                                                    'PagoNS':0
                                                },
                                                "Cesantias": {
                                                    "Pago": 0,
                                                    "Porcentaje": 0,
                                                    "PagoIntereses": 0
                                                },
                                                "Incapacidades": [],
                                                "Licencias": {
                                                    "LicenciaMP": [],
                                                    "LicenciaR": [],
                                                    "LicenciaNR": [],
                                                },
                                                'Dotacion': 0,
                                                'ApoyoSost': 0,
                                                'Teletrabajo': 0,
                                                'BonifRetiro': 0,
                                                'Indemnizacion': 0,
                                                'Reintegro': 0,
                                            },
                                            'Deducciones': {
                                                "Salud": {
                                                    "Porcentaje": 0,
                                                    "Deduccion": 0
                                                },
                                                "FondoPension": {
                                                    "Porcentaje": 0,
                                                    "Deduccion": 0
                                                },
                                                "FondoSP": {
                                                    "Porcentaje": 0,
                                                    "DeduccionSP": 0,
                                                    "PorcentajeSub": 0,
                                                    "DeduccionSub": 0
                                                },
                                                "Sindicatos": [],
                                                "Sanciones": [],
                                                "Libranzas": [],
                                                "PagosTerceros": [],
                                                "Anticipos": [],
                                                "OtrasDeducciones": [],
                                                'PensionVoluntaria': 0,
                                                'RetencionFuente': 0,
                                                'AFC': 0,
                                                'Cooperativa': 0,
                                                'EmbargoFiscal': 0,
                                                'PlanComplementarios': 0,
                                                'Educacion': 0,
                                                'Reintegro': 0,
                                                'Deuda': 0,
                                            }
                                        }
                                        datos['documento']['detalle'] = detalle
                                        documento_detalles = GenDocumentoDetalle.objects.filter(
                                            documento__documento_referencia_id=id
                                        ).values(
                                            'id', 'operacion', 'porcentaje', 'dias', 'cantidad', 'pago','concepto_id', 'concepto__nombre', 'concepto__concepto_tipo_id'
                                        )
                                        #print(documento_detalles.query)
                                        for documento_detalle in documento_detalles:                                        
                                            fecha = documento.fecha
                                            fecha_desde_extra = fecha.strftime('%Y-%m-%d') + "T11:00:00"
                                            fecha_hasta_extra = fecha.strftime('%Y-%m-%d') + "T12:00:00"                                        
                                            if documento_detalle['operacion'] == 1:
                                                if documento_detalle['concepto__concepto_tipo_id'] == 1:
                                                    dias_basico = documento_detalle['dias'] if documento_detalle['dias'] != 0 else 1
                                                    detalle['Devengados']['Basico']['DiasTrabajados'] += dias_basico
                                                    detalle['Devengados']['Basico']['SueldoTrabajado'] += documento_detalle['pago']                                                

                                                if documento_detalle['concepto__concepto_tipo_id'] == 7:
                                                    detalle['Devengados']['Transporte']['AuxilioTransporte'] += documento_detalle['pago']

                                                if documento_detalle['concepto__concepto_tipo_id'] == 3:                                                                                                                                            
                                                    detalle['Devengados']['OtrosConceptos'].append({
                                                            'DescripcionConcepto': documento_detalle['concepto__nombre'],
                                                            'ConceptoS': documento_detalle['pago'],
                                                            'ConceptoNS': 0
                                                            })
                                                    
                                                if documento_detalle['concepto__concepto_tipo_id'] == 4:                                                                                                                                            
                                                    detalle['Devengados']['OtrosConceptos'].append({
                                                            'DescripcionConcepto': documento_detalle['concepto__nombre'],
                                                            'ConceptoS': 0,
                                                            'ConceptoNS': documento_detalle['pago']
                                                            })
                                                
                                                if documento_detalle['concepto__concepto_tipo_id'] == 2:
                                                    # No se pueden cambiar los porcentajes
                                                    if documento_detalle['concepto_id'] in [5,3]:
                                                        detalle['Devengados']['HorasExtras']['HED'].append({
                                                            "Id":0,
                                                            "HoraInicio":fecha_desde_extra,
                                                            "HoraFin":fecha_hasta_extra,
                                                            "Cantidad":documento_detalle['cantidad'],
                                                            "Porcentaje":25,
                                                            "Pago":documento_detalle['pago']
                                                            })                                                    
                                                    if documento_detalle['concepto_id'] in [6,4]:
                                                        detalle['Devengados']['HorasExtras']['HEN'].append({
                                                            "Id":0,
                                                            "HoraInicio":fecha_desde_extra,
                                                            "HoraFin":fecha_hasta_extra,
                                                            "Cantidad":documento_detalle['cantidad'],
                                                            "Porcentaje":75,
                                                            "Pago":documento_detalle['pago']
                                                            }) 
                                                    if documento_detalle['concepto_id'] in [9,2]:
                                                        detalle['Devengados']['HorasExtras']['HRN'].append({
                                                            "Id":0,
                                                            "HoraInicio":fecha_desde_extra,
                                                            "HoraFin":fecha_hasta_extra,
                                                            "Cantidad":documento_detalle['cantidad'],
                                                            "Porcentaje":35,
                                                            "Pago":documento_detalle['pago']
                                                            }) 
                                                    if documento_detalle['concepto_id'] in [7]:
                                                        detalle['Devengados']['HorasExtras']['HEDDF'].append({
                                                            "Id":0,
                                                            "HoraInicio":fecha_desde_extra,
                                                            "HoraFin":fecha_hasta_extra,
                                                            "Cantidad":documento_detalle['cantidad'],
                                                            "Porcentaje":100,
                                                            "Pago":documento_detalle['pago']
                                                            }) 
                                                    if documento_detalle['concepto_id'] in [10]:
                                                        detalle['Devengados']['HorasExtras']['HRDDF'].append({
                                                            "Id":0,
                                                            "HoraInicio":fecha_desde_extra,
                                                            "HoraFin":fecha_hasta_extra,
                                                            "Cantidad":documento_detalle['cantidad'],
                                                            "Porcentaje":75,
                                                            "Pago":documento_detalle['pago']
                                                            }) 
                                                    if documento_detalle['concepto_id'] in [8]:
                                                        detalle['Devengados']['HorasExtras']['HENDF'].append({
                                                            "Id":0,
                                                            "HoraInicio":fecha_desde_extra,
                                                            "HoraFin":fecha_hasta_extra,
                                                            "Cantidad":documento_detalle['cantidad'],
                                                            "Porcentaje":150,
                                                            "Pago":documento_detalle['pago']
                                                            }) 
                                                    if documento_detalle['concepto_id'] in [11]:
                                                        detalle['Devengados']['HorasExtras']['HRNDF'].append({
                                                            "Id":0,
                                                            "HoraInicio":fecha_desde_extra,
                                                            "HoraFin":fecha_hasta_extra,
                                                            "Cantidad":documento_detalle['cantidad'],
                                                            "Porcentaje":110,
                                                            "Pago":documento_detalle['pago']
                                                            })                                                                                                                                                                                                                                                                                                 

                                                if documento_detalle['concepto__concepto_tipo_id'] == 18:
                                                    detalle['Devengados']['Vacaciones']['VacacionesComunes'].append({
                                                            'FechaInicio': None,
                                                            'FechaFin': None,
                                                            'Cantidad': documento_detalle['dias'],
                                                            'Pago': documento_detalle['pago']
                                                            })

                                                if documento_detalle['concepto__concepto_tipo_id'] == 19:
                                                    detalle['Devengados']['Vacaciones']['VacacionesCompensadas'].append({
                                                            'Cantidad': 0,
                                                            'Pago': documento_detalle['pago']
                                                    })

                                                if documento_detalle['concepto__concepto_tipo_id'] in [20,21]:
                                                    detalle['Devengados']['Primas']['PagoNS'] += documento_detalle['pago']

                                                if documento_detalle['concepto__concepto_tipo_id'] in [22,23]:
                                                    detalle['Devengados']['Cesantias']['Pago'] += documento_detalle['pago']

                                                if documento_detalle['concepto__concepto_tipo_id'] in [24,25]:
                                                    detalle['Devengados']['Cesantias']['PagoIntereses'] += documento_detalle['pago']

                                                if documento_detalle['concepto__concepto_tipo_id'] == 12:
                                                    detalle['Devengados']['Incapacidades'].append({
                                                        'FechaInicio': None,
                                                        'FechaFin': None,
                                                        'Cantidad':documento_detalle['dias'],
                                                        'Tipo':1,
                                                        'Pago':documento_detalle['pago']
                                                    })
                                                
                                                if documento_detalle['concepto__concepto_tipo_id'] == 13:
                                                    detalle['Devengados']['Incapacidades'].append({
                                                        'FechaInicio': None,
                                                        'FechaFin': None,
                                                        'Cantidad':documento_detalle['dias'],
                                                        'Tipo':3,
                                                        'Pago':documento_detalle['pago']
                                                    })  

                                                if documento_detalle['concepto__concepto_tipo_id'] == 14:
                                                    detalle['Devengados']['Licencias']['LicenciaMP'].append({
                                                        "FechaInicio":None,
                                                        "FechaFin":None,
                                                        "Cantidad":0,
                                                        "Pago":documento_detalle['pago']
                                                    })

                                                if documento_detalle['concepto__concepto_tipo_id'] in [15,16]:
                                                    detalle['Devengados']['Licencias']['LicenciaR'].append({
                                                        "FechaInicio":None,
                                                        "FechaFin":None,
                                                        "Cantidad":0,
                                                        "Pago":documento_detalle['pago']
                                                    })

                                                if documento_detalle['concepto__concepto_tipo_id'] in [17]:
                                                    detalle['Devengados']['Licencias']['LicenciaNR'].append({
                                                        "FechaInicio":None,
                                                        "FechaFin":None,
                                                        "Cantidad":0,
                                                        "Pago":documento_detalle['pago']
                                                    })                                                
                                                    
                                            if documento_detalle['operacion'] == -1:        
                                                if documento_detalle['concepto__concepto_tipo_id'] == 5:
                                                    detalle['Deducciones']['Salud']['Porcentaje'] = documento_detalle['porcentaje']
                                                    detalle['Deducciones']['Salud']['Deduccion'] += documento_detalle['pago']

                                                if documento_detalle['concepto__concepto_tipo_id'] == 6:
                                                    detalle['Deducciones']['FondoPension']['Porcentaje'] = documento_detalle['porcentaje']
                                                    detalle['Deducciones']['FondoPension']['Deduccion'] += documento_detalle['pago']

                                                if documento_detalle['concepto__concepto_tipo_id'] == 8:                                                                                                                                            
                                                    detalle['Deducciones']['Deuda'] += documento_detalle['pago']

                                                if documento_detalle['concepto__concepto_tipo_id'] == 9:
                                                    detalle['Deducciones']['FondoSP']['Porcentaje'] = documento_detalle['porcentaje']
                                                    detalle['Deducciones']['FondoSP']['DeduccionSP'] += documento_detalle['pago']

                                                if documento_detalle['concepto__concepto_tipo_id'] == 10:
                                                    detalle['Deducciones']['RetencionFuente'] += documento_detalle['pago']

                                                if documento_detalle['concepto__concepto_tipo_id'] == 11:                                                                                                                                            
                                                    detalle['Deducciones']['OtrasDeducciones'].append(documento_detalle['pago'])

                                        datos = transformar_decimal(datos)
                                        wolframio = Wolframio()
                                        respuesta = wolframio.emitir(datos)
                                        if respuesta['error'] == False: 
                                            documento.estado_electronico_enviado = True
                                            documento.electronico_id = respuesta['id']
                                            documento.save()                                        
                                        else:                                        
                                            return {'error':True, 'mensaje':respuesta['mensaje']}
                                    else:                                   
                                        return {'error':True, 'mensaje':respuesta['mensaje']}
                                return {'error':False}
                            else:                                
                                return {'error':True, 'mensaje':'El documento no tiene contacto'}
                        else:                                
                            return {'error':True, 'mensaje':'El tipo de documento no esta habilitado para enviar a la DIAN'}
                    else:                        
                        return {'error':True, 'mensaje':'La factura no cuenta con un número'}
                else:                                            
                    return {'error':True, 'mensaje':'El documento ya fue enviado'}
            else:                  
                return {'error':True, 'mensaje':'La empresa no se encuentra activada para emitir documentos electronicos'}
        else:
            return {'error':True, 'mensaje':'El documento no se puede emitir ya que no está aprobado'}
            
    @staticmethod
    def notificar(documento_id):
        try:
            documento = GenDocumento.objects.get(id=documento_id)
            if documento.estado_electronico_notificado == False:
                if documento.documento_tipo.documento_clase_id in [100, 101, 102]:                                        
                    formatoFactura = FormatoFactura()
                    pdf = formatoFactura.generar_pdf(documento_id)   
                    pdf_base64 = "data:application/pdf;base64," + base64.b64encode(pdf).decode('utf-8')            
                    wolframio = Wolframio()
                    respuesta = wolframio.notificar(documento.electronico_id, pdf_base64)
                    if respuesta['error'] == False: 
                        documento.estado_electronico_notificado = True                
                        documento.save()     
                    return respuesta                        
                else:
                    documento.estado_electronico_notificado = True                
                    documento.save()
                    return {'error':False}
            else:  
                return {'error':True, 'mensaje':'El documento ya fue notificado con anterioridad pruebe re-notificando'}
        except GenDocumento.DoesNotExist:
            return {'error':True, 'mensaje':'El documento no existe'}        
    
    @staticmethod
    def validacion_aprobar(documento: GenDocumento, documento_detalles: GenDocumentoDetalle, consecutivo):
        fecha = date.today()
        if documento.estado_aprobado == False:
            if documento.total >= 0:
                # Factura o documento soporte
                if documento.documento_tipo.documento_clase_id in (100,303):
                    # Si tiene detalles o es nomina electronica
                    if documento_detalles or documento.documento_tipo_id == 15:
                        if documento.resolucion:
                            if isinstance(documento.resolucion.fecha_hasta, date):
                                if documento.resolucion.fecha_hasta <= fecha:
                                    return {'error':True, 'mensaje':'La fecha de la resolucion esta vencida', 'codigo':1}
                            if consecutivo < documento.resolucion.consecutivo_desde or consecutivo > documento.resolucion.consecutivo_hasta:
                                return {'error':True, 'mensaje':f'El consecutivo {consecutivo} no corresponde con la resolucion desde {documento.resolucion.consecutivo_desde} hasta {documento.resolucion.consecutivo_hasta}', 'codigo':1}
                    else:
                        return {'error':True, 'mensaje':'El documento no tiene detalles', 'codigo':1} 
                
                # Pago
                if documento.documento_tipo.documento_clase_id == 200:
                    resultado = (
                        GenDocumentoDetalle.objects
                        .filter(documento_id=documento.id)
                        .exclude(documento_afectado_id__isnull=True)
                        .values('documento_afectado_id')
                        .annotate(
                            total_precio=Sum('precio'),
                            pendiente=F('documento_afectado__pendiente')))                        
                    for registro in resultado:
                        if registro['pendiente'] < registro['total_precio']:
                            return {'error':True, 'mensaje':f"El documento {registro['documento_afectado_id']} tiene saldo pendiente {registro['pendiente']} y se va afectar {registro['total_precio']}", 'codigo':1}
                        
                # Nomina
                if documento.documento_tipo.documento_clase_id == 702:
                    if documento.contacto.nombre1 and documento.contacto.apellido1:
                        pass
                    else:                        
                        return {'error':True, 'mensaje':'El contacto no tiene nombre1 o apellido1', 'codigo':1}
                    
                # Validar inventario
                for documento_detalle in documento_detalles:
                    if documento_detalle.operacion_inventario == -1:
                        if documento_detalle.almacen and documento_detalle.item:
                            existencia = InvExistencia.objects.filter(item_id=documento_detalle.item_id, almacen_id=documento_detalle.almacen_id).first()                            
                            if not existencia:
                                existencia = InvExistencia(item=documento_detalle.item, almacen=documento_detalle.almacen)
                                existencia.save()
                            disponible = existencia.disponible + documento_detalle.cantidad_operada
                            if disponible < 0 and documento_detalle.item.negativo == False:
                                return {'error':True, 'mensaje':f"En el detalle {documento_detalle.id} el item {documento_detalle.item_id} supera la cantidad disponible {existencia.disponible}", 'codigo':1}                                                            
                        else:
                            return {'error':True, 'mensaje':'El detalle afecta inventario no tiene almacen o item', 'codigo':1}                                   
                
                # Notas credito
                if documento.documento_referencia:
                    if documento.documento_referencia.pendiente < documento.total:
                        return {'error':True, 'mensaje':f"El documento referencia tiene saldo pendiente {documento.documento_referencia.pendiente} y se va afectar {documento.total}", 'codigo':1}

                return {'error':False}         
            else:
                return {'error':True, 'mensaje':'El total del documento no puede ser menor a cero', 'codigo':1}            
        else:
            return {'error':True, 'mensaje':'El documento ya esta aprobado', 'codigo':1}   

    @staticmethod
    def validacion_desaprobar(documento: GenDocumento):
        fecha = date.today()                                 
        documento_detalles = GenDocumentoDetalle.objects.filter(documento_id=documento.id)
        for documento_detalle in documento_detalles:
            if documento_detalle.operacion_inventario == 1:
                existencia = InvExistencia.objects.filter(item_id=documento_detalle.item_id, almacen_id=documento_detalle.almacen_id).first()
                disponible = existencia.disponible - documento_detalle.cantidad_operada
                if disponible < 0:
                    return {'error':True, 'mensaje':f"En el detalle {documento_detalle.id} el item {documento_detalle.item_id} dejaria el item en existencia negativa {disponible}", 'codigo':1}
        return {'error':False}  

    @staticmethod
    def validacion_anular(documento: GenDocumento):
        fecha = date.today()                                 
        documento_detalles = GenDocumentoDetalle.objects.filter(documento_id=documento.id)
        for documento_detalle in documento_detalles:
            if documento_detalle.operacion_inventario == 1:
                existencia = InvExistencia.objects.filter(item_id=documento_detalle.item_id, almacen_id=documento_detalle.almacen_id).first()
                disponible = existencia.disponible - documento_detalle.cantidad_operada
                if disponible < 0:
                    return {'error':True, 'mensaje':f"En el detalle {documento_detalle.id} el item {documento_detalle.item_id} dejaria el item en existencia negativa {disponible}", 'codigo':1}
        return {'error':False}  

    @staticmethod
    def validacion_nomina_emitir(documento: GenDocumento):
        if documento.contacto:
            if documento.contacto.banco:
                return {'error':False}
            else:
                return {'error':True, 'mensaje':f"El contacto no tiene un banco asignado", 'codigo':1}                
        else:
            return {'error':True, 'mensaje':f"El documento no tiene contacto", 'codigo':1}            
        
                    
        
             

